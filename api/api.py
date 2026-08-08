#!/usr/bin/env python3
"""
Coaches Station API — a thin FastAPI layer over nhl.db.

Design intent: every endpoint is a small, self-contained function reading
from the existing schema (plays/games/on_ice/players — see schema.sql).
Adding a new endpoint for a new coach request should mean writing one new
function + one new route, not touching anything else. Nothing here bakes
in a fixed set of "features" — it's just queryable access to the full
play-by-play data already sitting in nhl.db.

Run locally:
    pip install fastapi uvicorn
    uvicorn api:app --reload --port 8000

Then Coaches Station's JS can hit e.g. http://localhost:8000/api/game/20240/lines
"""

import json
import sqlite3
import asyncio
import datetime as dt
from pathlib import Path
from collections import defaultdict
from itertools import combinations
import httpx

from fastapi import FastAPI, HTTPException, Query, Depends, Request, Response
from fastapi.middleware.cors import CORSMiddleware

import os

# Locally (on your Mac), this just looks for nhl.db sitting next to api.py,
# same as always. On Fly.io, we point this at the persistent volume mount
# instead (via the DB_PATH environment variable set in fly.toml) so the
# database survives restarts and deploys.
DB_PATH = Path(os.environ.get("DB_PATH", Path(__file__).parent / "nhl.db"))

# plays.team stores full names ("Edmonton Oilers") as parsed straight from
# the PSF CSV, while games.homeTeamAbbrev/awayTeamAbbrev use abbreviations.
# Every endpoint below takes abbreviations (matches how the rest of
# Coaches Station addresses teams), so this bridges the two.
TEAM_NAMES = {
    "ANA": "Anaheim Ducks", "BOS": "Boston Bruins", "BUF": "Buffalo Sabres",
    "CAR": "Carolina Hurricanes", "CBJ": "Columbus Blue Jackets", "CGY": "Calgary Flames",
    "CHI": "Chicago Blackhawks", "COL": "Colorado Avalanche", "DAL": "Dallas Stars",
    "DET": "Detroit Red Wings", "EDM": "Edmonton Oilers", "FLA": "Florida Panthers",
    "LAK": "Los Angeles Kings", "MIN": "Minnesota Wild", "MTL": "Montreal Canadiens",
    "NJD": "New Jersey Devils", "NSH": "Nashville Predators", "NYI": "New York Islanders",
    "NYR": "New York Rangers", "OTT": "Ottawa Senators", "PHI": "Philadelphia Flyers",
    "PIT": "Pittsburgh Penguins", "SEA": "Seattle Kraken", "SJS": "San Jose Sharks",
    "STL": "St. Louis Blues", "TBL": "Tampa Bay Lightning", "TOR": "Toronto Maple Leafs",
    "UTA": "Utah Mammoth", "VAN": "Vancouver Canucks", "VGK": "Vegas Golden Knights",
    "WPG": "Winnipeg Jets", "WSH": "Washington Capitals",
}


def full_name(abbr: str) -> str:
    name = TEAM_NAMES.get(abbr.upper())
    if not name:
        raise HTTPException(400, f"Unknown team abbreviation '{abbr}'")
    return name

# Regular-season/playoff cutoff — matches the same constant used in
# Coaches Station's own JS, so "Regular Season"/"Playoffs" mean the same
# thing whether a date is being filtered client-side or here.
RS_END = "2026-04-16"


app = FastAPI(title="Coaches Station API", version="0.1")


@app.on_event("startup")
def ensure_schema():
    """Applies schema.sql if the tables don't exist yet — matters on a
    brand new deployment where the persistent volume starts out empty.
    Safe to run every time (schema.sql uses CREATE TABLE IF NOT EXISTS
    throughout), so this doesn't touch anything on repeat startups.

    If the existing file is corrupt or incomplete (e.g. an interrupted
    upload/transfer got cut off partway through), this moves it aside
    rather than crash-looping on a database the app can never open —
    that's exactly what happened once already: a dropped SFTP transfer
    left a partial file, and the app died on every single restart trying
    to read it, with no way to SSH in and fix it since it never stayed
    up long enough. This makes that failure mode impossible going
    forward, regardless of what caused the corruption."""
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    schema_path = Path(__file__).parent / "schema.sql"

    def apply_schema():
        conn = sqlite3.connect(DB_PATH)
        with open(schema_path) as f:
            conn.executescript(f.read())
        conn.commit()
        conn.close()

    try:
        apply_schema()
    except sqlite3.DatabaseError as e:
        if DB_PATH.exists():
            backup_path = DB_PATH.with_name(DB_PATH.name + ".corrupt")
            DB_PATH.rename(backup_path)
            print(f"[startup] {DB_PATH} was corrupt/unreadable ({e}) — "
                  f"moved aside to {backup_path}, starting fresh.")
        apply_schema()

    # Small migrations for columns added after a table already existed on
    # the live database — CREATE TABLE IF NOT EXISTS alone won't add a new
    # column to a table that's already there from before this feature.
    conn = sqlite3.connect(DB_PATH)
    existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if "isAdmin" not in existing_cols:
        conn.execute("ALTER TABLE users ADD COLUMN isAdmin INTEGER NOT NULL DEFAULT 0")
        conn.commit()
        print("[startup] migrated: added isAdmin column to users table")
    conn.close()

# Wide open for local dev — tighten this (specific origin) before hosting
# this anywhere public.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


import bcrypt
from fastapi.security import HTTPBasic, HTTPBasicCredentials

security = HTTPBasic()

class CurrentUser:
    def __init__(self, username: str, is_admin: bool):
        self.username = username
        self.is_admin = is_admin

def require_login(credentials: HTTPBasicCredentials = Depends(security)) -> CurrentUser:
    """Every protected endpoint depends on this. Checks the submitted
    username/password against the bcrypt hash stored in the users table —
    the plaintext password itself is never stored anywhere, only what's
    needed to verify a login attempt matches it."""
    conn = get_db()
    row = conn.execute(
        "SELECT passwordHash, isAdmin FROM users WHERE username = ?", (credentials.username,)
    ).fetchone()
    conn.close()
    valid = row and bcrypt.checkpw(credentials.password.encode("utf-8"), row["passwordHash"].encode("utf-8"))
    if not valid:
        raise HTTPException(401, detail="Invalid username or password", headers={"WWW-Authenticate": "Basic"})
    return CurrentUser(credentials.username, bool(row["isAdmin"]))

def require_admin(user: CurrentUser = Depends(require_login)) -> CurrentUser:
    """Same as require_login, but additionally requires the account to be
    an admin — used for user management and data-upload endpoints."""
    if not user.is_admin:
        raise HTTPException(403, detail=f"'{user.username}' is not an admin — this action is restricted to admin accounts.")
    return user


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def game_or_404(conn, game_id: int):
    row = conn.execute("SELECT * FROM games WHERE gameReferenceId = ?", (game_id,)).fetchone()
    if not row:
        raise HTTPException(404, f"Game {game_id} not found in nhl.db")
    return dict(row)


# ── Basic game/team lookups ─────────────────────────────────────────────

from pydantic import BaseModel

class NewUserRequest(BaseModel):
    username: str
    password: str
    isAdmin: bool = False

class ResetPasswordRequest(BaseModel):
    password: str


@app.get("/api/admin/users")
def list_users(admin: CurrentUser = Depends(require_admin)):
    """Every account's username and admin status — never returns password
    hashes at all, there's no legitimate reason for this list to include them."""
    conn = get_db()
    rows = conn.execute("SELECT username, isAdmin, createdAt FROM users ORDER BY username").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/admin/users")
def add_user(req: NewUserRequest, admin: CurrentUser = Depends(require_admin)):
    """Adds a new account, or fully replaces an existing one with this
    username (INSERT OR REPLACE) — used both for adding new staff and for
    resetting someone's password by re-adding them."""
    if not req.username or not req.password:
        raise HTTPException(400, "Both username and password are required")
    hashed = bcrypt.hashpw(req.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    conn = get_db()
    conn.execute(
        "INSERT OR REPLACE INTO users (username, passwordHash, isAdmin) VALUES (?, ?, ?)",
        (req.username, hashed, 1 if req.isAdmin else 0),
    )
    conn.commit()
    conn.close()
    return {"username": req.username, "isAdmin": req.isAdmin, "status": "created"}


@app.put("/api/admin/users/{username}/password")
def reset_password(username: str, req: ResetPasswordRequest, admin: CurrentUser = Depends(require_admin)):
    """Resets an existing user's password without touching their admin
    status or creating a duplicate account."""
    conn = get_db()
    existing = conn.execute("SELECT username FROM users WHERE username = ?", (username,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(404, f"No account named '{username}'")
    hashed = bcrypt.hashpw(req.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    conn.execute("UPDATE users SET passwordHash = ? WHERE username = ?", (hashed, username))
    conn.commit()
    conn.close()
    return {"username": username, "status": "password reset"}


@app.delete("/api/admin/users/{username}")
def remove_user(username: str, admin: CurrentUser = Depends(require_admin)):
    """Removes an account. An admin can't delete their own account this
    way — a small guard against accidentally locking every admin out."""
    if username == admin.username:
        raise HTTPException(400, "Can't delete your own account while logged in as it")
    conn = get_db()
    existing = conn.execute("SELECT username FROM users WHERE username = ?", (username,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(404, f"No account named '{username}'")
    conn.execute("DELETE FROM users WHERE username = ?", (username,))
    conn.commit()
    conn.close()
    return {"username": username, "status": "removed"}


@app.get("/api/login")
def check_login(user: CurrentUser = Depends(require_login)):
    """Dedicated endpoint for the frontend's login screen to verify a
    username/password before unlocking the rest of the app. Also bundles
    in whatever Pre-Scout Prep data is already cached (see _prep_cache)
    so the frontend gets it in this same response instead of needing a
    separate round trip right after — this is a plain in-memory dict
    read, not a live computation, so it doesn't add any real time to
    this endpoint even when the cache holds a big payload."""
    return {"username": user.username, "isAdmin": user.is_admin, "status": "ok", "prep": _prep_cache}


@app.get("/api/seasons")
def list_seasons(user: CurrentUser = Depends(require_login)):
    """Every distinct season present in the database (e.g. '20252026'),
    most recent first — powers the season selector dropdown."""
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT season FROM games WHERE season IS NOT NULL ORDER BY season DESC").fetchall()
    conn.close()
    return [r["season"] for r in rows]


# ── NHL API passthrough ───────────────────────────────────────────────────
# The dashboard needs a few things straight from the NHL's own public API
# (roster jersey numbers/positions, schedule-to-game-ID mapping, boxscore
# goalie data) that aren't in our own database. Calling api-web.nhle.com
# directly from the browser was tried and confirmed broken: the NHL API
# doesn't send an Access-Control-Allow-Origin header, so every browser-side
# fetch to it gets blocked by CORS (verified against the live dashboard on
# 2026-08-01 — this wasn't just our new lineup feature, it also broke the
# pre-existing "data freshness" badge check the same way). A request made
# by THIS server, however, is a normal server-to-server HTTP call and isn't
# subject to browser CORS at all — so the dashboard routes NHL API calls
# through here instead of hitting api-web.nhle.com directly.
NHL_API_BASE = "https://api-web.nhle.com/v1"

# In-memory caches for NHL API responses. None of this data has any reason
# to change within a server process's lifetime for the cases we hit it:
# a past game's boxscore is immutable, a team's schedule-to-game-id mapping
# for games already played doesn't change, and a roster snapshot going
# briefly stale (rather than re-fetched on every single request) is a
# perfectly fine tradeoff for speed. Before this, opening one game recap
# triggered up to 6 fresh NHL API network round trips (roster x2 teams,
# schedule x2 teams, boxscore x2 teams) with zero reuse across requests —
# a real, measurable chunk of the "lineups load slowly" complaint. Boxscore
# is keyed by the NHL's own game id (not team+date) specifically so that
# the primary and opponent team's requests for the SAME game share one
# cached fetch instead of each independently re-fetching the identical JSON.
_nhl_roster_cache: dict = {}
_nhl_schedule_cache: dict = {}
_nhl_boxscore_cache: dict = {}

# Same CORS reasoning as NHL_API_BASE above, but for the NHL's OTHER
# public stats host — a completely different domain from api-web.nhle.com,
# used for stats that api-web.nhle.com's club-stats endpoint doesn't carry
# (e.g. real PP points — club-stats only has powerPlayGoals, no
# powerPlayAssists/powerPlayPoints at all; confirmed by inspecting a live
# response). Needs its own proxy route, not just reuse of nhl_proxy above,
# because this one has to forward the query string (cayenneExp filtering,
# sort, pagination) — nhl_proxy never needed that for any of its existing
# path-only callers.
NHL_STATS_API_BASE = "https://api.nhle.com"

@app.get("/api/nhl-stats-proxy/{path:path}")
async def nhl_stats_proxy(path: str, request: Request, user: CurrentUser = Depends(require_login)):
    """Passthrough to https://api.nhle.com/{path}, forwarding the query
    string as-is. E.g. path='stats/rest/en/skater/summary' with
    ?cayenneExp=gameTypeId=2 and seasonId=20252026 and teamAbbrevs="VAN".
    Same login requirement as nhl_proxy, same reasoning (keeps this from
    being an open proxy; the data itself is public)."""
    url = f"{NHL_STATS_API_BASE}/{path}"
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.get(url, params=list(request.query_params.multi_items()))
    except httpx.RequestError as e:
        raise HTTPException(502, f"Couldn't reach the NHL stats API ({path}): {e}")
    if r.status_code != 200:
        raise HTTPException(r.status_code, f"NHL stats API returned {r.status_code} for {path}")
    return r.json()


@app.get("/api/nhl-proxy/{path:path}")
async def nhl_proxy(path: str, user: CurrentUser = Depends(require_login)):
    """Passthrough to https://api-web.nhle.com/v1/{path}. E.g. path=
    'roster/EDM/current', 'club-schedule-season/EDM/20252026', or
    'gamecenter/2025020006/boxscore'. Requires login same as everything
    else here, purely to keep this from being an open proxy anyone could
    hit — the NHL data itself is public either way."""
    url = f"{NHL_API_BASE}/{path}"
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.get(url)
    except httpx.RequestError as e:
        raise HTTPException(502, f"Couldn't reach the NHL API ({path}): {e}")
    if r.status_code != 200:
        raise HTTPException(r.status_code, f"NHL API returned {r.status_code} for {path}")
    return r.json()


@app.get("/api/teams")
def list_teams(season: str | None = Query(None), user: CurrentUser = Depends(require_login)):
    """Every distinct team abbreviation with at least one game in the
    database — used for team dropdowns/segment lookups that shouldn't
    depend on an Excel upload having happened first."""
    conn = get_db()
    q = "SELECT DISTINCT homeTeamAbbrev AS abbr FROM games"
    params = []
    if season:
        q += " WHERE season = ?"
        params.append(season)
    q += " UNION SELECT DISTINCT awayTeamAbbrev FROM games"
    if season:
        q += " WHERE season = ?"
        params.append(season)
    q += " ORDER BY abbr"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [r["abbr"] for r in rows]


@app.get("/api/games")
def list_games(user: CurrentUser = Depends(require_login)):
    """Every game currently loaded into nhl.db."""
    conn = get_db()
    rows = conn.execute(
        "SELECT gameReferenceId, gameDate, awayTeamAbbrev, homeTeamAbbrev, matchup, numPlays "
        "FROM games ORDER BY gameDate"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/team/{abbr}/games")
def team_games(abbr: str, user: CurrentUser = Depends(require_login)):
    """A team's full game log (as home or away)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT gameReferenceId, gameDate, awayTeamAbbrev, homeTeamAbbrev, matchup "
        "FROM games WHERE homeTeamAbbrev = ? OR awayTeamAbbrev = ? ORDER BY gameDate",
        (abbr.upper(), abbr.upper()),
    ).fetchall()
    conn.close()
    if not rows:
        raise HTTPException(404, f"No games found for team '{abbr}'")
    return [dict(r) for r in rows]


@app.get("/api/matchup/{team_a}/{team_b}")
def matchup_games(team_a: str, team_b: str, user: CurrentUser = Depends(require_login)):
    """Every game between two specific teams — the 'any team vs any other
    team' case. Same underlying table as team_games, just filtered to
    games where BOTH teams appear."""
    conn = get_db()
    a, b = team_a.upper(), team_b.upper()
    rows = conn.execute(
        "SELECT gameReferenceId, gameDate, awayTeamAbbrev, homeTeamAbbrev, matchup FROM games "
        "WHERE (homeTeamAbbrev = ? AND awayTeamAbbrev = ?) OR (homeTeamAbbrev = ? AND awayTeamAbbrev = ?) "
        "ORDER BY gameDate",
        (a, b, b, a),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Lines / pairs / special-teams units ──────────────────────────────────
# Same icetime-based approach used earlier in this project by hand: walk
# events in order, sum the time each on-ice combo holds until the next
# combo change, filtered to the requested strength state.

def _clean_refs(raw):
    if not raw:
        return []
    cleaned = raw.replace("\t", "").strip()
    if not cleaned:
        return []
    return [p.strip() for p in cleaned.split(",") if p.strip()]


def _player_names(conn):
    rows = conn.execute("SELECT playerReferenceId, firstName, lastName, position, jersey FROM players").fetchall()
    return {r["playerReferenceId"]: dict(r) for r in rows}


def _player_names_for_team(conn, team_full: str) -> dict:
    """Same shape as _player_names(), but jersey number (and position)
    scoped to a SPECIFIC team instead of the global players dimension
    table. That table's jersey comes from load_nhl.py's rebuild_players(),
    which takes MAX(playerJersey) across every game a player has EVER
    appeared in — for someone who worked with the same number their whole
    career that's harmless, but for a player who wore different numbers
    on different teams, MAX() just picks whichever number sorts highest,
    regardless of which team's game is actually being viewed.

    Real example that surfaced this: Brett Kulak wore #27 with Edmonton,
    got traded to Pittsburgh and wore #77, then to Colorado and wore #27
    again. MAX(playerJersey) always returns 77 (since 77 > 27), so every
    Oilers game showed him as #77 — wrong for that team/context.

    Starts from the global _player_names() as a fallback (so a player
    with no rows tagged to this specific team for some reason still gets
    a reasonable name/jersey instead of nothing), then overrides with
    whatever jersey/position that player actually wore for team_full
    specifically."""
    result = dict(_player_names(conn))
    rows = conn.execute(
        "SELECT playerReferenceId, MAX(playerFirstName) as firstName, MAX(playerLastName) as lastName, "
        "MAX(playerPosition) as position, MAX(playerJersey) as jersey "
        "FROM plays WHERE team = ? AND playerReferenceId IS NOT NULL "
        "GROUP BY playerReferenceId",
        (team_full,),
    ).fetchall()
    for r in rows:
        result[r["playerReferenceId"]] = dict(r)
    return result


def _name(players, pid):
    p = players.get(int(pid)) if pid else None
    if not p:
        return {"id": pid, "name": pid}
    return {"id": pid, "name": f"{p['firstName']} {p['lastName']}", "position": p["position"], "jersey": p["jersey"]}


def _manpower_combo_seconds(conn, game_id: int, team_abbr: str, situation: str):
    """Core walk building raw {tuple-of-player-id-strings: cumulative
    seconds} dicts for forwards and D, for a given 5v5/5v4/4v5 situation.
    Shared by compute_units (which formats/sorts this for the existing
    /lines and /special-teams endpoints, unchanged) and the line/pair
    assignment logic in the /lineup endpoint, which needs the raw
    per-combo seconds to detect duplicate/overlapping groupings and
    identify a 7th defenseman on an irregular-roster night."""
    situation_map = {
        "5v5": ("evenStrength", 5, 5),
        "5v4": ("powerPlay", 5, 4),
        "4v5": ("shortHanded", 4, 5),
    }
    if situation not in situation_map:
        raise HTTPException(400, f"situation must be one of {list(situation_map)}")
    manpower, want_skaters, want_opp = situation_map[situation]
    team_full = full_name(team_abbr)

    rows = conn.execute(
        "SELECT gameTime, team, manpowerSituation, teamSkatersOnIce, opposingTeamSkatersOnIce, "
        "teamForwardsOnIceRefs, teamDefencemenOnIceRefs, "
        "opposingTeamForwardsOnIceRefs, opposingTeamDefencemenOnIceRefs "
        "FROM plays WHERE gameReferenceId = ? ORDER BY gameTime",
        (game_id,),
    ).fetchall()

    home_away = conn.execute(
        "SELECT homeTeamAbbrev, awayTeamAbbrev FROM games WHERE gameReferenceId = ?", (game_id,)
    ).fetchone()
    if not home_away:
        raise HTTPException(404, f"Game {game_id} not found in the games table")
    if not home_away["homeTeamAbbrev"] or not home_away["awayTeamAbbrev"]:
        raise HTTPException(
            422,
            f"Game {game_id} has incomplete team data in the database "
            f"(homeTeamAbbrev={home_away['homeTeamAbbrev']!r}, awayTeamAbbrev={home_away['awayTeamAbbrev']!r}). "
            "This looks like a data quality gap in how this specific game was originally loaded, "
            "not a problem with this request.",
        )
    home_abbr, away_abbr = home_away["homeTeamAbbrev"], home_away["awayTeamAbbrev"]
    if team_abbr.upper() not in (home_abbr, away_abbr):
        raise HTTPException(
            400,
            f"Team '{team_abbr.upper()}' did not play in game {game_id} "
            f"(this game was {away_abbr} @ {home_abbr}). Double-check the game ID.",
        )
    other_abbr = away_abbr if home_abbr == team_abbr.upper() else home_abbr
    other_full = full_name(other_abbr)

    fwd_time = defaultdict(float)
    dmen_time = defaultdict(float)
    n = len(rows)
    CAP = 10.0

    for i in range(n - 1):
        r, nxt = rows[i], rows[i + 1]
        dt = nxt["gameTime"] - r["gameTime"]
        if dt <= 0 or dt > CAP:
            continue
        if r["manpowerSituation"] != manpower:
            continue
        team = r["team"]
        if team == team_full:
            fwd, dmen = _clean_refs(r["teamForwardsOnIceRefs"]), _clean_refs(r["teamDefencemenOnIceRefs"])
            skaters, opp = r["teamSkatersOnIce"], r["opposingTeamSkatersOnIce"]
        elif team == other_full:
            fwd, dmen = _clean_refs(r["opposingTeamForwardsOnIceRefs"]), _clean_refs(r["opposingTeamDefencemenOnIceRefs"])
            skaters, opp = r["opposingTeamSkatersOnIce"], r["teamSkatersOnIce"]
        else:
            continue
        if skaters != want_skaters or opp != want_opp:
            continue
        if fwd:
            fwd_time[tuple(sorted(fwd))] += dt
        if dmen:
            dmen_time[tuple(sorted(dmen))] += dt

    return fwd_time, dmen_time


def _manpower_combo_xg(conn, game_id: int, team_abbr: str, situation: str):
    """Sibling to _manpower_combo_seconds: instead of cumulative icetime,
    accumulates on-ice xGF/xGA per exact forward/D combo for a game, keyed
    the same way (tuple of sorted player-ref strings) so a line/pair's
    displayed roster and its measured xG% can never drift apart.

    Unlike the icetime walk, this doesn't need consecutive-row diffing:
    every play row already carries its own on-ice snapshot (the same
    teamForwardsOnIceRefs/etc. columns _manpower_combo_seconds reads), so
    a qualifying shot is attributed in a single pass, no interval math
    needed.

    Deliberately NOT the same shot-quality convention as the per-skater
    ES xG% used elsewhere (_skater_woi_data): this uses
    expectedGoalsAllShots (not expectedGoalsOnNet) and does NOT exclude
    blocked shots. expectedGoalsOnNet is null for anything that didn't
    reach the goalie (a miss or a block), so restricting a per-line/pair
    metric to it silently drops most attempts for lines with only a
    handful of shots that game, which swings the percentage hard on
    small samples. Confirmed by hand against 2025-10-08 EDM @ CGY (game
    20006): 3 of 4 forward lines matched a known-correct external xG%
    reading exactly once switched to expectedGoalsAllShots with blocked
    shots included, versus being off by double digits under the
    on-net/exclude-blocked convention. period<=3 only applies for
    situation='5v5' since that's the only caller today."""
    situation_map = {
        "5v5": ("evenStrength", 5, 5),
        "5v4": ("powerPlay", 5, 4),
        "4v5": ("shortHanded", 4, 5),
    }
    if situation not in situation_map:
        raise HTTPException(400, f"situation must be one of {list(situation_map)}")
    manpower, want_skaters, want_opp = situation_map[situation]
    team_full = full_name(team_abbr)

    rows = conn.execute(
        "SELECT team, manpowerSituation, teamSkatersOnIce, opposingTeamSkatersOnIce, "
        "teamForwardsOnIceRefs, teamDefencemenOnIceRefs, "
        "opposingTeamForwardsOnIceRefs, opposingTeamDefencemenOnIceRefs, "
        "expectedGoalsAllShots "
        "FROM plays WHERE gameReferenceId = ? AND name = 'shot' AND period <= 3",
        (game_id,),
    ).fetchall()

    home_away = conn.execute(
        "SELECT homeTeamAbbrev, awayTeamAbbrev FROM games WHERE gameReferenceId = ?", (game_id,)
    ).fetchone()
    if not home_away:
        raise HTTPException(404, f"Game {game_id} not found in the games table")
    if not home_away["homeTeamAbbrev"] or not home_away["awayTeamAbbrev"]:
        raise HTTPException(
            422,
            f"Game {game_id} has incomplete team data in the database "
            f"(homeTeamAbbrev={home_away['homeTeamAbbrev']!r}, awayTeamAbbrev={home_away['awayTeamAbbrev']!r}).",
        )
    home_abbr, away_abbr = home_away["homeTeamAbbrev"], home_away["awayTeamAbbrev"]
    if team_abbr.upper() not in (home_abbr, away_abbr):
        raise HTTPException(
            400,
            f"Team '{team_abbr.upper()}' did not play in game {game_id} "
            f"(this game was {away_abbr} @ {home_abbr}). Double-check the game ID.",
        )
    other_abbr = away_abbr if home_abbr == team_abbr.upper() else home_abbr
    other_full = full_name(other_abbr)

    fwd_xg = defaultdict(lambda: {"xgf": 0.0, "xga": 0.0})
    dmen_xg = defaultdict(lambda: {"xgf": 0.0, "xga": 0.0})

    for r in rows:
        if r["manpowerSituation"] != manpower:
            continue
        team = r["team"]
        xg = r["expectedGoalsAllShots"] or 0.0
        if team == team_full:
            fwd, dmen = _clean_refs(r["teamForwardsOnIceRefs"]), _clean_refs(r["teamDefencemenOnIceRefs"])
            skaters, opp = r["teamSkatersOnIce"], r["opposingTeamSkatersOnIce"]
            side = "xgf"
        elif team == other_full:
            fwd, dmen = _clean_refs(r["opposingTeamForwardsOnIceRefs"]), _clean_refs(r["opposingTeamDefencemenOnIceRefs"])
            skaters, opp = r["opposingTeamSkatersOnIce"], r["teamSkatersOnIce"]
            side = "xga"
        else:
            continue
        if skaters != want_skaters or opp != want_opp:
            continue
        if len(fwd) == 3:
            fwd_xg[tuple(sorted(fwd))][side] += xg
        if len(dmen) == 2:
            dmen_xg[tuple(sorted(dmen))][side] += xg

    return fwd_xg, dmen_xg


def _combo_xg_pct(xg_map, combo):
    """esXgPct for one specific trio/pair: None (not 0%) when the combo
    never saw a 5v5 shot at either end, so the frontend can render '—'
    instead of a misleading percentage built from zero shots."""
    entry = xg_map.get(tuple(sorted(str(p) for p in combo)))
    if not entry:
        return None
    xgf, xga = entry["xgf"], entry["xga"]
    if xgf + xga <= 0:
        return None
    return round(xgf / (xgf + xga) * 100, 1)


def _combo_toi_seconds(seconds_map, combo):
    """5v5 icetime for one specific trio/pair, in seconds — same lookup
    convention as _combo_xg_pct, keyed off the exact combo the line/pair
    was assigned from in the first place (_manpower_combo_seconds), so
    this can never disagree with the icetime that put the combo there."""
    return round(seconds_map.get(tuple(sorted(str(p) for p in combo)), 0.0), 1)


def _first_shift_times(conn, game_id: int, team_abbr: str, situation: str):
    """Finds the EARLIEST gameTime each forward/D combo appears together
    on the ice, for a given 5v5/5v4/4v5 situation.

    This exists specifically for 'what lines/pairs did the team START
    the game with' — a coach's alignment can and does change over 60
    minutes (injury, matchups, a bad period), but this feature only
    cares about the opening configuration, so ranking combos by their
    FIRST appearance (not by how much cumulative time they racked up
    over the whole game) is the right sort key here — a line that
    started the game but got split up by the third period would rank
    low on total icetime but should still be correctly identified as a
    starting line.

    Deliberately just 'first appearance,' NOT 'first appearance that
    persists for some minimum continuous shift': an earlier version
    required a combo to hold together unbroken for 10+ seconds before
    it counted, on the theory that a fleeting overlap shouldn't count as
    'the team started with this group.' That turned out to actively
    discard real, heavily-used lines — confirmed on a real game (EDM @
    VGK, 2026-03-08): McDavid's actual trio (530+ total seconds
    together, the #2 most-used forward combo in the whole game per
    /lines) never had a single individual continuous run reach 10
    unbroken seconds in the underlying event stream — the recorded
    on-ice tuple changes from event to event more often than an actual
    substitution would, which reads like a data-granularity artifact
    rather than real line changes — so the whole trio silently
    disappeared from the starting lineup instead of just being ranked
    differently. Simple first-occurrence timing doesn't have that
    failure mode, and the group-size filter applied by the caller
    (exactly 3 forwards / exactly 2 D) already screens out most
    genuinely bogus transitional combos.

    Returns (fwd_first, dmen_first): {combo_tuple: first_gameTime}."""
    situation_map = {
        "5v5": ("evenStrength", 5, 5),
        "5v4": ("powerPlay", 5, 4),
        "4v5": ("shortHanded", 4, 5),
    }
    if situation not in situation_map:
        raise HTTPException(400, f"situation must be one of {list(situation_map)}")
    manpower, want_skaters, want_opp = situation_map[situation]
    team_full = full_name(team_abbr)

    rows = conn.execute(
        "SELECT gameTime, team, manpowerSituation, teamSkatersOnIce, opposingTeamSkatersOnIce, "
        "teamForwardsOnIceRefs, teamDefencemenOnIceRefs, "
        "opposingTeamForwardsOnIceRefs, opposingTeamDefencemenOnIceRefs "
        "FROM plays WHERE gameReferenceId = ? ORDER BY gameTime",
        (game_id,),
    ).fetchall()

    home_away = conn.execute(
        "SELECT homeTeamAbbrev, awayTeamAbbrev FROM games WHERE gameReferenceId = ?", (game_id,)
    ).fetchone()
    if not home_away:
        raise HTTPException(404, f"Game {game_id} not found in the games table")
    home_abbr, away_abbr = home_away["homeTeamAbbrev"], home_away["awayTeamAbbrev"]
    other_abbr = away_abbr if home_abbr == team_abbr.upper() else home_abbr
    other_full = full_name(other_abbr)

    fwd_first, dmen_first = {}, {}
    for r in rows:
        if r["manpowerSituation"] != manpower:
            continue
        team = r["team"]
        if team == team_full:
            fwd, dmen = _clean_refs(r["teamForwardsOnIceRefs"]), _clean_refs(r["teamDefencemenOnIceRefs"])
            skaters, opp = r["teamSkatersOnIce"], r["opposingTeamSkatersOnIce"]
        elif team == other_full:
            fwd, dmen = _clean_refs(r["opposingTeamForwardsOnIceRefs"]), _clean_refs(r["opposingTeamDefencemenOnIceRefs"])
            skaters, opp = r["opposingTeamSkatersOnIce"], r["teamSkatersOnIce"]
        else:
            continue
        if skaters != want_skaters or opp != want_opp:
            continue
        gtime = r["gameTime"]
        if fwd:
            combo = tuple(sorted(fwd))
            if combo not in fwd_first or gtime < fwd_first[combo]:
                fwd_first[combo] = gtime
        if dmen:
            combo = tuple(sorted(dmen))
            if combo not in dmen_first or gtime < dmen_first[combo]:
                dmen_first[combo] = gtime

    return fwd_first, dmen_first


# ── Opponent PK report (McFarland's Pre-Scout Prep) ──────────────────────
# Everything here is 4v5 (the opponent shorthanded) and purely PSF-derived
# — no NHL API dependency, so it's scoped to whatever games actually exist
# in nhl.db for this team (which may be fewer than their real full
# schedule, if not every game's been coded yet — that's expected and
# normal here, unlike the NHL-API-sourced scoring stats elsewhere in the
# prep payload).

def _resolve_game_ids(conn, team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> list[int]:
    """Same season-XOR-dates convention as goalie_gsax/skater_woi above —
    dates scopes to specific games (e.g. a last-5 cut) by gameDate rather
    than game ID, since PSF's own gameReferenceId numbering doesn't match
    the NHL's."""
    game_q = "SELECT gameReferenceId FROM games WHERE (homeTeamAbbrev = ? OR awayTeamAbbrev = ?)"
    game_params = [team_abbr, team_abbr]
    if dates:
        placeholders = ",".join("?" for _ in dates)
        game_q += f" AND gameDate IN ({placeholders})"
        game_params += dates
    else:
        game_q += " AND season = ?"
        game_params.append(season)
    return [r["gameReferenceId"] for r in conn.execute(game_q, game_params).fetchall()]


def _pk_units_data(conn, team_abbr: str, game_ids: list[int]) -> dict:
    """Aggregates _manpower_combo_seconds (already validated for single
    games via the /lineup feature) across every game in scope, to find
    the opponent's most-used PK personnel groupings over a stretch rather
    than just one night.

    Forward combo size is 2, not 3: standard 4v5 PK deployment is 2F+2D
    (unlike a 5v5 line, which is 3F). Using 3 here (copied from the 5v5
    convention without adjusting it) meant the forward-combo filter never
    matched anything — confirmed live: D pairs showed up fine, forwards
    always came back empty, on real data."""
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)
    fwd_totals: dict = defaultdict(float)
    dmen_totals: dict = defaultdict(float)
    for gid in game_ids:
        try:
            fwd_time, dmen_time = _manpower_combo_seconds(conn, gid, team_abbr, "4v5")
        except HTTPException:
            continue  # e.g. incomplete team data for that specific game — skip, don't fail the whole report
        for combo, secs in fwd_time.items():
            fwd_totals[combo] += secs
        for combo, secs in dmen_time.items():
            dmen_totals[combo] += secs

    def top_combos(totals, size, n):
        matching = [(combo, secs) for combo, secs in totals.items() if len(combo) == size]
        matching.sort(key=lambda x: -x[1])
        return [{"players": [_name(players, pid) for pid in combo], "seconds": round(secs, 1)} for combo, secs in matching[:n]]

    return {
        "forwards": top_combos(fwd_totals, 2, 3),
        "defense": top_combos(dmen_totals, 2, 3),
    }


def _faceoff_pct_data(conn, team_abbr: str, game_ids: list[int]) -> dict:
    """Faceoff win% for team_abbr specifically while shorthanded (4v5).
    manpowerSituation is relative to the acting player's own team, so this
    is simply team=team_full AND manpowerSituation='shortHanded'."""
    if not game_ids:
        return {"won": 0, "total": 0, "pct": None}
    team_full = full_name(team_abbr)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'faceoff' AND team = ? AND manpowerSituation = 'shortHanded'
            GROUP BY outcome""",
        game_ids + [team_full],
    ).fetchall()
    total = sum(r["n"] for r in rows)
    won = sum(r["n"] for r in rows if r["outcome"] == "successful")
    return {"won": won, "total": total, "pct": round(won / total * 100, 1) if total else None}


def _faceoff_breakdown_data(conn, team_abbr: str, game_ids: list[int]) -> list:
    """Per-player faceoff record while shorthanded (4v5) — the top 3
    players by draws taken, each with won/lost/pct. Same manpower/team
    filter as _faceoff_pct_data, just grouped by the taking player
    (playerReferenceId, the acting player on a faceoff row) instead of
    summed team-wide."""
    if not game_ids:
        return []
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT playerReferenceId, outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'faceoff' AND team = ? AND manpowerSituation = 'shortHanded'
              AND playerReferenceId IS NOT NULL
            GROUP BY playerReferenceId, outcome""",
        game_ids + [team_full],
    ).fetchall()
    totals: dict = {}
    for r in rows:
        pid = r["playerReferenceId"]
        t = totals.setdefault(pid, {"won": 0, "lost": 0})
        if r["outcome"] == "successful":
            t["won"] += r["n"]
        else:
            t["lost"] += r["n"]
    result = []
    for pid, t in totals.items():
        total = t["won"] + t["lost"]
        result.append({
            "player": _name(players, pid),
            "won": t["won"], "lost": t["lost"], "total": total,
            "pct": round(t["won"] / total * 100, 1) if total else None,
        })
    result.sort(key=lambda x: -x["total"])
    return result[:3]


def _zone_entry_defense_data(conn, team_abbr: str, game_ids: list[int]) -> dict:
    """Controlled zone entries allowed vs. denied while this team is
    shorthanded (4v5). Per Scoring_Chance_Derivation_Reference.md's
    documented convention: entries are logged as a 'controlledentryagainst'
    row under the DEFENDING team, and outcome=='failed' counterintuitively
    means the entry SUCCEEDED (the defense's attempt to stop it failed) —
    anything else means the entry was denied.

    NOTE: unlike the scoring-chance and lineup logic elsewhere in this
    file, this specific convention hasn't been re-verified against a real
    game's actual entry count here — it's taken directly from that
    reference doc. Worth spot-checking against one game with a known
    entry-defense tally before trusting this number in front of the
    room."""
    if not game_ids:
        return {"allowed": 0, "denied": 0, "pct": None}
    team_full = full_name(team_abbr)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'controlledentryagainst' AND team = ? AND manpowerSituation = 'shortHanded'
            GROUP BY outcome""",
        game_ids + [team_full],
    ).fetchall()
    allowed = sum(r["n"] for r in rows if r["outcome"] == "failed")
    denied = sum(r["n"] for r in rows if r["outcome"] != "failed")
    total = allowed + denied
    return {"allowed": allowed, "denied": denied, "pct": round(allowed / total * 100, 1) if total else None}


def _zone_entry_defense_breakdown_data(conn, team_abbr: str, game_ids: list[int]) -> list:
    """Top 5 of team_abbr's own defenders who face the most controlled
    zone entries against them while shorthanded, with each player's
    personal denial rate (denied / (denied + allowed)).

    Confirmed directly (via a reported bug on the PP-offense mirror of
    this function, below): a 'controlledentryagainst' row's own
    playerReferenceId is tagged to the DEFENDING side, not the attacking
    carrier — the opposite of what an earlier version of this assumed
    (that version joined on_ice to work around a wrong assumption; this
    is the simpler, now-confirmed-correct direct query). Same
    controlledentryagainst/outcome convention as _zone_entry_defense_data
    above, including its same not-yet-independently-verified caveat on
    the entry aggregate itself."""
    if not game_ids:
        return []
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT playerReferenceId, outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'controlledentryagainst' AND team = ? AND manpowerSituation = 'shortHanded'
              AND playerReferenceId IS NOT NULL
            GROUP BY playerReferenceId, outcome""",
        game_ids + [team_full],
    ).fetchall()
    totals: dict = {}
    for r in rows:
        pid = r["playerReferenceId"]
        t = totals.setdefault(pid, {"denied": 0, "allowed": 0})
        if r["outcome"] == "failed":
            t["allowed"] += r["n"]  # entry succeeded despite this player being tagged on it
        else:
            t["denied"] += r["n"]   # entry denied
    result = []
    for pid, t in totals.items():
        total = t["denied"] + t["allowed"]
        result.append({
            "player": _name(players, pid),
            "denied": t["denied"], "allowed": t["allowed"], "total": total,
            "pct": round(t["denied"] / total * 100, 1) if total else None,
        })
    result.sort(key=lambda x: -x["total"])
    return result[:5]


def _pk_report_threaded(team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> dict:
    """Combines units/faceoffs/zone-entry-defense into one payload; run via
    asyncio.to_thread from the async prep-cache builder, same reasoning as
    _goalie_gsax_threaded/_skater_woi_threaded — these are blocking SQLite
    calls and must not run directly on the event loop."""
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=dates)
        return {
            "gamesInDatabase": len(game_ids),
            "units": _pk_units_data(conn, team_abbr, game_ids),
            "faceoffs": _faceoff_pct_data(conn, team_abbr, game_ids),
            "faceoffBreakdown": _faceoff_breakdown_data(conn, team_abbr, game_ids),
            "zoneEntryDefense": _zone_entry_defense_data(conn, team_abbr, game_ids),
            "zoneEntryDefenseBreakdown": _zone_entry_defense_breakdown_data(conn, team_abbr, game_ids),
        }
    finally:
        conn.close()


def _manpower_full_unit_seconds(conn, game_id: int, team_abbr: str, situation: str):
    """Like _manpower_combo_seconds, but keys on the FULL on-ice skater
    unit — forwards + D together as one group — instead of tracking
    forward combos and D combos independently. A power play is deployed
    and coached as one group of 5, not two separate position groups, so
    Smith's PP units need this instead of the PK's forwards/D split."""
    situation_map = {
        "5v5": ("evenStrength", 5, 5),
        "5v4": ("powerPlay", 5, 4),
        "4v5": ("shortHanded", 4, 5),
    }
    if situation not in situation_map:
        raise HTTPException(400, f"situation must be one of {list(situation_map)}")
    manpower, want_skaters, want_opp = situation_map[situation]
    team_full = full_name(team_abbr)

    rows = conn.execute(
        "SELECT gameTime, team, manpowerSituation, teamSkatersOnIce, opposingTeamSkatersOnIce, "
        "teamForwardsOnIceRefs, teamDefencemenOnIceRefs, "
        "opposingTeamForwardsOnIceRefs, opposingTeamDefencemenOnIceRefs "
        "FROM plays WHERE gameReferenceId = ? ORDER BY gameTime",
        (game_id,),
    ).fetchall()

    home_away = conn.execute(
        "SELECT homeTeamAbbrev, awayTeamAbbrev FROM games WHERE gameReferenceId = ?", (game_id,)
    ).fetchone()
    if not home_away:
        raise HTTPException(404, f"Game {game_id} not found in the games table")
    if not home_away["homeTeamAbbrev"] or not home_away["awayTeamAbbrev"]:
        raise HTTPException(
            422,
            f"Game {game_id} has incomplete team data in the database "
            f"(homeTeamAbbrev={home_away['homeTeamAbbrev']!r}, awayTeamAbbrev={home_away['awayTeamAbbrev']!r}). "
            "This looks like a data quality gap in how this specific game was originally loaded, "
            "not a problem with this request.",
        )
    home_abbr, away_abbr = home_away["homeTeamAbbrev"], home_away["awayTeamAbbrev"]
    if team_abbr.upper() not in (home_abbr, away_abbr):
        raise HTTPException(
            400,
            f"Team '{team_abbr.upper()}' did not play in game {game_id} "
            f"(this game was {away_abbr} @ {home_abbr}). Double-check the game ID.",
        )
    other_abbr = away_abbr if home_abbr == team_abbr.upper() else home_abbr
    other_full = full_name(other_abbr)

    unit_time = defaultdict(float)
    n = len(rows)
    CAP = 10.0

    for i in range(n - 1):
        r, nxt = rows[i], rows[i + 1]
        dt = nxt["gameTime"] - r["gameTime"]
        if dt <= 0 or dt > CAP:
            continue
        if r["manpowerSituation"] != manpower:
            continue
        team = r["team"]
        if team == team_full:
            fwd, dmen = _clean_refs(r["teamForwardsOnIceRefs"]), _clean_refs(r["teamDefencemenOnIceRefs"])
            skaters, opp = r["teamSkatersOnIce"], r["opposingTeamSkatersOnIce"]
        elif team == other_full:
            fwd, dmen = _clean_refs(r["opposingTeamForwardsOnIceRefs"]), _clean_refs(r["opposingTeamDefencemenOnIceRefs"])
            skaters, opp = r["opposingTeamSkatersOnIce"], r["teamSkatersOnIce"]
        else:
            continue
        if skaters != want_skaters or opp != want_opp:
            continue
        if fwd or dmen:
            key = (tuple(sorted(fwd or [])), tuple(sorted(dmen or [])))
            unit_time[key] += dt

    return unit_time


def _pp_units_data(conn, team_abbr: str, game_ids: list[int]) -> list:
    """Top 3 full 5-skater PP units (forwards + D together, not tracked
    separately) by cumulative combined seconds across every game in
    scope — this is how a power play is actually coached and deployed,
    unlike the PK's 2F+2D convention where forwards/D are more
    interchangeable across pairings. Forwards are listed before
    defensemen within each unit."""
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)
    unit_totals: dict = defaultdict(float)  # (fwd_tuple, dmen_tuple) -> seconds
    for gid in game_ids:
        try:
            combo_time = _manpower_full_unit_seconds(conn, gid, team_abbr, "5v4")
        except HTTPException:
            continue
        for key, secs in combo_time.items():
            unit_totals[key] += secs

    ranked = sorted(unit_totals.items(), key=lambda x: -x[1])[:3]
    return [
        {
            "players": [_name(players, pid) for pid in fwd_combo] + [_name(players, pid) for pid in dmen_combo],
            "seconds": round(secs, 1),
        }
        for (fwd_combo, dmen_combo), secs in ranked
    ]


def _pp_faceoff_pct_data(conn, team_abbr: str, game_ids: list[int]) -> dict:
    """Mirror of _faceoff_pct_data for the power play (manpowerSituation='powerPlay')."""
    if not game_ids:
        return {"won": 0, "total": 0, "pct": None}
    team_full = full_name(team_abbr)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'faceoff' AND team = ? AND manpowerSituation = 'powerPlay'
            GROUP BY outcome""",
        game_ids + [team_full],
    ).fetchall()
    total = sum(r["n"] for r in rows)
    won = sum(r["n"] for r in rows if r["outcome"] == "successful")
    return {"won": won, "total": total, "pct": round(won / total * 100, 1) if total else None}


def _pp_faceoff_breakdown_data(conn, team_abbr: str, game_ids: list[int]) -> list:
    """Mirror of _faceoff_breakdown_data for the power play."""
    if not game_ids:
        return []
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT playerReferenceId, outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'faceoff' AND team = ? AND manpowerSituation = 'powerPlay'
              AND playerReferenceId IS NOT NULL
            GROUP BY playerReferenceId, outcome""",
        game_ids + [team_full],
    ).fetchall()
    totals: dict = {}
    for r in rows:
        pid = r["playerReferenceId"]
        t = totals.setdefault(pid, {"won": 0, "lost": 0})
        if r["outcome"] == "successful":
            t["won"] += r["n"]
        else:
            t["lost"] += r["n"]
    result = []
    for pid, t in totals.items():
        total = t["won"] + t["lost"]
        result.append({
            "player": _name(players, pid),
            "won": t["won"], "lost": t["lost"], "total": total,
            "pct": round(t["won"] / total * 100, 1) if total else None,
        })
    result.sort(key=lambda x: -x["total"])
    return result[:3]


def _zone_entry_offense_data(conn, team_abbr: str, game_ids: list[int]) -> dict:
    """Inverse of _zone_entry_defense_data: controlled zone entries this
    team achieves (successful/denied) while ON the power play, across its
    own games. Entries are logged as a 'controlledentryagainst' row under
    the DEFENDING team (per Scoring_Chance_Derivation_Reference.md), so to
    find entries FOR team_abbr we look at rows tagged under the OTHER
    team in each of team_abbr's games (team != team_full) while that
    other team is shorthanded (manpowerSituation='shortHanded' — their
    situation, since they're defending against team_abbr's power play).
    outcome=='failed' still counterintuitively means the entry succeeded."""
    if not game_ids:
        return {"successful": 0, "denied": 0, "pct": None}
    team_full = full_name(team_abbr)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""SELECT outcome, COUNT(*) as n FROM plays
            WHERE gameReferenceId IN ({placeholders})
              AND name = 'controlledentryagainst' AND team != ? AND manpowerSituation = 'shortHanded'
            GROUP BY outcome""",
        game_ids + [team_full],
    ).fetchall()
    successful = sum(r["n"] for r in rows if r["outcome"] == "failed")
    denied = sum(r["n"] for r in rows if r["outcome"] != "failed")
    total = successful + denied
    return {"successful": successful, "denied": denied, "pct": round(successful / total * 100, 1) if total else None}


def _zone_entry_offense_breakdown_data(conn, team_abbr: str, game_ids: list[int]) -> list:
    """Top 5 players who most often carry/pass the puck into the zone on
    team_abbr's power play, with each player's personal success rate.
    The point of this stat is who they like to GIVE the puck TO in order
    to enter the zone, so for a pass-based entry, credit goes to whoever
    actually receives it and crosses the line — not whoever threw the
    pass.

    Two earlier versions of this got the outcome wrong in different ways:
    - Using 'controlledentryagainst' directly tagged the DEFENDING
      player, not the attacking carrier (confirmed by a reported bug
      where it surfaced the wrong team's players entirely).
    - Using each attacking-side event's OWN outcome (carry/reception)
      turned out not to work either — confirmed directly: across 2,141
      sampled 'carry' events, EVERY single one is outcome=='successful'.
      A carry only seems to get logged when the puck-handling itself
      completes; a failed/broken-up entry attempt apparently doesn't
      generate one at all, so there's no real success/denied signal on
      the attacking event's own outcome field.

    Fixed by TIME-MATCHING instead: 'controlledentryagainst' (tagged to
    the defender) is still the authoritative record of whether an entry
    attempt actually succeeded or got denied (outcome=='failed' means
    the entry succeeded, the confusing convention documented in
    Scoring_Chance_Derivation_Reference.md) — it just can't identify the
    attacking player. Spot-checked in the raw data: a
    'controlledentryagainst' row and its corresponding attacking-side
    'carry'/'reception' row land within a fraction of a second of each
    other in the same game (e.g. gameTime 29.267 vs. 29.3 for the same
    physical entry, logged from both sides). So this pairs each
    'controlledentryagainst' event with whichever qualifying attacking
    event (a 'carry' landing in the OZ with the previous event in the
    neutral/defensive zone — an actual crossing, not just a carry that
    happens to occur while already in the o-zone — or a 'reception' of
    type 'rush'/'ozentry') is CLOSEST in gameTime, within a 3-second
    window, and credits that player with the authoritative outcome.
    Entries with no attacking-side event inside that window (most likely
    genuinely denied attempts that never got established well enough to
    log a carry/reception at all) are simply not attributable to a
    specific player and are excluded — there's no player to credit if
    nothing on the attacking side was ever logged for that attempt."""
    if not game_ids:
        return []
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)
    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""WITH seq AS (
                SELECT id, gameReferenceId, team, manpowerSituation, playerReferenceId,
                       outcome, name, type, zone, gameTime,
                       LAG(zone) OVER (PARTITION BY gameReferenceId ORDER BY id) AS prevZone
                FROM plays
                WHERE gameReferenceId IN ({placeholders})
            ),
            entries AS (
                SELECT gameReferenceId, id AS entry_id, gameTime AS entry_time, outcome
                FROM seq
                WHERE name = 'controlledentryagainst' AND team != ? AND manpowerSituation = 'shortHanded'
            ),
            attackers AS (
                SELECT gameReferenceId, id AS attacker_id, gameTime AS attacker_time, playerReferenceId
                FROM seq
                WHERE team = ? AND manpowerSituation = 'powerPlay' AND playerReferenceId IS NOT NULL
                  AND (
                        (name = 'carry' AND zone = 'oz' AND prevZone IN ('nz', 'dz'))
                     OR (name = 'reception' AND type IN ('rush', 'ozentry'))
                  )
            ),
            matched AS (
                SELECT e.entry_id, e.outcome, a.playerReferenceId,
                       ROW_NUMBER() OVER (
                           PARTITION BY e.gameReferenceId, e.entry_id
                           ORDER BY ABS(a.attacker_time - e.entry_time)
                       ) AS rn
                FROM entries e
                JOIN attackers a
                  ON a.gameReferenceId = e.gameReferenceId
                 AND ABS(a.attacker_time - e.entry_time) <= 3.0
            )
            SELECT playerReferenceId, outcome, COUNT(*) as n
            FROM matched
            WHERE rn = 1
            GROUP BY playerReferenceId, outcome""",
        game_ids + [team_full, team_full],
    ).fetchall()
    totals: dict = {}
    for r in rows:
        pid = r["playerReferenceId"]
        t = totals.setdefault(pid, {"successful": 0, "denied": 0})
        if r["outcome"] == "failed":
            t["successful"] += r["n"]
        else:
            t["denied"] += r["n"]
    result = []
    for pid, t in totals.items():
        total = t["successful"] + t["denied"]
        result.append({
            "player": _name(players, pid),
            "successful": t["successful"], "denied": t["denied"], "total": total,
            "pct": round(t["successful"] / total * 100, 1) if total else None,
        })
    result.sort(key=lambda x: -x["total"])
    return result[:5]


def _pp_report_threaded(team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> dict:
    """Smith's tab is the inverse of McFarland's _pk_report_threaded —
    same shape, but power-play personnel/faceoffs/zone-entry-offense
    instead of PK. Run via asyncio.to_thread for the same blocking-SQLite
    reason as the PK version."""
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=dates)
        return {
            "gamesInDatabase": len(game_ids),
            "units": _pp_units_data(conn, team_abbr, game_ids),
            "faceoffs": _pp_faceoff_pct_data(conn, team_abbr, game_ids),
            "faceoffBreakdown": _pp_faceoff_breakdown_data(conn, team_abbr, game_ids),
            "zoneEntryOffense": _zone_entry_offense_data(conn, team_abbr, game_ids),
            "zoneEntryOffenseBreakdown": _zone_entry_offense_breakdown_data(conn, team_abbr, game_ids),
        }
    finally:
        conn.close()





def compute_units(conn, game_id: int, team_abbr: str, situation: str, players: dict | None = None):
    """situation: '5v5' | '5v4' | '4v5' — returns forward/D combo icetime for team_abbr.

    players: pass in an already-fetched _player_names_for_team() result to
    skip re-querying it (that query scans the team's entire play history,
    not just this game — callers building multiple unit breakdowns for the
    same team/game, like _build_lineup_sync_data, should fetch it once and
    share it instead of paying for it again per situation)."""
    fwd_time, dmen_time = _manpower_combo_seconds(conn, game_id, team_abbr, situation)
    if players is None:
        team_full = full_name(team_abbr)
        players = _player_names_for_team(conn, team_full)

    def fmt(combo_time):
        return [
            {"players": [_name(players, p) for p in combo], "icetimeSeconds": round(t, 1)}
            for combo, t in sorted(combo_time.items(), key=lambda x: -x[1])
        ]

    return {"forwardUnits": fmt(fwd_time), "defenseUnits": fmt(dmen_time)}


@app.get("/api/game/{game_id}/lines")
def game_lines(game_id: int, team: str = Query(..., description="Team abbreviation, e.g. EDM"), user: CurrentUser = Depends(require_login)):
    conn = get_db()
    game_or_404(conn, game_id)
    result = compute_units(conn, game_id, team.upper(), "5v5")
    conn.close()
    return {"gameId": game_id, "team": team.upper(), "situation": "5v5", **result}


@app.get("/api/game/{game_id}/special-teams")
def game_special_teams(game_id: int, team: str = Query(...), situation: str = Query("5v4"), user: CurrentUser = Depends(require_login)):
    conn = get_db()
    game_or_404(conn, game_id)
    result = compute_units(conn, game_id, team.upper(), situation)
    conn.close()
    return {"gameId": game_id, "team": team.upper(), "situation": situation, **result}


# ── Full game lineup (lines / pairs / goalies / PP units) ────────────────
# Assembles compute_units' raw icetime groupings into what a coach
# actually wants for a game recap: 4 labeled forward lines (with a
# best-guess center), 3 labeled D pairs (with LD/RD), starting/backup
# goalie, and PP1/PP2 as full 5-skater groups. Each derivation below was
# validated by hand against a real game before being folded in here —
# see the 2026-08-01 session notes.

async def _fetch_nhl_position_map(team_abbr: str, season: str | None = None):
    """Fetches the NHL roster's own official position code (C/L/R/D/G)
    per player, keyed by NHL player id — which is the same id space as
    our PSF playerReferenceId (confirmed elsewhere in this project, e.g.
    McDavid = 8478402 in both). Used alongside faceoff counts to pick a
    trio's center: a team's own official center designation is a
    steadier signal than a single game's faceoff split, which can
    occasionally point at a winger who happened to take more draws that
    night — confirmed directly: EDM @ VGK on 2026-03-08, McDavid (an
    every-night center) was out-drawn by Nugent-Hopkins on their own
    line, and faceoff-count-only logic picked Nugent-Hopkins as center
    for that game, which is wrong. Returns {} (falls back to faceoffs
    alone) if the fetch fails for any reason — this is a refinement,
    not a hard dependency.

    Pass `season` (e.g. "20252026") whenever the caller has one — uses
    /roster/{team}/{season} instead of /roster/{team}/current, which
    matters a lot in the offseason: "current" is either empty or
    reflects in-progress free agency, not the roster as it stood for
    the season actually being reported on. Confirmed directly: this
    caused every player in a season report to show as "no longer on
    the team" during the summer, since /current came back essentially
    empty at the time. Falls back to /current only when no season is
    given at all."""
    cache_key = (team_abbr, season or "current")
    if cache_key in _nhl_roster_cache:
        return _nhl_roster_cache[cache_key]
    endpoint = f"{team_abbr}/{season}" if season else f"{team_abbr}/current"
    try:
        async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
            r = await client.get(f"{NHL_API_BASE}/roster/{endpoint}")
        if r.status_code != 200:
            return {}
        data = r.json()
    except httpx.RequestError:
        return {}
    out = {}
    for grp in ("forwards", "defensemen", "goalies"):
        for p in data.get(grp, []):
            if "id" in p:
                out[p["id"]] = p.get("positionCode")
    _nhl_roster_cache[cache_key] = out
    return out


async def _faceoff_centers(conn, game_id: int, team_full: str, team_abbr: str, trios: list):
    """For each forward trio, picks a center by combining two signals:

    1. Faceoffs taken (for team_full, in this game) — the primary
       signal, and deliberately kept scoped to THIS game rather than
       smoothed across the season: who actually centers a line on a
       given night can genuinely differ from a player's usual role for
       real, in-game reasons (an injury, a tactical change), and this
       feature should reflect what actually happened that night, not a
       season-long average papering over it. (A season-long faceoff
       tiebreaker was tried and reverted for exactly this reason — see
       git history / 2026-08-01 session notes: it looked like the right
       fix for what appeared to be noisy single-game data, but risked
       overriding real, legitimate in-game situations, like a usual 1C
       playing through an injury and taking few or no draws that
       night.)
    2. The NHL roster's own official position code (see
       _fetch_nhl_position_map) — used only to break ties when faceoffs
       alone don't distinguish a real center from a winger who happened
       to win a draw or two (PSF's own playerPosition tag on individual
       events isn't reliable enough for this — mislabels wingers as 'C'
       often enough that a real EDM line — Savoie/Roslovic/Dach/Samanski
       — were ALL tagged 'C' despite clearly playing wing on their
       trios).

    If exactly one real trio member is roster-listed as 'C', that's
    used directly. If zero or more-than-one are (e.g. two natural
    centers sharing a trio), THIS game's faceoff count breaks the tie
    among whichever candidates are relevant — all real trio members if
    no roster signal, or just the roster-tagged centers if there's more
    than one.

    Returns a list of center ids aligned with `trios` by index (not a
    dict keyed by trio content) — trios can contain a blanked (None)
    slot on an irregular-roster night, so trio contents alone aren't a
    safe/stable dict key here. None entries are ignored when picking the
    center; a trio reduced to a single real player is trivially "center"
    by default (there's no one else to compare against).

    Split into a sync DB piece (_faceoff_counts_for_game) and a pure
    center-picking piece (_pick_centers) below — game_lineup calls those
    directly so the DB read can run inside its threaded builder instead
    of blocking the event loop; this combined async version is kept for
    any future caller that doesn't need that split."""
    faceoff_counts = _faceoff_counts_for_game(conn, game_id, team_full)
    position_map = await _fetch_nhl_position_map(team_abbr)
    return _pick_centers(trios, faceoff_counts, position_map)


def _faceoff_counts_for_game(conn, game_id: int, team_full: str) -> dict:
    """Just the sync DB half of _faceoff_centers — a single cheap
    GROUP BY on one game's faceoff rows, safe to call from inside a
    threaded builder."""
    rows = conn.execute(
        "SELECT playerReferenceId, COUNT(*) as n FROM plays "
        "WHERE gameReferenceId = ? AND name = 'faceoff' AND team = ? "
        "GROUP BY playerReferenceId",
        (game_id, team_full),
    ).fetchall()
    return {r["playerReferenceId"]: r["n"] for r in rows}


def _pick_centers(trios: list, faceoff_counts: dict, position_map: dict):
    """Pure center-selection logic (no I/O) — see _faceoff_centers'
    docstring above for the actual signal-combining rules."""
    centers = []
    for trio in trios:
        real = [p for p in trio if p is not None]
        if not real:
            centers.append(None)
            continue
        roster_centers = [p for p in real if position_map.get(int(p)) == "C"]
        candidates = roster_centers if len(roster_centers) == 1 else real if not roster_centers else roster_centers
        centers.append(max(candidates, key=lambda pid: faceoff_counts.get(int(pid), 0)))
    return centers


def _dpair_sides(conn, game_id: int, team_full: str, pairs: list):
    """LD/RD assignment per D pair, from whole-game pooled LPR yAdjCoord
    averages (see Coordinate_Adjustment_Wing_Side_Reference.md): negative
    mean y = left side, positive = right side. Validated against 3 known
    real-world pairs (2 clean, 1 weak-but-correct-side). Unlike forward
    wing side — which coordinate data can't resolve reliably — this
    works because D structurally hold a side (DZ retrievals, blue-line
    coverage) rather than crossing the middle the way wingers do.

    Returns a list aligned with `pairs` by index, same reasoning as
    _faceoff_centers above (pairs are now guaranteed-clean by
    _assign_stable_groups, so a None slot shouldn't occur here in
    practice, but this stays index-based for consistency)."""
    out = []
    for pair in pairs:
        avg_y = {}
        for pid in pair:
            if pid is None:
                continue
            row = conn.execute(
                "SELECT AVG(yAdjCoord) as avg_y FROM plays "
                "WHERE gameReferenceId = ? AND name = 'lpr' AND team = ? "
                "AND playerReferenceId = ? AND yAdjCoord IS NOT NULL",
                (game_id, team_full, int(pid)),
            ).fetchone()
            avg_y[pid] = row["avg_y"] if row and row["avg_y"] is not None else 0.0
        ordered = sorted(avg_y.keys(), key=lambda pid: avg_y[pid])
        out.append({"LD": ordered[0] if ordered else None, "RD": ordered[-1] if len(ordered) > 1 else None})
    return out


def _dedupe_ordered_groups(groups: list):
    """groups: list of player-id lists already ordered by descending
    combo icetime (as the forward-lines selection produces). If a player
    appears in more than one group — a real possibility on an irregular-
    roster night (11 forwards instead of 12, or a mid-game line shuffle
    that gives one skater meaningful icetime with two different trios)
    — keeps him only in the first (highest-icetime, i.e. earlier-listed)
    group and blanks (sets to None) the redundant slot in every later
    group, rather than silently showing him as if he belonged to
    multiple lines at once."""
    seen = set()
    out = []
    for group in groups:
        new_group = []
        for p in group:
            if p in seen:
                new_group.append(None)
            else:
                new_group.append(p)
                seen.add(p)
        out.append(new_group)
    return out


def _assign_stable_groups_from_sorted(sorted_combos: list, group_size: int, max_groups: int):
    """Core of _assign_stable_groups (below), generalized to take an
    already-ordered list of (combo, sort_value) pairs — the priority
    order can be 'most icetime first' (used for PK/PP-style 'who plays
    the most') or 'earliest confirmed shift first' (used for 'who did
    the team START the game with'), and this function doesn't need to
    care which, since it never looks at sort_value itself past using
    the given order.

    Greedily accepts a combo only if every member is still unused,
    guaranteeing the resulting groups never repeat a player — which
    naturally surfaces anyone who never had one clean, consistent
    group as leftover (a 7th defenseman, an extra forward on an
    irregular-roster night) instead of forcing them into a misleading
    grouping.

    Returns (groups, leftover): groups is a list of player-id lists;
    leftover is player ids never captured in a clean group, ordered by
    their own first (highest-priority) appearance in sorted_combos."""
    all_players = set()
    first_appearance = {}
    for idx, (combo, _val) in enumerate(sorted_combos):
        all_players.update(combo)
        for p in combo:
            if p not in first_appearance:
                first_appearance[p] = idx

    used = set()
    groups = []
    for combo, _val in sorted_combos:
        if len(groups) >= max_groups:
            break
        if len(combo) == group_size and all(p not in used for p in combo):
            groups.append(list(combo))
            used.update(combo)

    leftover = sorted(all_players - used, key=lambda p: first_appearance[p])
    return groups, leftover


def _assign_stable_groups(combo_time: dict, group_size: int, max_groups: int):
    """Icetime-ranked convenience wrapper around
    _assign_stable_groups_from_sorted — used wherever 'who plays the
    most together' (rather than 'who started together') is the right
    question, e.g. PK pairs. See that function's docstring for the
    core algorithm; see _first_shift_times for the starting-lineup use
    case, which sorts by first-shift time instead of total icetime."""
    sorted_combos = sorted(combo_time.items(), key=lambda x: -x[1])
    return _assign_stable_groups_from_sorted(sorted_combos, group_size, max_groups)


def _starting_backup_goalie(conn, game_id: int, team_full: str):
    """Starting goalie = whoever is in net at the FIRST on-ice event of
    the game for this team — deliberately NOT whoever logs the most
    icetime, since a starter pulled early can end up with less recorded
    time than the goalie who relieves them. Backup = the next distinct
    goalie who appears on this team's side at any later point."""
    rows = conn.execute(
        "SELECT gameTime, team, teamGoalieOnIceRef, opposingTeamGoalieOnIceRef "
        "FROM plays WHERE gameReferenceId = ? ORDER BY gameTime",
        (game_id,),
    ).fetchall()
    seen, seen_set = [], set()
    for r in rows:
        g = r["teamGoalieOnIceRef"] if r["team"] == team_full else r["opposingTeamGoalieOnIceRef"]
        if g and g not in seen_set:
            seen_set.add(g)
            seen.append(g)
    starter = seen[0] if seen else None
    backup = seen[1] if len(seen) > 1 else None
    return starter, backup


async def _nhl_boxscore_stats_side(team_abbr: str, game_date: str, season: str):
    """Looks up the NHL's own game id for team_abbr on game_date (via
    club-schedule-season) and fetches that game's boxscore, returning
    just this team's playerByGameStats side (forwards/defense/goalies
    arrays with real dressed-roster stats). Used for two things PSF
    alone can't fully answer:
      1. A backup goalie who dressed but never played — PSF's play-by-
         play has zero record of a player who never touches the ice, but
         the NHL boxscore's goalies array lists both, with a `starter`
         flag.
      2. A self-check on how many defensemen actually dressed for this
         game, to catch our own derivation missing/mis-clustering a
         defenseman on an irregular-roster night.
    Returns None if anything in this chain fails (network issue, no
    schedule match, unexpected shape) — this is a cross-check /
    fallback, not a hard dependency; the rest of the lineup response
    still works fine without it."""
    try:
        sched_key = (team_abbr, season)
        if sched_key in _nhl_schedule_cache:
            games = _nhl_schedule_cache[sched_key]
        else:
            async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                sched = await client.get(f"{NHL_API_BASE}/club-schedule-season/{team_abbr}/{season}")
            if sched.status_code != 200:
                return None
            games = sched.json().get("games", [])
            _nhl_schedule_cache[sched_key] = games
        match = next((g for g in games if g.get("gameDate") == game_date), None)
        if not match:
            return None
        nhl_game_id = match["id"]
        if nhl_game_id in _nhl_boxscore_cache:
            data = _nhl_boxscore_cache[nhl_game_id]
        else:
            async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                box = await client.get(f"{NHL_API_BASE}/gamecenter/{nhl_game_id}/boxscore")
            if box.status_code != 200:
                return None
            data = box.json()
            _nhl_boxscore_cache[nhl_game_id] = data
        is_home = data.get("homeTeam", {}).get("abbrev") == team_abbr.upper()
        return data.get("playerByGameStats", {}).get("homeTeam" if is_home else "awayTeam")
    except (httpx.RequestError, KeyError, TypeError, ValueError):
        return None


def _pp_units(conn, game_id: int, team_abbr: str, players: dict | None = None):
    """Full 5-skater (forwards + D combined) on-ice groupings while
    team_abbr is on the power play — kept as ONE group per combo rather
    than split by position, since PP units mix forwards/D freely (unlike
    5v5 lines/pairs, which stay position-pure). Sorted by icetime: the
    top group is PP1, the next is PP2. In games with limited PP time,
    PP2 (or beyond) can rest on a very small sample — treat a low
    icetimeSeconds value on a group as tentative, not a confident call
    (confirmed in practice: one EDM game had two different PP2-ish
    groupings at only 13-16 seconds each, not enough to trust)."""
    team_full = full_name(team_abbr)
    home_away = conn.execute(
        "SELECT homeTeamAbbrev, awayTeamAbbrev FROM games WHERE gameReferenceId = ?", (game_id,)
    ).fetchone()
    if not home_away:
        raise HTTPException(404, f"Game {game_id} not found in the games table")
    other_abbr = home_away["awayTeamAbbrev"] if home_away["homeTeamAbbrev"] == team_abbr.upper() else home_away["homeTeamAbbrev"]
    other_full = full_name(other_abbr)

    rows = conn.execute(
        "SELECT gameTime, team, manpowerSituation, teamSkatersOnIce, opposingTeamSkatersOnIce, "
        "teamForwardsOnIceRefs, teamDefencemenOnIceRefs, "
        "opposingTeamForwardsOnIceRefs, opposingTeamDefencemenOnIceRefs "
        "FROM plays WHERE gameReferenceId = ? ORDER BY gameTime",
        (game_id,),
    ).fetchall()

    unit_time = defaultdict(float)
    n = len(rows)
    CAP = 10.0
    for i in range(n - 1):
        r, nxt = rows[i], rows[i + 1]
        dt = nxt["gameTime"] - r["gameTime"]
        if dt <= 0 or dt > CAP:
            continue
        if r["manpowerSituation"] != "powerPlay":
            continue
        team = r["team"]
        if team == team_full:
            fwd, dmen = _clean_refs(r["teamForwardsOnIceRefs"]), _clean_refs(r["teamDefencemenOnIceRefs"])
            skaters, opp = r["teamSkatersOnIce"], r["opposingTeamSkatersOnIce"]
        elif team == other_full:
            fwd, dmen = _clean_refs(r["opposingTeamForwardsOnIceRefs"]), _clean_refs(r["opposingTeamDefencemenOnIceRefs"])
            skaters, opp = r["opposingTeamSkatersOnIce"], r["teamSkatersOnIce"]
        else:
            continue
        if skaters != 5 or opp != 4:  # team_abbr specifically on the man advantage
            continue
        group = tuple(sorted(fwd + dmen))
        if group:
            unit_time[group] += dt

    if players is None:
        players = _player_names_for_team(conn, team_full)
    return [
        {"players": [_name(players, p) for p in combo], "icetimeSeconds": round(t, 1)}
        for combo, t in sorted(unit_time.items(), key=lambda x: -x[1])
    ]


@app.get("/api/game/{game_id}/lineup")
async def game_lineup(game_id: int, team: str = Query(..., description="Team abbreviation, e.g. EDM"), user: CurrentUser = Depends(require_login)):
    """Assembled game-recap lineup for one team: the 4 forward lines and
    3 D pairs the team STARTED the game with (each line/pair with a
    best-guess center + LD/RD), starting/backup goalie, and PP1/PP2 as
    full 5-skater groups.

    'Started with' is a deliberate choice, not an approximation of
    'played the most together': lines and pairs change over 60 minutes
    (injury, matchups, a bad period), but this feature only cares about
    the opening alignment. Concretely, lines/pairs are ranked by the
    EARLIEST gameTime each combo reaches a genuine, sustained shift (see
    _first_shift_times), not by total icetime accumulated over the whole
    game — a line that started the game but got broken up in the third
    period still correctly shows up as a starting line here.

    Handles irregular roster nights (e.g. 11 forwards / 7 defensemen
    instead of the usual 12/6) explicitly rather than just papering over
    them with whatever the raw combos happen to produce:
      - Forwards: if a skater has a genuine early shift with two
        different trios (an extra forward with no fixed line), he's
        kept only in the earlier-starting one — the redundant slot in
        the other line is left blank (None) rather than showing his
        name twice. There should never be a duplicate player across the
        returned lines/pairs.
      - Defensemen: pairs are built to be guaranteed non-overlapping
        (see _assign_stable_groups_from_sorted), so a 7th defenseman who
        never had one consistent starting partner falls out naturally
        as 'extraDefensemen' — listed separately below the (up to) 3
        clean pairs, rather than forced into a misleading pairing.

    Self-checks against the NHL's own boxscore where possible (see
    _nhl_boxscore_stats_side): fills in a backup goalie who dressed but
    never played (PSF has zero record of a player who never touches the
    ice), and flags a 'defensemenCountWarning' if our derived pair/extra
    count doesn't match how many defensemen the NHL boxscore lists as
    dressed for this game — a real signal something was missed or
    mis-clustered, not just a cosmetic mismatch.

    Center is picked by combining two signals (see _faceoff_centers):
    faceoffs taken this game, and the NHL roster's own official
    position code — confirmed necessary after EDM @ VGK 2026-03-08,
    where faceoff count alone wrongly picked Nugent-Hopkins over McDavid
    (an every-night center) as the trio's center.

    Deliberately NOT resolved here: LW vs RW side for forwards. PSF's
    per-event playerPosition tag is unreliable for this, and coordinate-
    based inference — which does work for D pairs — was tested for
    forward wing side and came back at-or-below chance (see
    Coordinate_Adjustment_Wing_Side_Reference.md). The frontend resolves
    LW/RW from each player's actual roster position instead, not from
    anything derived here.

    Jersey numbers come from the players table (players.jersey, sourced
    from PSF's own playerJersey column) — already loaded, no NHL API
    round-trip required for this specifically.

    Lines and D pairs are additionally numbered 1st/2nd/3rd/4th by
    depth-chart role, not by which combo started earliest: the line/pair
    containing this SEASON's single highest-point player is #1, then
    whichever line/pair has the highest point total among players not
    already claimed by an earlier line/pair, and so on (see
    _rank_units_by_top_scorer) — this only reorders which position in
    the response counts as "1st"/"2nd"/etc.; it doesn't change who's
    actually grouped together, which is still the earliest-sustained-
    shift logic described above."""
    team_abbr = team.upper()
    data = await asyncio.to_thread(_build_lineup_sync_data, game_id, team_abbr)

    async def _fetch_stats_side():
        if data["game_date"] and data["game_season"]:
            return await _nhl_boxscore_stats_side(team_abbr, data["game_date"], data["game_season"])
        return None

    async def _fetch_points_map():
        if data["game_season"]:
            return await _fetch_season_points_map(team_abbr, data["game_season"])
        return {}

    # These three NHL API calls are independent of each other (roster,
    # schedule/boxscore, season stats) — running them concurrently
    # instead of one-after-another keeps the added points-ranking lookup
    # from adding its own extra round-trip's worth of wait.
    position_map, stats_side, points_map = await asyncio.gather(
        _fetch_nhl_position_map(team_abbr, data["game_season"]),
        _fetch_stats_side(),
        _fetch_points_map(),
    )
    centers = _pick_centers(data["trios"], data["faceoff_counts"], position_map)

    players = data["players"]
    trios = data["trios"]
    extra_fwd_ids = data["extra_fwd_ids"]
    pairs = data["pairs"]
    extra_d_ids = data["extra_d_ids"]
    fwd_xg = data["fwd_xg"]
    dmen_xg = data["dmen_xg"]
    fwd_total = data["fwd_total"]
    dmen_total = data["dmen_total"]
    dpair_sides = data["dpair_sides"]
    starter_id = data["starter_id"]
    backup_id = data["backup_id"]
    pp_groups = data["pp_groups"]
    pk_forward_pairs = data["pk_forward_pairs"]
    pk_d_pairs = data["pk_d_pairs"]

    # Reorder lines/pairs (and their parallel centers/dpair_sides arrays,
    # index-for-index) by top scorer — see docstring above. Falls back to
    # the original earliest-shift order harmlessly if points_map is empty
    # (e.g. season stats fetch failed) since every unit's best_points is
    # then 0 and the sort is stable.
    fwd_rank = _rank_units_by_top_scorer(trios, points_map)
    trios = [trios[i] for i in fwd_rank]
    centers = [centers[i] for i in fwd_rank]

    d_rank = _rank_units_by_top_scorer(pairs, points_map)
    pairs = [pairs[i] for i in d_rank]
    dpair_sides = [dpair_sides[i] for i in d_rank]

    def line_out(trio, center_id, i):
        wings = [p for p in trio if p != center_id]
        return {
            "line": i + 1,
            "center": _name(players, int(center_id)) if center_id else None,
            "wings": [_name(players, int(p)) if p is not None else None for p in wings],
            "esXgPct": _combo_xg_pct(fwd_xg, trio),
            "esToiSeconds": _combo_toi_seconds(fwd_total, trio),
        }

    def pair_out(sides, i):
        return {
            "pair": i + 1,
            "LD": _name(players, int(sides["LD"])) if sides.get("LD") else None,
            "RD": _name(players, int(sides["RD"])) if sides.get("RD") else None,
            "esXgPct": _combo_xg_pct(dmen_xg, pairs[i]),
            "esToiSeconds": _combo_toi_seconds(dmen_total, pairs[i]),
        }

    def _goalie_from_stats_side(is_starter: bool):
        """NHL boxscore fallback for a goalie's name/jersey — goalies
        almost never appear as the 'acting player' on a play row (they're
        tracked via on-ice reference columns instead, not as someone
        passing/shooting/entering), so plays.playerJersey structurally
        can't be relied on for them the way it can for a skater. Confirmed
        directly: Demko's name resolved fine but his jersey came back
        blank — the PSF-derived players lookup apparently had a name for
        him from some rare acting-player row but no jersey on it."""
        if not stats_side:
            return None
        g = next((g for g in stats_side.get("goalies", []) if g.get("starter") is is_starter), None)
        if not g:
            return None
        return {
            "id": g.get("playerId"),
            "name": g.get("name", {}).get("default"),
            "jersey": str(g.get("sweaterNumber", "")),
            "position": "G",
        }

    starter_out = _name(players, int(starter_id)) if starter_id else None
    if starter_out and not starter_out.get("jersey"):
        fb = _goalie_from_stats_side(True)
        if fb:
            starter_out["jersey"] = fb["jersey"]
            if not starter_out.get("name"):
                starter_out["name"] = fb["name"]
    elif starter_out is None and stats_side:
        starter_out = _goalie_from_stats_side(True)

    backup_out = _name(players, int(backup_id)) if backup_id else None
    if backup_out and not backup_out.get("jersey"):
        fb = _goalie_from_stats_side(False)
        if fb:
            backup_out["jersey"] = fb["jersey"]
            if not backup_out.get("name"):
                backup_out["name"] = fb["name"]
    elif backup_out is None and stats_side:
        backup_out = _goalie_from_stats_side(False)

    forwards_count_warning = None
    defensemen_count_warning = None
    if stats_side:
        nhl_fwd_count = len(stats_side.get("forwards", []))
        our_fwd_count = len(trios) * 3 + len(extra_fwd_ids)
        if nhl_fwd_count != our_fwd_count:
            forwards_count_warning = (
                f"NHL boxscore lists {nhl_fwd_count} forwards dressed for this game; "
                f"this derivation accounted for {our_fwd_count}. Worth double-checking for "
                f"a missed or mis-clustered line."
            )
        nhl_d_count = len(stats_side.get("defense", []))
        our_d_count = len(pairs) * 2 + len(extra_d_ids)
        if nhl_d_count != our_d_count:
            defensemen_count_warning = (
                f"NHL boxscore lists {nhl_d_count} defensemen dressed for this game; "
                f"this derivation accounted for {our_d_count}. Worth double-checking for "
                f"a missed or mis-clustered pairing."
            )

    return {
        "gameId": game_id,
        "team": team_abbr,
        "forwardLines": [line_out(t, centers[i], i) for i, t in enumerate(trios)],
        "extraForwards": [_name(players, int(p)) for p in extra_fwd_ids],
        "dPairs": [pair_out(dpair_sides[i], i) for i in range(len(pairs))],
        "extraDefensemen": [_name(players, int(p)) for p in extra_d_ids],
        "goalies": {
            "starter": starter_out,
            "backup": backup_out,
        },
        "ppUnits": {
            "PP1": pp_groups[0] if len(pp_groups) > 0 else None,
            "PP2": pp_groups[1] if len(pp_groups) > 1 else None,
        },
        "pkUnits": {
            "forwardPairs": pk_forward_pairs,
            "dPairs": pk_d_pairs,
        },
        "forwardsCountWarning": forwards_count_warning,
        "defensemenCountWarning": defensemen_count_warning,
    }


def _build_lineup_sync_data(game_id: int, team_abbr: str) -> dict:
    """All of game_lineup's blocking DB/CPU work, self-contained with its
    own connection so it can run via asyncio.to_thread instead of on the
    event loop. This was previously done directly inside the async route
    — while it ran, it blocked EVERY other pending request on the
    server (single uvicorn worker, no threads for async routes), not
    just this one. Confirmed live: multiple unrelated requests (other
    games' lineups, roster fetches) queuing up for 50+ seconds behind
    one of these calls in the exact same way the goalie-gsax/skater-woi
    functions did before that fix — same bug class, just in an older
    part of the codebase that hadn't been audited yet.

    Only the two genuine NHL API calls (_fetch_nhl_position_map,
    _nhl_boxscore_stats_side) stay as real awaits in game_lineup itself;
    everything else — including the faceoff-count query used for center
    selection — happens in here."""
    conn = get_db()
    try:
        game_or_404(conn, game_id)
        team_full = full_name(team_abbr)
        players = _player_names_for_team(conn, team_full)

        fwd_first, dmen_first = _first_shift_times(conn, game_id, team_abbr, "5v5")
        fwd_total, dmen_total = _manpower_combo_seconds(conn, game_id, team_abbr, "5v5")
        fwd_xg, dmen_xg = _manpower_combo_xg(conn, game_id, team_abbr, "5v5")

        # A combo's first-ever recorded appearance can occasionally be a
        # near-meaningless blip — a brief overlap during a change, or a
        # data-granularity artifact — that happens to have an earlier
        # timestamp than a real, heavily-used line's own first appearance,
        # purely by chance. Confirmed on a real game (EDM @ VGK 2026-03-08):
        # without this filter, McDavid's actual trio (530+ total seconds,
        # the #2 most-used forward combo all game) got crowded out of the
        # top 4 by combos that individually amounted to almost nothing.
        # Requiring a combo to have logged at least one real shift's worth
        # of total icetime before it's even eligible to be ranked by "first
        # appearance" keeps the starting-lineup logic robust to that noise
        # without reintroducing the opposite failure (the old continuous-
        # shift requirement, which discarded real lines outright — see
        # _first_shift_times' docstring).
        MIN_STARTING_COMBO_SECONDS = 30.0

        # Forwards: among combos with real, substantial 5v5 icetime, build
        # guaranteed-clean (non-overlapping) trios. Conflicts between
        # overlapping candidates (two combos sharing 1-2 players) are
        # resolved by ICETIME, not first-appearance — confirmed necessary on
        # a real game (VGK, same 2026-03-08 game as the McDavid fix): a
        # heavily-used real trio (Eichel/Barbashev/Bowman, 590 total
        # seconds) got split across two separate "lines" when an earlier
        # version resolved conflicts by first-appearance instead, because a
        # noisy 30-second shift-variant combo (sharing 2 of those 3 players
        # with a different third player) happened to have an EARLIER first
        # appearance than the real trio, so it won the conflict and locked
        # out the real 600-second trio entirely. Icetime doesn't have that
        # failure mode — a real line will always dwarf a noise blip in total
        # minutes, so greedily accepting the highest-icetime candidate first
        # can't let noise steal a player away from a real line. This does
        # NOT reintroduce "most-used over the whole game" as the ranking
        # concept, though: two genuinely different, non-overlapping trios
        # (no shared players) both get accepted regardless of which has
        # more icetime, since there's no conflict to resolve between them —
        # icetime only matters as the tiebreaker when trios actually
        # compete for the same player. Once the clean, correctly-grouped
        # trios are known, THEY are re-ordered by first-appearance
        # (ascending) purely for "line 1 / line 2 / etc" display purposes,
        # which is what actually captures "what did the team start with."
        # Anyone left over (an 11th forward on a genuinely irregular-roster
        # night, not just noisy variants of a real trio) is surfaced
        # separately as 'extraForwards', same treatment as extraDefensemen.
        fwd_candidates = [
            (combo, t) for combo, t in fwd_total.items()
            if len(combo) == 3 and t >= MIN_STARTING_COMBO_SECONDS
        ]
        fwd_by_icetime = sorted(fwd_candidates, key=lambda x: -x[1])
        trios, extra_fwd_ids = _assign_stable_groups_from_sorted(fwd_by_icetime, group_size=3, max_groups=4)
        trios.sort(key=lambda trio: fwd_first.get(tuple(sorted(trio)), float("inf")))

        # Defensemen: identical two-stage approach — icetime resolves
        # conflicts between overlapping pair candidates, first-appearance
        # only orders the resulting clean pairs for display. Anyone left
        # over (a 7th D on an irregular night) is surfaced separately.
        dmen_candidates = [
            (combo, t) for combo, t in dmen_total.items()
            if len(combo) == 2 and t >= MIN_STARTING_COMBO_SECONDS
        ]
        dmen_by_icetime = sorted(dmen_candidates, key=lambda x: -x[1])
        pairs, extra_d_ids = _assign_stable_groups_from_sorted(dmen_by_icetime, group_size=2, max_groups=3)
        pairs.sort(key=lambda pair: dmen_first.get(tuple(sorted(pair)), float("inf")))

        faceoff_counts = _faceoff_counts_for_game(conn, game_id, team_full)
        dpair_sides = _dpair_sides(conn, game_id, team_full, pairs)

        starter_id, backup_id = _starting_backup_goalie(conn, game_id, team_full)
        pp_groups = _pp_units(conn, game_id, team_abbr, players=players)
        # PK forward pairs / D pairs: still ranked by total icetime (who
        # plays the PK the most), NOT starting-shift order — "PK1/PK2" is
        # conventionally about usage, unlike the 5v5 lines/pairs above.
        pk = compute_units(conn, game_id, team_abbr, "4v5", players=players)
        pk_forward_pairs = pk["forwardUnits"][:3]
        pk_d_pairs = pk["defenseUnits"][:3]

        game_row = conn.execute(
            "SELECT gameDate, season FROM games WHERE gameReferenceId = ?", (game_id,)
        ).fetchone()

        return {
            "players": players,
            "trios": trios,
            "extra_fwd_ids": extra_fwd_ids,
            "pairs": pairs,
            "extra_d_ids": extra_d_ids,
            "fwd_xg": fwd_xg,
            "dmen_xg": dmen_xg,
            "fwd_total": fwd_total,
            "dmen_total": dmen_total,
            "faceoff_counts": faceoff_counts,
            "dpair_sides": dpair_sides,
            "starter_id": starter_id,
            "backup_id": backup_id,
            "pp_groups": pp_groups,
            "pk_forward_pairs": pk_forward_pairs,
            "pk_d_pairs": pk_d_pairs,
            "game_date": game_row["gameDate"] if game_row else None,
            "game_season": game_row["season"] if game_row else None,
        }
    finally:
        conn.close()


# ── Goalie GSAx (Goals Saved Above Expected) ─────────────────────────────
# Every shot event in nhl.db already carries expectedGoalsOnNet
# (SportLogiq's own per-shot xG model) — GSAx is just:
#
#     GSAx = SUM(expectedGoalsOnNet) for shots that goalie faced
#            - actual goals allowed while they were in net
#
# No NHL API round-trip needed; this is pure PSF-already-in-nhl.db math,
# same spirit as _starting_backup_goalie above.
#
# Split into a plain function (_goalie_gsax_data) the route just wraps,
# so the prep-cache builder below can call it directly without an HTTP
# round-trip to its own server.

def _goalie_gsax_data(conn, team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> dict:
    """Core GSAx computation — season XOR dates, exactly one required."""
    team_full = full_name(team_abbr)
    players = _player_names_for_team(conn, team_full)

    game_q = "SELECT gameReferenceId FROM games WHERE (homeTeamAbbrev = ? OR awayTeamAbbrev = ?)"
    game_params = [team_abbr, team_abbr]
    if dates:
        placeholders = ",".join("?" for _ in dates)
        game_q += f" AND gameDate IN ({placeholders})"
        game_params += dates
    else:
        game_q += " AND season = ?"
        game_params.append(season)

    game_ids = [r["gameReferenceId"] for r in conn.execute(game_q, game_params).fetchall()]
    if not game_ids:
        return {}

    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""
        SELECT opposingTeamGoalieOnIceRef AS goalieId,
               SUM(CASE WHEN name = 'shot' AND type NOT IN ('slotblocked', 'outsideblocked')
                        THEN COALESCE(expectedGoalsOnNet, 0) ELSE 0 END) AS xga,
               SUM(CASE WHEN name = 'goal' THEN 1 ELSE 0 END) AS ga
        FROM plays
        WHERE gameReferenceId IN ({placeholders})
          AND team != ?
          AND period <= 3
          AND opposingTeamGoalieOnIceRef IS NOT NULL
        GROUP BY opposingTeamGoalieOnIceRef
        """,
        game_ids + [team_full],
    ).fetchall()

    result = {}
    for r in rows:
        goalie_id = r["goalieId"]
        xga, ga = r["xga"] or 0.0, r["ga"] or 0
        p = players.get(int(goalie_id)) if goalie_id else None
        result[str(goalie_id)] = {
            "xga": round(xga, 2),
            "ga": ga,
            "gsax": round(xga - ga, 2),
            "name": f"{p['firstName']} {p['lastName']}" if p else None,
            "jersey": p["jersey"] if p else None,
        }
    return result


@app.get("/api/goalie-gsax/{team}")
def goalie_gsax(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026 — every game nhl.db has for this team in that season"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD) to scope to specific games, e.g. a last-5 cut"),
    user: CurrentUser = Depends(require_login),
):
    """Goals Saved Above Expected, per goalie, for one team — scoped to
    either a whole season (season=) or a specific set of games (dates=);
    exactly one is required.

    dates is used instead of a list of game IDs because PSF's own
    gameReferenceId numbering doesn't match the NHL's — the same problem
    _nhl_boxscore_stats_side already solves by matching on gameDate
    instead of ID, so this follows the same convention rather than
    introducing a second one.

    Returns {} if no games match the given season/dates for this team
    (not a 404 — an opponent simply not in nhl.db yet is a normal,
    expected case here, not an error)."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        return _goalie_gsax_data(conn, team.upper(), season=season, dates=date_list)
    finally:
        conn.close()


# ── ES xG% WOI (while on ice) ─────────────────────────────────────────────
# For each skater: their team's even-strength xGF and xGA specifically
# during the shifts they were on the ice for, from nhl.db's on_ice table
# (one row per event x on-ice player, teamSide 'for'/'against' relative to
# the event's acting team) joined back to plays for expectedGoalsOnNet.
# A shot tagged 'for' for player P means P's team took that shot (counts
# toward P's on-ice xGF); 'against' means the opponent took it while P was
# out there (counts toward P's on-ice xGA) — this falls out of teamSide's
# existing definition with no extra team-matching needed.

def _skater_woi_data(conn, team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> dict:
    """Core ES xG% WOI computation — season XOR dates, exactly one required."""
    game_q = "SELECT gameReferenceId FROM games WHERE (homeTeamAbbrev = ? OR awayTeamAbbrev = ?)"
    game_params = [team_abbr, team_abbr]
    if dates:
        placeholders = ",".join("?" for _ in dates)
        game_q += f" AND gameDate IN ({placeholders})"
        game_params += dates
    else:
        game_q += " AND season = ?"
        game_params.append(season)

    game_ids = [r["gameReferenceId"] for r in conn.execute(game_q, game_params).fetchall()]
    if not game_ids:
        return {}

    placeholders = ",".join("?" for _ in game_ids)
    rows = conn.execute(
        f"""
        SELECT oi.playerReferenceId AS playerId,
               SUM(CASE WHEN oi.teamSide = 'for' THEN COALESCE(p.expectedGoalsOnNet, 0) ELSE 0 END) AS xgf,
               SUM(CASE WHEN oi.teamSide = 'against' THEN COALESCE(p.expectedGoalsOnNet, 0) ELSE 0 END) AS xga
        FROM on_ice oi
        JOIN plays p ON p.gameReferenceId = oi.gameReferenceId AND p.id = oi.id
        WHERE oi.gameReferenceId IN ({placeholders})
          AND p.name = 'shot'
          AND p.type NOT IN ('slotblocked', 'outsideblocked')
          AND p.manpowerSituation = 'evenStrength'
          AND p.period <= 3
        GROUP BY oi.playerReferenceId
        """,
        game_ids,
    ).fetchall()

    result = {}
    for r in rows:
        xgf, xga = r["xgf"] or 0.0, r["xga"] or 0.0
        result[str(r["playerId"])] = {
            "xgf": round(xgf, 3),
            "xga": round(xga, 3),
            "xgPct": round(xgf / (xgf + xga) * 100, 1) if (xgf + xga) > 0 else None,
        }
    return result


@app.get("/api/skater-woi/{team}")
def skater_woi(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD), e.g. a last-5 cut"),
    user: CurrentUser = Depends(require_login),
):
    """ES xG% while on ice, per skater, for one team — scoped to either a
    whole season (season=) or a specific set of games (dates=); exactly
    one required. dates instead of game IDs for the same reason as
    goalie_gsax above: PSF's gameReferenceId doesn't match the NHL's own
    numbering.

    Returns {} if no games match (not a 404 — a team with no PSF data yet
    is a normal, expected case here)."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        return _skater_woi_data(conn, team.upper(), season=season, dates=date_list)
    finally:
        conn.close()


# _goalie_gsax_data/_skater_woi_data are plain synchronous SQLite calls —
# fine as-is inside the two `def` routes above (FastAPI runs sync routes
# in a thread pool automatically), but calling them directly from inside
# an `async def` (as the prep-cache builder below does) would run the
# query on the main event loop and block EVERY other request — including
# unrelated ones like login — for however long that query takes against
# an 8+ GB database. These two wrappers open their own connection and run
# in a worker thread via asyncio.to_thread so the event loop stays free.
def _goalie_gsax_threaded(team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> dict:
    conn = get_db()
    try:
        return _goalie_gsax_data(conn, team_abbr, season=season, dates=dates)
    finally:
        conn.close()


def _skater_woi_threaded(team_abbr: str, season: str | None = None, dates: list[str] | None = None) -> dict:
    conn = get_db()
    try:
        return _skater_woi_data(conn, team_abbr, season=season, dates=dates)
    finally:
        conn.close()


# ── Pre-Scout Prep cache (EDM's next game, precomputed) ──────────────────
# Pre-Scout Prep is always EDM vs. whoever's next — there's no reason to
# make a coach wait on ~15 NHL API round-trips every time they open it.
# This computes the whole payload server-side, once, and keeps it warm in
# memory; the frontend just asks for whatever's cached and gets it back
# instantly. Ported from the equivalent (now-removed) client-side fetch
# logic in index.html — same field names, same NHL API endpoints, same
# GSAx/WOI merge logic, just running here instead of in every coach's
# browser on every page load.
#
# "Rollover" needs no special bookkeeping: _find_next_edm_game always asks
# the live NHL schedule for the earliest game still in gameState 'FUT'.
# Once a game actually starts, its own gameState flips away from FUT, so
# the very next query for "what's next" naturally returns the following
# game — the rollover is just a property of always asking fresh, not
# something to track ourselves.

def _nhl_schedule_season_str(today: dt.date | None = None) -> str:
    """'20262027' for a date in Jul(inclusive)-Jun; Jan-Jun still belongs
    to the season that started the previous fall."""
    today = today or dt.date.today()
    start_year = today.year if today.month >= 7 else today.year - 1
    return f"{start_year}{start_year + 1}"


def _nhl_player_display_name(p: dict) -> str:
    def _default(v):
        if v is None:
            return ""
        return v.get("default", "") if isinstance(v, dict) else str(v)
    if p.get("name"):
        return _default(p["name"])
    return f"{_default(p.get('firstName'))} {_default(p.get('lastName'))}".strip()


def _parse_toi_to_seconds(toi) -> int:
    if toi is None:
        return 0
    if isinstance(toi, (int, float)):
        return int(toi)
    parts = str(toi).split(":")
    if len(parts) == 2:
        try:
            return int(parts[0]) * 60 + int(parts[1])
        except ValueError:
            return 0
    return 0


def _parse_save_shots_against(s) -> tuple[int, int]:
    if not s or "/" not in str(s):
        return 0, 0
    try:
        saves, shots_against = str(s).split("/")
        return int(saves), int(shots_against)
    except ValueError:
        return 0, 0


async def _find_next_n_edm_games(client: httpx.AsyncClient, n: int = 1) -> list[dict]:
    """The next N upcoming EDM games (regular season + playoffs, preseason
    excluded), earliest first. Same single NHL API call regardless of N —
    this is the schedule-only lookup, not the heavy per-opponent stats
    build (see _build_prep_payload for that)."""
    season = _nhl_schedule_season_str()
    r = await client.get(f"{NHL_API_BASE}/club-schedule-season/EDM/{season}")
    r.raise_for_status()
    games = r.json().get("games", [])
    reg_and_po = [g for g in games if g.get("gameType") != 1]  # exclude preseason only
    upcoming = sorted((g for g in reg_and_po if g.get("gameState") == "FUT"), key=lambda g: g["gameDate"])
    all_sorted = sorted(reg_and_po, key=lambda g: g["gameDate"])
    game_number_by_id = {g["id"]: i + 1 for i, g in enumerate(all_sorted)}
    total_games = len(all_sorted)

    result = []
    for target in upcoming[:n]:
        is_home = target["homeTeam"]["abbrev"] == "EDM"
        opponent = target["awayTeam"]["abbrev"] if is_home else target["homeTeam"]["abbrev"]
        result.append({
            "gameId": target["id"],
            "gameDate": target["gameDate"],
            "startTimeUTC": target.get("startTimeUTC"),
            "opponent": opponent,
            "isHome": is_home,
            "season": season,
            "gameNumber": game_number_by_id.get(target["id"]),
            "totalGames": total_games,
        })
    return result


async def _find_next_edm_game(client: httpx.AsyncClient) -> dict | None:
    games = await _find_next_n_edm_games(client, n=1)
    return games[0] if games else None


async def _fetch_jersey_map(client: httpx.AsyncClient, team_abbr: str, season: str | None = None) -> dict:
    """club-stats has no jersey numbers at all (confirmed live) — only the
    roster endpoint does. Pass `season` when available — see
    _fetch_nhl_position_map's docstring for why /current alone is
    unreliable outside the season being reported on (empty/unsettled
    during the offseason in particular)."""
    endpoint = f"{team_abbr}/{season}" if season else f"{team_abbr}/current"
    try:
        r = await client.get(f"{NHL_API_BASE}/roster/{endpoint}")
        if r.status_code != 200:
            return {}
        data = r.json()
        return {
            p["id"]: p.get("sweaterNumber")
            for p in (data.get("forwards", []) + data.get("defensemen", []) + data.get("goalies", []))
        }
    except httpx.RequestError:
        return {}


async def _fetch_pp_points_map(client: httpx.AsyncClient, team_abbr: str, season: str) -> dict:
    """club-stats has powerPlayGoals but no PP points/assists at all
    (confirmed live) — real PP points live on the NHL's other stats host,
    api.nhle.com (not api-web.nhle.com), hence a separate request base."""
    try:
        r = await client.get(
            f"{NHL_STATS_API_BASE}/stats/rest/en/skater/summary",
            params={
                "isAggregate": "false", "isGame": "false",
                "sort": '[{"property":"points","direction":"DESC"}]',
                "start": "0", "limit": "100",
                "cayenneExp": f'gameTypeId=2 and seasonId={season} and teamAbbrevs="{team_abbr}"',
            },
        )
        if r.status_code != 200:
            return {}
        return {row["playerId"]: row.get("ppPoints", 0) for row in r.json().get("data", [])}
    except httpx.RequestError:
        return {}


_IS_FORWARD_POS = {"C", "L", "R", "LW", "RW", "F"}


def _top_by_stat(players: list, position_check, stat_key: str, count: int) -> list:
    matching = [p for p in players if position_check(p.get("position"))]
    matching.sort(key=lambda p: (-(p.get(stat_key) or 0), -(p.get("g") or 0)))
    return matching[:count]


_season_points_cache: dict = {}


async def _fetch_season_points_map(team_abbr: str, season: str) -> dict:
    """Season point totals (playerId -> points), used to rank lineup
    lines/pairs by who their most productive player is. Deliberately its
    own lightweight fetch rather than reusing _fetch_season_scorer_data
    below, which also pulls jersey/PP/WOI/GSAx data this doesn't need —
    keeping this to a single cached call matters since it now runs on
    every /lineup request, right after the lineup endpoint's own
    round-trip count was just trimmed down."""
    key = (team_abbr, season)
    if key in _season_points_cache:
        return _season_points_cache[key]
    try:
        async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
            r = await client.get(f"{NHL_API_BASE}/club-stats/{team_abbr}/{season}/2")
        if r.status_code != 200:
            return {}
        data = r.json()
    except httpx.RequestError:
        return {}
    points_map = {}
    for s in data.get("skaters", []):
        pid = s.get("playerId")
        if pid is not None:
            points_map[pid] = s.get("points", (s.get("goals", 0) + s.get("assists", 0)))
    _season_points_cache[key] = points_map
    return points_map


def _rank_units_by_top_scorer(units: list, points_map: dict) -> list:
    """Returns the original indices of `units` (each a tuple/list of
    player-id strings, possibly with a None for a blanked irregular-
    roster slot) sorted so the unit containing the single highest
    season-point player comes first, then the unit with the highest
    point total among whoever's left, and so on.

    This is provably the same result as literally walking players in
    points order and assigning ranks as their unit first comes up
    unassigned: the next-highest remaining player can only belong to a
    not-yet-ranked unit (an already-ranked one's players are already
    accounted for), so sorting units by their own single highest-point
    member produces an identical ordering without the extra bookkeeping.
    Ties keep their original relative order (Python's sort is stable)."""
    def best_points(unit):
        pids = [p for p in unit if p is not None]
        return max((points_map.get(int(p), 0) for p in pids), default=0)
    return sorted(range(len(units)), key=lambda i: -best_points(units[i]))


async def _fetch_season_scorer_data(client: httpx.AsyncClient, team_abbr: str, season: str) -> dict:
    empty = {"skaters": [], "forwards": [], "defense": [], "ppForwards": [], "ppDefense": [], "goalies": [], "nameMap": {}}
    stats_res, jersey_map, pp_map, woi_map, gsax_map = await asyncio.gather(
        client.get(f"{NHL_API_BASE}/club-stats/{team_abbr}/{season}/2"),
        _fetch_jersey_map(client, team_abbr, season),
        _fetch_pp_points_map(client, team_abbr, season),
        asyncio.to_thread(_skater_woi_threaded, team_abbr, season),
        asyncio.to_thread(_goalie_gsax_threaded, team_abbr, season),
    )
    if stats_res.status_code != 200:
        return empty
    data = stats_res.json()

    skaters = []
    for s in data.get("skaters", []):
        pid = s["playerId"]
        skaters.append({
            "playerId": pid,
            "name": _nhl_player_display_name(s),
            "jersey": jersey_map.get(pid),
            "position": s.get("positionCode"),
            "gp": s.get("gamesPlayed", 0),
            "g": s.get("goals", 0),
            "a": s.get("assists", 0),
            "points": s.get("points", (s.get("goals", 0) + s.get("assists", 0))),
            "plusMinus": s.get("plusMinus", 0),
            "pim": s.get("penaltyMinutes", 0),
            "toiSec": s.get("avgTimeOnIcePerGame"),
            "ppGoals": s.get("powerPlayGoals", 0),
            "ppPoints": pp_map.get(pid, 0),
            "esXgPctWoi": (woi_map.get(str(pid)) or {}).get("xgPct"),
            # In the season stats but not on the current roster fetch —
            # traded/waived/etc. since. Surfaced in the UI as a small red
            # X next to their name rather than silently showing stale
            # numbers for someone no longer on the team.
            "onRoster": pid in jersey_map,
        })

    goalies = []
    for g in data.get("goalies", []):
        pid = g["playerId"]
        goalies.append({
            "playerId": pid,
            "name": _nhl_player_display_name(g),
            "jersey": jersey_map.get(pid),
            "gp": g.get("gamesPlayed", 0),
            "gs": g.get("gamesStarted", g.get("gamesPlayed", 0)),
            "gaa": g.get("goalsAgainstAverage"),
            "svPct": g.get("savePercentage"),
            "gsax": (gsax_map.get(str(pid)) or {}).get("gsax"),
            "onRoster": pid in jersey_map,
        })
    goalies.sort(key=lambda g: g["gs"], reverse=True)
    goalies = goalies[:3]

    name_map = {p["playerId"]: {"name": p["name"], "jersey": p["jersey"], "onRoster": p["onRoster"]} for p in skaters + goalies}
    is_fwd = lambda pos: pos in _IS_FORWARD_POS
    is_def = lambda pos: pos == "D"

    return {
        "skaters": skaters,
        "forwards": _top_by_stat(skaters, is_fwd, "points", 6),
        "defense": _top_by_stat(skaters, is_def, "points", 4),
        "ppForwards": _top_by_stat(skaters, is_fwd, "ppPoints", 6),
        "ppDefense": _top_by_stat(skaters, is_def, "ppPoints", 4),
        "goalies": goalies,
        "nameMap": name_map,
    }


async def _fetch_last5_scorer_data(client: httpx.AsyncClient, team_abbr: str, season: str) -> dict:
    empty = {"forwards": [], "defense": [], "goalies": [], "discoveredJerseys": {}}
    sched_res = await client.get(f"{NHL_API_BASE}/club-schedule-season/{team_abbr}/{season}")
    if sched_res.status_code != 200:
        return empty
    games = sched_res.json().get("games", [])
    today_str = dt.date.today().isoformat()
    played = sorted(
        (g for g in games if g.get("gameType") != 1 and g.get("gameDate", "") < today_str and g.get("gameState") in ("OFF", "FINAL")),
        key=lambda g: g["gameDate"], reverse=True,
    )[:5]
    game_dates = [g["gameDate"] for g in played]

    skater_totals: dict = {}
    goalie_totals: dict = {}
    discovered_jerseys: dict = {}
    discovered_names: dict = {}

    async def process_game(g):
        box_res, land_res = await asyncio.gather(
            client.get(f"{NHL_API_BASE}/gamecenter/{g['id']}/boxscore"),
            client.get(f"{NHL_API_BASE}/gamecenter/{g['id']}/landing"),
            return_exceptions=True,
        )
        if isinstance(box_res, Exception) or box_res.status_code != 200:
            return
        box = box_res.json()
        is_home = (box.get("homeTeam", {}).get("abbrev") or g.get("homeTeam", {}).get("abbrev")) == team_abbr
        side = box.get("playerByGameStats", {}).get("homeTeam" if is_home else "awayTeam")
        if not side:
            return

        # Real per-game PP points (goals AND assists) from landing's
        # scoring summary — each goal there carries a strength field
        # ('ev'/'pp'/'sh') plus the full assists list, so a PP goal
        # credits the scorer and every assister with a PP point. The
        # boxscore's own skater stats only expose powerPlayGoals (no PP
        # assists), which is why this second fetch is needed.
        pp_points_this_game: dict = {}
        if not isinstance(land_res, Exception) and land_res.status_code == 200:
            landing = land_res.json()
            goals = [gl for period in landing.get("summary", {}).get("scoring", []) for gl in period.get("goals", [])]
            for gl in goals:
                if gl.get("strength") != "pp":
                    continue
                scorer = gl.get("playerId")
                if scorer is not None:
                    pp_points_this_game[scorer] = pp_points_this_game.get(scorer, 0) + 1
                for a in gl.get("assists") or []:
                    aid = a.get("playerId")
                    if aid is not None:
                        pp_points_this_game[aid] = pp_points_this_game.get(aid, 0) + 1

        for p in side.get("forwards", []) + side.get("defense", []):
            pid = p.get("playerId")
            if pid is None:
                continue
            if p.get("sweaterNumber") is not None:
                discovered_jerseys[pid] = p["sweaterNumber"]
            if p.get("name"):
                discovered_names[pid] = _nhl_player_display_name(p)
            t = skater_totals.setdefault(pid, {
                "position": p.get("position") or p.get("positionCode"),
                "gp": 0, "g": 0, "a": 0, "points": 0, "ppGoals": 0, "ppPoints": 0,
                "plusMinus": 0, "pim": 0, "toiSecTotal": 0,
            })
            t["gp"] += 1
            t["g"] += p.get("goals", 0)
            t["a"] += p.get("assists", 0)
            t["points"] += p.get("points", p.get("goals", 0) + p.get("assists", 0))
            t["ppGoals"] += p.get("powerPlayGoals", 0)
            t["ppPoints"] += pp_points_this_game.get(pid, 0)
            t["plusMinus"] += p.get("plusMinus", 0)
            t["pim"] += p.get("pim", p.get("penaltyMinutes", 0))
            t["toiSecTotal"] += _parse_toi_to_seconds(p.get("toi"))

        # The boxscore's goalies array carries an explicit starter boolean
        # per game (confirmed live) — used directly for GS, no TOI-based
        # guessing. Backups who dressed but never played still show up
        # here with "00:00" TOI and starter:false; the toi<=0 guard below
        # is what excludes them from GP.
        for gk in side.get("goalies", []):
            toi_sec = _parse_toi_to_seconds(gk.get("toi"))
            if gk.get("sweaterNumber") is not None:
                discovered_jerseys[gk["playerId"]] = gk["sweaterNumber"]
            if gk.get("name"):
                discovered_names[gk["playerId"]] = _nhl_player_display_name(gk)
            if toi_sec <= 0:
                continue
            pid = gk["playerId"]
            t = goalie_totals.setdefault(pid, {"gp": 0, "gs": 0, "saves": 0, "shotsAgainst": 0, "goalsAgainst": 0, "toiSec": 0})
            t["gp"] += 1
            if gk.get("starter"):
                t["gs"] += 1
            saves, shots_against = _parse_save_shots_against(gk.get("saveShotsAgainst"))
            t["saves"] += saves
            t["shotsAgainst"] += shots_against
            t["goalsAgainst"] += gk.get("goalsAgainst", shots_against - saves)
            t["toiSec"] += toi_sec

    await asyncio.gather(*(process_game(g) for g in played))

    gsax_map, woi_map = ({}, {})
    if game_dates:
        gsax_map, woi_map = await asyncio.gather(
            asyncio.to_thread(_goalie_gsax_threaded, team_abbr, None, game_dates),
            asyncio.to_thread(_skater_woi_threaded, team_abbr, None, game_dates),
        )
    roster_map = await _fetch_jersey_map(client, team_abbr, season)

    skaters = []
    for pid, s in skater_totals.items():
        skaters.append({
            "playerId": pid,
            "name": discovered_names.get(pid, "Unknown"),
            "jersey": discovered_jerseys.get(pid),
            "position": s["position"],
            "gp": s["gp"], "g": s["g"], "a": s["a"], "points": s["points"],
            "ppGoals": s["ppGoals"],
            "ppPoints": s["ppPoints"],
            "plusMinus": s["plusMinus"], "pim": s["pim"],
            "toiSec": (s["toiSecTotal"] / s["gp"]) if s["gp"] else None,
            "esXgPctWoi": (woi_map.get(str(pid)) or {}).get("xgPct"),
            "onRoster": pid in roster_map,
        })
    goalies = []
    for pid, g in goalie_totals.items():
        goalies.append({
            "playerId": pid,
            "name": discovered_names.get(pid, "Unknown"),
            "jersey": discovered_jerseys.get(pid),
            "gp": g["gp"], "gs": g["gs"],
            "gaa": (g["goalsAgainst"] * 3600 / g["toiSec"]) if g["toiSec"] > 0 else None,
            "svPct": (g["saves"] / g["shotsAgainst"]) if g["shotsAgainst"] > 0 else None,
            "gsax": (gsax_map.get(str(pid)) or {}).get("gsax"),
            "onRoster": pid in roster_map,
        })
    goalies.sort(key=lambda g: g["gp"], reverse=True)
    goalies = goalies[:3]

    is_fwd = lambda pos: pos in _IS_FORWARD_POS
    is_def = lambda pos: pos == "D"
    return {
        "forwards": _top_by_stat(skaters, is_fwd, "points", 6),
        "defense": _top_by_stat(skaters, is_def, "points", 4),
        "goalies": goalies,
        "discoveredJerseys": discovered_jerseys,
        "gameDates": game_dates,
    }


def _prev_season_str(season: str) -> str:
    """'20262027' -> '20252026'."""
    start = int(season[:4])
    return f"{start - 1}{start}"


async def _build_prep_payload(next_game: dict | None = None) -> dict:
    async with httpx.AsyncClient(timeout=15.0) as client:
        if next_game is None:
            next_game = await _find_next_edm_game(client)
        if not next_game:
            return {"opponent": None, "computedAt": dt.datetime.now(dt.timezone.utc).isoformat()}

        opponent, season = next_game["opponent"], next_game["season"]
        stats_season = season

        async def try_season(season_to_try):
            return await asyncio.gather(
                _fetch_season_scorer_data(client, opponent, season_to_try),
                _fetch_last5_scorer_data(client, opponent, season_to_try),
            )

        season_data, last5_data = await try_season(season)

        # The upcoming game's own season can have zero games played
        # yet — e.g. right now, in the offseason, before 2026-27 has
        # actually started (the schedule is published well before puck
        # drop, but club-stats for a season with no games played comes
        # back empty). Fall back to the most recently completed season
        # so Pre-Scout Prep still shows something meaningful instead of
        # going blank for the entire offseason.
        if not season_data["skaters"]:
            prev_season = _prev_season_str(season)
            fb_season, fb_last5 = await try_season(prev_season)
            if fb_season["skaters"]:
                season_data, last5_data = fb_season, fb_last5
                stats_season = prev_season

    # Self-heal names/jerseys both directions, same idea as the old
    # client-side version: last-5's boxscore-abbreviated names get
    # upgraded to full names from the season lookup; season jersey gaps
    # (e.g. a player traded in since the roster snapshot) get patched from
    # whatever the last-5 boxscores actually showed.
    for p in last5_data["forwards"] + last5_data["defense"] + last5_data["goalies"]:
        info = season_data["nameMap"].get(p["playerId"])
        if info:
            if info.get("name"):
                p["name"] = info["name"]
            if p.get("jersey") is None and info.get("jersey") is not None:
                p["jersey"] = info["jersey"]
    for p in season_data["skaters"] + season_data["goalies"]:
        if p.get("jersey") is None:
            jersey = last5_data["discoveredJerseys"].get(p["playerId"])
            if jersey is not None:
                p["jersey"] = jersey

    # Opponent PK report (McFarland's tab) + opponent PP report (Smith's
    # tab, the inverse) — both purely PSF-derived, no NHL API involved,
    # scoped to whatever games actually exist in nhl.db for this team
    # (season: every coded game that season; last-5: whichever of the
    # same 5 game dates above happen to have PSF data — could be fewer
    # than 5, or even 0, if not everything's been coded). These four are
    # fully independent of each other, so they run as one combined
    # gather rather than two sequential ones — this is the main lever on
    # cold-start rebuild time (i.e. right after a deploy, before the
    # background refresh loop has warmed the cache).
    season_pk, last5_pk, season_pp, last5_pp = await asyncio.gather(
        asyncio.to_thread(_pk_report_threaded, opponent, stats_season, None),
        asyncio.to_thread(_pk_report_threaded, opponent, None, last5_data.get("gameDates") or []),
        asyncio.to_thread(_pp_report_threaded, opponent, stats_season, None),
        asyncio.to_thread(_pp_report_threaded, opponent, None, last5_data.get("gameDates") or []),
    )
    season_data["pk"] = season_pk
    last5_data["pk"] = last5_pk
    season_data["pp"] = season_pp
    last5_data["pp"] = last5_pp

    return {
        **next_game,
        "computedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "statsSeason": stats_season,  # next_game's own "season" key gets overwritten below —
                                       # this is the one to use for anything scoping to the
                                       # same season as season.forwards/season.pk/etc.
        "season": {k: v for k, v in season_data.items() if k not in ("skaters", "nameMap")},
        "last5": {k: v for k, v in last5_data.items() if k != "discoveredJerseys"},
    }


_prep_cache: dict = {}
_prep_lock = asyncio.Lock()

# Multi-game Pre-Scout Prep support: the game picker (next PREP_UPCOMING_COUNT
# EDM games) needs a lightweight schedule-only list instantly, plus the full
# season/last5/pk/pp payload for whichever specific game a coach clicks into.
# _prep_cache above is kept exactly as before (always mirrors game #1, i.e.
# the literal next game) so /api/prep/next-game needs zero changes and stays
# backward compatible. Games 2-N build lazily on first request rather than
# unconditionally every background cycle — nobody's paying for ~15 NHL API
# calls per game, every 10 minutes, for games a coach never opens — but once
# a game HAS been opened once, it's added to _prep_warm_game_ids and the
# background loop keeps it warm from then on, same "instant on repeat visits"
# guarantee the single-game version always had.
PREP_UPCOMING_COUNT = 5
_prep_upcoming: list = []          # schedule-only list, next PREP_UPCOMING_COUNT games
_prep_cache_by_game: dict = {}     # gameId -> full payload
_prep_warm_game_ids: set = set()   # games 2-N that have been requested at least once


async def _refresh_prep_cache():
    global _prep_cache, _prep_upcoming
    async with _prep_lock:
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                games = await _find_next_n_edm_games(client, n=PREP_UPCOMING_COUNT)
            _prep_upcoming = games
            if not games:
                _prep_cache = {"opponent": None, "computedAt": dt.datetime.now(dt.timezone.utc).isoformat()}
                _prep_cache_by_game.clear()
                _prep_warm_game_ids.clear()
                return

            # Always rebuild game #1 (unchanged original behavior), plus
            # any of games 2-N that's actually been opened before.
            to_build = [games[0]] + [g for g in games[1:] if g["gameId"] in _prep_warm_game_ids]

            async def build_one(g):
                try:
                    return g["gameId"], await _build_prep_payload(g)
                except Exception as e:
                    print(f"[prep-cache] game {g['gameId']} ({g['opponent']}) refresh failed: {e}")
                    return g["gameId"], None

            results = await asyncio.gather(*(build_one(g) for g in to_build))
            for gid, payload in results:
                if payload is not None:
                    _prep_cache_by_game[gid] = payload

            # Drop anything that's rolled out of the upcoming-N window
            # entirely (game played, or bumped out by a schedule change).
            live_ids = {g["gameId"] for g in games}
            for gid in list(_prep_cache_by_game.keys()):
                if gid not in live_ids:
                    _prep_cache_by_game.pop(gid, None)
                    _prep_warm_game_ids.discard(gid)

            _prep_cache = _prep_cache_by_game.get(games[0]["gameId"]) or _prep_cache
        except Exception as e:
            print(f"[prep-cache] refresh failed: {e}")


async def _prep_cache_background_loop():
    """Keeps the cache warm proactively so a user request almost never has
    to wait on the full rebuild. Runs immediately on startup, then every
    10 minutes — 10 minutes covers both "new games finished, refresh
    last-5/season stats" and, worst case, up to a 10-minute lag on
    rollover to a new opponent — the per-request stale-while-revalidate
    check in prep_next_game below closes most of that gap faster in
    practice."""
    while True:
        await _refresh_prep_cache()
        await asyncio.sleep(600)


@app.on_event("startup")
async def _start_prep_cache_loop():
    # Fired as a background task, NOT awaited — this must never block the
    # app from accepting traffic. A previous version of this awaited the
    # first build directly in the startup handler, which in turn blocked
    # EVERY request (including login, which has nothing to do with
    # Pre-Scout Prep) until ~15 external NHL API calls finished. That
    # traded "Pre-Scout Prep is instant even the very first time" for
    # "login is slow right after every deploy," which is the wrong trade —
    # unrelated functionality should never wait on this. The real cost of
    # going back to fire-and-forget: the very first hit to
    # /api/prep/next-game after a fresh deploy can still be slow (whatever
    # request lands before this background task finishes triggers a live
    # rebuild) — but that's isolated to Pre-Scout Prep specifically, and
    # only for the first user unlucky enough to hit it right after a
    # deploy, not the whole app.
    asyncio.create_task(_prep_cache_background_loop())


@app.get("/api/prep/next-game")
async def prep_next_game(user: CurrentUser = Depends(require_login)):
    """Precomputed Pre-Scout Prep payload for EDM's next game. Instant
    once the background loop has warmed it (which it does automatically —
    see _prep_cache_background_loop). Only blocks on a live rebuild if the
    cache is completely empty (e.g. moments after a fresh deploy, before
    the first background pass lands)."""
    if not _prep_cache:
        await _refresh_prep_cache()
        return _prep_cache

    # Stale-while-revalidate: a single cheap schedule check (not the full
    # rebuild) to see if the opponent has actually changed since this was
    # cached — e.g. the previous game just started. If so, kick off the
    # full rebuild in the background and still answer THIS request
    # instantly with what's cached; the next request gets the fresh data.
    async def check_and_refresh():
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                current = await _find_next_edm_game(client)
            if current and current.get("gameId") != _prep_cache.get("gameId"):
                await _refresh_prep_cache()
        except Exception:
            pass
    asyncio.create_task(check_and_refresh())

    return _prep_cache


@app.get("/api/prep/upcoming-games")
async def prep_upcoming_games(user: CurrentUser = Depends(require_login)):
    """Schedule-only list of EDM's next PREP_UPCOMING_COUNT games — just
    enough to render the game-picker boxes instantly (gameId, date,
    opponent, home/away, real season game number). Deliberately doesn't
    wait on any of the heavier per-opponent stats (see /api/prep/game/{id}
    for those), so this is fast even on a completely cold cache: worst
    case it's the one schedule call _find_next_n_edm_games always needs,
    not a ~15-call-per-game stats build."""
    if _prep_upcoming:
        return {"games": _prep_upcoming}
    async with httpx.AsyncClient(timeout=15.0) as client:
        games = await _find_next_n_edm_games(client, n=PREP_UPCOMING_COUNT)
    return {"games": games}


@app.get("/api/prep/game/{game_id}")
async def prep_game(game_id: int, user: CurrentUser = Depends(require_login)):
    """Full Pre-Scout Prep payload (same shape as /api/prep/next-game) for
    one specific game out of the current picker list, addressed by its
    NHL gameId. Instant once warmed; the first time a given game is opened
    it triggers a live rebuild (same cold-start tradeoff /api/prep/next-game
    has always had) and from then on the background loop keeps it warm
    alongside game #1."""
    cached = _prep_cache_by_game.get(game_id)
    if cached:
        _prep_warm_game_ids.add(game_id)
        return cached

    match = next((g for g in _prep_upcoming if g["gameId"] == game_id), None)
    if not match:
        raise HTTPException(404, f"Game {game_id} isn't one of the currently listed upcoming games — refresh the game list and try again.")

    payload = await _build_prep_payload(match)
    _prep_cache_by_game[game_id] = payload
    _prep_warm_game_ids.add(game_id)
    return payload


# ── Scoring chances ──────────────────────────────────────────────────────
# Implements the classification in Scoring_Chance_Derivation_Reference.md:
# possession-chain construction backward from each slot shot, then
# Rush / Forecheck / Cycle / 2nd Chance / OZ Play (dedup union).

CYCLE_PASS_TYPES = {
    "north", "south", "east", "west", "eastwest", "northsouth",
    "northoffboards", "southoffboards", "eastoffboards", "westoffboards",
    "northeastoffboards", "northwestoffboards", "southeastoffboards", "southwestoffboards",
}
OPPONENT_RECOVERY_NAMES = {"lpr", "carry", "pass", "reception", "dumpin", "dumpout", "puckprotection"}


def _build_chain(rows_by_time, idx, team):
    """Walk backward from event idx, returning the possession chain (list of
    row dicts, oldest-first) up to (not including) the boundary event."""
    chain = []
    j = idx - 1
    while j >= 0:
        ev = rows_by_time[j]
        if ev["name"] in ("whistle", "faceoff") or ev["name"] == "penalty":
            break
        if ev["team"] != team and ev["name"] in OPPONENT_RECOVERY_NAMES:
            break
        if rows_by_time[j + 1]["gameTime"] - ev["gameTime"] > 8.0:
            break
        chain.append(ev)
        j -= 1
    chain.reverse()
    return chain


def _classify_shot(chain, shot):
    team = shot["team"]

    # Rush: controlled entry in chain, shot within ~4s, <=1 cycle pass after entry
    entry_idx = None
    for i, ev in enumerate(chain):
        if ev["team"] != team:
            continue
        is_carry_entry = ev["name"] == "carry" and ev["zone"] == "oz"
        is_rush_pass = ev["name"] == "pass" and ev["type"] == "rush"
        is_ozentry = ev["name"] in ("reception", "pass") and ev["type"] == "ozentry"
        if is_carry_entry or is_rush_pass or is_ozentry:
            entry_idx = i
    is_rush = False
    if entry_idx is not None:
        entry_ev = chain[entry_idx]
        time_since_entry = shot["gameTime"] - entry_ev["gameTime"]
        cycle_passes_after = sum(
            1 for ev in chain[entry_idx + 1:]
            if ev["team"] == team and ev["name"] == "pass" and ev["type"] in CYCLE_PASS_TYPES
        )
        if time_since_entry <= 4.0 and cycle_passes_after <= 1:
            is_rush = True

    # Forecheck: no controlled entry, first OZ LPR by team preceded by team check/block
    is_forecheck = False
    if entry_idx is None:
        for i, ev in enumerate(chain):
            if ev["team"] == team and ev["name"] == "lpr" and ev["zone"] == "oz":
                preceding = chain[:i]
                if any(p["team"] == team and p["name"] in ("check", "block") for p in preceding):
                    is_forecheck = True
                break

    # Cycle: team perimeter pass in OZ after the most recent PRIOR shot in the chain
    last_shot_idx = -1
    for i, ev in enumerate(chain):
        if ev["name"] == "shot":
            last_shot_idx = i
    is_cycle = any(
        ev["team"] == team and ev["name"] == "pass" and ev["type"] in CYCLE_PASS_TYPES and ev["zone"] == "oz"
        for ev in chain[last_shot_idx + 1:]
    )

    # 2nd Chance: rebound LPR by team within 3s of shot, no intervening shot/cycle pass
    is_second_chance = False
    if chain:
        last = chain[-1]
        if (last["team"] == team and last["name"] == "lpr" and last["type"] == "rebound"
                and shot["gameTime"] - last["gameTime"] <= 3.0):
            is_second_chance = True

    return {
        "rush": is_rush,
        "forecheck": is_forecheck,
        "cycle": is_cycle,
        "secondChance": is_second_chance,
        "ozPlay": is_forecheck or is_cycle or is_second_chance,
    }


def compute_chances(conn, game_id: int, team_abbr: str):
    team_full = full_name(team_abbr)
    rows = conn.execute(
        "SELECT id, gameTime, period, team, name, type, zone FROM plays "
        "WHERE gameReferenceId = ? ORDER BY gameTime",
        (game_id,),
    ).fetchall()
    rows = [dict(r) for r in rows]

    totals = {"total": 0, "rush": 0, "forecheck": 0, "cycle": 0, "secondChance": 0, "ozPlay": 0}
    shots_detail = []

    for idx, ev in enumerate(rows):
        if ev["team"] != team_full or ev["name"] != "shot" or ev["type"] not in ("slot", "slotblocked"):
            continue
        totals["total"] += 1
        chain = _build_chain(rows, idx, team_full)
        cls = _classify_shot(chain, ev)
        for k in ("rush", "forecheck", "cycle", "secondChance", "ozPlay"):
            totals[k] += 1 if cls[k] else 0
        shots_detail.append({
            "eventId": ev["id"], "period": ev["period"], "gameTime": ev["gameTime"],
            "blocked": ev["type"] == "slotblocked", **cls,
        })

    return {"totals": totals, "shots": shots_detail}


@app.get("/api/game/{game_id}/chances")
def game_chances(game_id: int, team: str = Query(..., description="Team abbreviation, e.g. EDM"), user: CurrentUser = Depends(require_login)):
    """Scoring-chance classification (rush/forecheck/cycle/2nd chance/OZ play)
    for one team in one game. Known accuracy per the derivation reference:
    Total/Rush exact, Forecheck/2nd Chance within ~1, Cycle/OZ Play softer
    (tracking-layer limits on 'heading to the net' / buttonhook judgments)."""
    conn = get_db()
    game_or_404(conn, game_id)
    result = compute_chances(conn, game_id, team.upper())
    conn.close()
    return {"gameId": game_id, "team": team.upper(), **result}


# ── Coordinates (shots, entries) ─────────────────────────────────────────

@app.get("/api/game/{game_id}/coordinates")
def game_coordinates(
    game_id: int,
    team: str | None = Query(None, description="Filter to one team's events"),
    event: str = Query("shot", description="Event name: shot, carry, pass, lpr, etc."),
    strength: str | None = Query(None, description="evenStrength | powerPlay | shortHanded"),
    user: CurrentUser = Depends(require_login),
):
    """Raw (adjusted) coordinates for a given event type — the building
    block for shot charts / zone-entry maps. xAdjCoord/yAdjCoord are
    already normalized so attack direction is consistent across periods
    (see Coordinate_Adjustment_Wing_Side_Reference.md for how that's
    verified)."""
    conn = get_db()
    game_or_404(conn, game_id)
    q = ("SELECT id, period, gameTime, team, name, type, zone, xAdjCoord, yAdjCoord, "
         "manpowerSituation, playerReferenceId, playerFirstName, playerLastName "
         "FROM plays WHERE gameReferenceId = ? AND name = ?")
    params = [game_id, event]
    if team:
        q += " AND team = ?"
        params.append(full_name(team))
    if strength:
        q += " AND manpowerSituation = ?"
        params.append(strength)
    q += " AND xAdjCoord IS NOT NULL ORDER BY gameTime"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/pk-entry-locations/{team}")
def pk_entry_locations(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD), e.g. a last-5 cut"),
    user: CurrentUser = Depends(require_login),
):
    """Raw locations for every controlled zone entry attempt against this
    team's PK (4v5), across a season or a specific set of games — the
    data feed for a future heat map showing WHERE they allow entries.
    xAdjCoord/yAdjCoord are already normalized so attack direction is
    consistent across periods (teams switch ends for period 2) — no
    extra flip logic needed on the consuming side.

    Not yet wired to a visual (waiting on rink diagram PNGs); this just
    exposes the underlying data so that hookup is fast once those exist."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    team_abbr = team.upper()
    team_full = full_name(team_abbr)
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=date_list)
        if not game_ids:
            return []
        placeholders = ",".join("?" for _ in game_ids)
        rows = conn.execute(
            f"""SELECT gameReferenceId, xAdjCoord, yAdjCoord, outcome
                FROM plays
                WHERE gameReferenceId IN ({placeholders})
                  AND name = 'controlledentryagainst' AND team = ? AND manpowerSituation = 'shortHanded'
                  AND xAdjCoord IS NOT NULL AND yAdjCoord IS NOT NULL""",
            game_ids + [team_full],
        ).fetchall()
        return [
            {"gameId": r["gameReferenceId"], "x": r["xAdjCoord"], "y": r["yAdjCoord"], "allowed": r["outcome"] == "failed"}
            for r in rows
        ]
    finally:
        conn.close()


@app.get("/api/pk-shot-locations/{team}")
def pk_shot_locations(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD), e.g. a last-5 cut"),
    user: CurrentUser = Depends(require_login),
):
    """Raw locations for every shot ATTEMPT against this team's PK (4v5),
    across a season or a specific set of games — the data feed for a
    density heat map showing WHERE the opponent generates shots against
    this team's penalty kill. Includes blocked attempts (the release
    point still matters for showing volume-by-location) since this is
    about attempt density, not xG accuracy — unlike goalie_gsax, which
    correctly excludes blocked shots because they never reached the
    goalie.

    Returns {} if no games match (a team with no PSF data yet is a normal,
    expected case here, not an error)."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    team_abbr = team.upper()
    team_full = full_name(team_abbr)
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=date_list)
        if not game_ids:
            return []
        placeholders = ",".join("?" for _ in game_ids)
        rows = conn.execute(
            f"""SELECT xAdjCoord, yAdjCoord FROM plays
                WHERE gameReferenceId IN ({placeholders})
                  AND team != ? AND manpowerSituation = 'powerPlay'
                  AND name = 'shot'
                  AND xAdjCoord IS NOT NULL AND yAdjCoord IS NOT NULL""",
            game_ids + [team_full],
        ).fetchall()
        return [{"x": r["xAdjCoord"], "y": r["yAdjCoord"]} for r in rows]
    finally:
        conn.close()


@app.get("/api/pp-entry-locations/{team}")
def pp_entry_locations(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD), e.g. a last-5 cut"),
    user: CurrentUser = Depends(require_login),
):
    """Inverse of pk_entry_locations: raw locations for every controlled
    zone entry this team ACHIEVES while on the power play (5v4), across a
    season or a specific set of games — Smith's tab, since he's prepping
    our PK against the opponent's PP rather than our PP against their PK.
    Entries are tagged under the DEFENDING team, so this queries rows
    tagged under team's opponent (team != team_full) while that opponent
    is shorthanded, within team's own games."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    team_abbr = team.upper()
    team_full = full_name(team_abbr)
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=date_list)
        if not game_ids:
            return []
        placeholders = ",".join("?" for _ in game_ids)
        rows = conn.execute(
            f"""SELECT gameReferenceId, xAdjCoord, yAdjCoord, outcome
                FROM plays
                WHERE gameReferenceId IN ({placeholders})
                  AND name = 'controlledentryagainst' AND team != ? AND manpowerSituation = 'shortHanded'
                  AND xAdjCoord IS NOT NULL AND yAdjCoord IS NOT NULL""",
            game_ids + [team_full],
        ).fetchall()
        return [
            {"gameId": r["gameReferenceId"], "x": r["xAdjCoord"], "y": r["yAdjCoord"], "allowed": r["outcome"] == "failed"}
            for r in rows
        ]
    finally:
        conn.close()


@app.get("/api/pp-shot-locations/{team}")
def pp_shot_locations(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD), e.g. a last-5 cut"),
    user: CurrentUser = Depends(require_login),
):
    """Inverse of pk_shot_locations: raw locations for every shot attempt
    this team generates while ON the power play (5v4), across a season or
    a specific set of games — the data feed for Smith's density heat map
    showing where the upcoming opponent likes to shoot from on their own
    power play, so our PK knows what to take away. Includes blocked
    attempts for the same volume-by-location reasoning as the PK version."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    team_abbr = team.upper()
    team_full = full_name(team_abbr)
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=date_list)
        if not game_ids:
            return []
        placeholders = ",".join("?" for _ in game_ids)
        rows = conn.execute(
            f"""SELECT xAdjCoord, yAdjCoord FROM plays
                WHERE gameReferenceId IN ({placeholders})
                  AND team = ? AND manpowerSituation = 'powerPlay'
                  AND name = 'shot'
                  AND xAdjCoord IS NOT NULL AND yAdjCoord IS NOT NULL""",
            game_ids + [team_full],
        ).fetchall()
        return [{"x": r["xAdjCoord"], "y": r["yAdjCoord"]} for r in rows]
    finally:
        conn.close()


@app.get("/api/pk-highest-xga-shots/{team}")
def pk_highest_xga_shots(
    team: str,
    season: str | None = Query(None, description="e.g. 20252026"),
    dates: str | None = Query(None, description="Comma-separated gameDate list (YYYY-MM-DD), e.g. a last-5 cut"),
    limit: int = Query(10, description="How many top-xG shots to return"),
    user: CurrentUser = Depends(require_login),
):
    """The highest-expectedGoalsOnNet shot attempts taken against this
    team's PK (4v5), each paired with the immediately preceding event
    (usually the setup pass) — the data feed for a future diagram showing
    the path into their most dangerous chances allowed, not just the shot
    location alone.

    Not yet wired to a visual (waiting on rink diagram PNGs); this just
    exposes the underlying data so that hookup is fast once those exist."""
    if not season and not dates:
        raise HTTPException(400, "Provide either 'season' or 'dates' (comma-separated YYYY-MM-DD)")
    team_abbr = team.upper()
    team_full = full_name(team_abbr)
    date_list = [d.strip() for d in dates.split(",") if d.strip()] if dates else None
    conn = get_db()
    try:
        game_ids = _resolve_game_ids(conn, team_abbr, season=season, dates=date_list)
        if not game_ids:
            return []
        placeholders = ",".join("?" for _ in game_ids)
        shots = conn.execute(
            f"""SELECT gameReferenceId, id, xAdjCoord, yAdjCoord, expectedGoalsOnNet, type
                FROM plays
                WHERE gameReferenceId IN ({placeholders})
                  AND team != ? AND manpowerSituation = 'powerPlay'
                  AND name = 'shot' AND type NOT IN ('slotblocked', 'outsideblocked')
                  AND expectedGoalsOnNet IS NOT NULL
                  AND xAdjCoord IS NOT NULL AND yAdjCoord IS NOT NULL
                ORDER BY expectedGoalsOnNet DESC
                LIMIT ?""",
            game_ids + [team_full, limit],
        ).fetchall()

        results = []
        for s in shots:
            prev = conn.execute(
                """SELECT name, type, xAdjCoord, yAdjCoord FROM plays
                   WHERE gameReferenceId = ? AND id < ? ORDER BY id DESC LIMIT 1""",
                (s["gameReferenceId"], s["id"]),
            ).fetchone()
            results.append({
                "gameId": s["gameReferenceId"],
                "shot": {"x": s["xAdjCoord"], "y": s["yAdjCoord"], "xg": s["expectedGoalsOnNet"], "type": s["type"]},
                "setup": ({"name": prev["name"], "type": prev["type"], "x": prev["xAdjCoord"], "y": prev["yAdjCoord"]}
                          if prev and prev["xAdjCoord"] is not None else None),
            })
        return results
    finally:
        conn.close()


@app.get("/")
def index():
    return {
        "service": "Coaches Station API",
        "endpoints": [
            "/api/games",
            "/api/team/{abbr}/games",
            "/api/matchup/{team_a}/{team_b}",
            "/api/game/{game_id}/lines?team=EDM",
            "/api/game/{game_id}/special-teams?team=EDM&situation=5v4",
            "/api/game/{game_id}/lineup?team=EDM",
            "/api/game/{game_id}/chances?team=EDM",
            "/api/game/{game_id}/coordinates?team=EDM&event=shot&strength=evenStrength",
            "POST /api/goal-categories/upload (multipart file upload, .xlsx)",
            "/api/goal-categories/export (downloads a regenerated GC.xlsx)",
            "/api/goal-categories/summary?team=EDM",
        ],
    }


# ── Goal Category import/export ──────────────────────────────────────────
# Parsing logic here mirrors processGoalCategoryData in Coaches_Station.html
# (Timeline/Row regex, team-name normalization, corrupted-date-value fix) —
# validated against a real season's GC.xlsx earlier in this project. Kept
# as a second, independent implementation (Python here, JS in the browser)
# rather than sharing code, since the two run in different languages; if
# the browser-side parsing logic ever changes, this needs updating too.

import re
import io
import datetime as dt
import openpyxl
from fastapi import UploadFile, File

TIMELINE_RE = re.compile(r'^(\d{4}-\d{2}-\d{2})\s+(.+?)\s+vs\s+(.+)$')
SCORER_RE = re.compile(r'^(.+?)\s+Goal$')
FULL_NAME_TO_ABBR = {v.upper(): k for k, v in TEAM_NAMES.items()}
GC_NAME_ALIASES = {"MONTR\u221A\u00C2AL CANADIENS": "MTL"}  # known mojibake in some exports

# Same (macroCategory, month, day) -> corrected value table as Coaches_Station.html's
# GC_SYSTEM_FIX, for Excel cells where a system value like "2-2" or "1-3" got
# auto-converted into a date by Excel.
GC_SYSTEM_FIX = {
    ("LRA", 2, 2): "1-2-2", ("LRA", 1, 3): "1-1-3", ("FC", 2, 2): "1-2-2",
    ("PKFC", 1, 2): "1-1-2", ("PKFC", 1, 3): "1-3", ("PKFC", 3, 1): "3-1",
}


def gc_name_to_abbr(full_name: str):
    cleaned = re.sub(r"\s+R\d+G\d+\s*$", "", full_name or "").strip().upper()
    return FULL_NAME_TO_ABBR.get(cleaned) or GC_NAME_ALIASES.get(cleaned)


def gc_fix_system_value(macro_category, raw_value):
    if raw_value is None or raw_value == "":
        return None
    if isinstance(raw_value, (dt.datetime, dt.date)):
        key = (macro_category, raw_value.month, raw_value.day)
        return GC_SYSTEM_FIX.get(key, f"[unrecognized {raw_value.month}/{raw_value.day}]")
    return str(raw_value).strip()


class ParsedGoalRecord(BaseModel):
    gameDate: str
    homeTeamAbbrev: str
    awayTeamAbbrev: str
    scoringTeamAbbrev: str
    concedingTeamAbbrev: str
    instanceNum: int
    gaCat: str | None = None
    gaSys: str | None = None
    gfCat: str | None = None
    gfSys: str | None = None
    sourceFile: str | None = None


@app.post("/api/goal-categories/import-parsed")
def import_parsed_goals(records: list[ParsedGoalRecord], admin: CurrentUser = Depends(require_admin)):
    """Accepts goal records already parsed and validated client-side from
    raw .SCTimeline exports (see the GC Uploader page) — the browser does
    all the parsing/validation (same logic GC Updater always used, plus
    the NHL boxscore cross-check), and this endpoint just needs to insert
    the final, clean results. Same dedup behavior as the Excel-upload
    endpoint: safe to re-submit, duplicates are skipped rather than erroring."""
    conn = get_db()
    inserted, skipped = 0, 0
    for r in records:
        game_row = conn.execute(
            "SELECT gameReferenceId FROM games WHERE gameDate=? AND homeTeamAbbrev=? AND awayTeamAbbrev=?",
            (r.gameDate, r.homeTeamAbbrev, r.awayTeamAbbrev),
        ).fetchone()
        game_ref_id = game_row["gameReferenceId"] if game_row else None
        try:
            conn.execute(
                "INSERT INTO goal_categories "
                "(gameReferenceId, gameDate, homeTeamAbbrev, awayTeamAbbrev, scoringTeamAbbrev, "
                " concedingTeamAbbrev, instanceNum, gaCat, gaSys, gfCat, gfSys, sourceFile) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (game_ref_id, r.gameDate, r.homeTeamAbbrev, r.awayTeamAbbrev, r.scoringTeamAbbrev,
                 r.concedingTeamAbbrev, r.instanceNum, r.gaCat, r.gaSys, r.gfCat, r.gfSys, r.sourceFile),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            skipped += 1
    conn.commit()
    conn.close()
    return {"goalsInserted": inserted, "goalsSkipped_alreadyImported": skipped}


@app.post("/api/goal-categories/upload")
async def upload_goal_categories(file: UploadFile = File(...), user: CurrentUser = Depends(require_admin)):
    """Accepts a GC.xlsx (the GC Updater's output, or any GC.xlsx built the
    same way) and imports every goal into the goal_categories table.
    Duplicate goals (same date/matchup/scorer/instance) are silently
    skipped — safe to re-upload the same file, or a file with some games
    already present."""
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Couldn't read this as an Excel file: {e}")
    if "GF" not in wb.sheetnames:
        raise HTTPException(400, "No 'GF' sheet found — expected the same GF/GA sheet structure as GC.xlsx")

    ws = wb["GF"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx = {h: headers.index(h) for h in
               ["Timeline", "Row", "Original Instance", "GA Goal Category", "GA System", "GF Goal Category", "GF System"]}
    except ValueError as e:
        raise HTTPException(400, f"Missing expected column: {e}")

    instance_re = re.compile(r":\s*(\d+)\s*$")
    conn = get_db()
    inserted, skipped, unresolved_teams, unresolved_rows = 0, 0, set(), 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        timeline = str(row[idx["Timeline"]] or "").strip()
        row_label = str(row[idx["Row"]] or "").strip()
        orig_instance = str(row[idx["Original Instance"]] or "").strip()
        tl_match = TIMELINE_RE.match(timeline)
        sc_match = SCORER_RE.match(row_label)
        inst_match = instance_re.search(orig_instance)
        if not tl_match or not sc_match or not inst_match:
            unresolved_rows += 1
            continue

        game_date, home_name, away_name = tl_match.groups()
        scorer_name = sc_match.group(1)
        instance_num = int(inst_match.group(1))
        home_abbr = gc_name_to_abbr(home_name)
        away_abbr = gc_name_to_abbr(away_name)
        scorer_abbr = gc_name_to_abbr(scorer_name)
        if not home_abbr or not away_abbr or not scorer_abbr:
            for n in (home_name, away_name, scorer_name):
                if not gc_name_to_abbr(n):
                    unresolved_teams.add(n)
            unresolved_rows += 1
            continue
        conceding_abbr = away_abbr if scorer_abbr == home_abbr else home_abbr

        ga_cat = row[idx["GA Goal Category"]] or None
        gf_cat = row[idx["GF Goal Category"]] or None
        ga_sys = gc_fix_system_value(ga_cat, row[idx["GA System"]])
        gf_sys = gc_fix_system_value(gf_cat, row[idx["GF System"]])

        game_row = conn.execute(
            "SELECT gameReferenceId FROM games WHERE gameDate=? AND homeTeamAbbrev=? AND awayTeamAbbrev=?",
            (game_date, home_abbr, away_abbr),
        ).fetchone()
        game_ref_id = game_row["gameReferenceId"] if game_row else None

        try:
            conn.execute(
                "INSERT INTO goal_categories "
                "(gameReferenceId, gameDate, homeTeamAbbrev, awayTeamAbbrev, scoringTeamAbbrev, "
                " concedingTeamAbbrev, instanceNum, gaCat, gaSys, gfCat, gfSys, sourceFile) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (game_ref_id, game_date, home_abbr, away_abbr, scorer_abbr, conceding_abbr,
                 instance_num, ga_cat, ga_sys, gf_cat, gf_sys, file.filename),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            skipped += 1  # already imported (unique index on date+matchup+scorer+instance)

    conn.commit()
    conn.close()

    return {
        "filename": file.filename,
        "goalsInserted": inserted,
        "goalsSkipped_alreadyImported": skipped,
        "rowsUnresolved": unresolved_rows,
        "unresolvedTeamNames": sorted(unresolved_teams) if unresolved_teams else [],
    }


@app.get("/api/goal-categories/export")
def export_goal_categories(team: str | None = Query(None), season_from: str | None = Query(None), season_to: str | None = Query(None), user: CurrentUser = Depends(require_login)):
    """Regenerates a GC.xlsx (GF/GA sheets, same column layout as the
    original) straight from whatever's stored in goal_categories — the
    reverse direction of /upload."""
    conn = get_db()
    q = "SELECT * FROM goal_categories WHERE 1=1"
    params = []
    if team:
        q += " AND (scoringTeamAbbrev = ? OR concedingTeamAbbrev = ?)"
        params += [team.upper(), team.upper()]
    if season_from:
        q += " AND gameDate >= ?"
        params.append(season_from)
    if season_to:
        q += " AND gameDate <= ?"
        params.append(season_to)
    rows = conn.execute(q, params).fetchall()
    conn.close()

    if not rows:
        raise HTTPException(404, "No goal-category data matches these filters")

    wb = openpyxl.Workbook()
    ws_gf = wb.active
    ws_gf.title = "GF"
    ws_ga = wb.create_sheet("GA")
    header = ["Timeline", "Row", "Original Instance", "GA Goal Category", "GA System", "GF Goal Category", "GF System"]
    ws_gf.append(header)
    ws_ga.append(header)

    for r in rows:
        home_full = TEAM_NAMES.get(r["homeTeamAbbrev"], r["homeTeamAbbrev"])
        away_full = TEAM_NAMES.get(r["awayTeamAbbrev"], r["awayTeamAbbrev"])
        scorer_full = TEAM_NAMES.get(r["scoringTeamAbbrev"], r["scoringTeamAbbrev"])
        conceder_full = TEAM_NAMES.get(r["concedingTeamAbbrev"], r["concedingTeamAbbrev"])
        timeline_str = f"{r['gameDate']} {home_full} vs {away_full}"
        gf_row_label = f"{scorer_full} Goal"
        ga_row_label = f"{conceder_full} Goal Against"
        ws_gf.append([timeline_str, gf_row_label, f"{gf_row_label}: {r['instanceNum']}",
                      r["gaCat"], r["gaSys"], r["gfCat"], r["gfSys"]])
        ws_ga.append([timeline_str, ga_row_label, f"{ga_row_label}: {r['instanceNum']}",
                      r["gaCat"], r["gaSys"], r["gfCat"], r["gfSys"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=GC_export.xlsx"},
    )


@app.get("/api/goal-categories/summary")
def goal_categories_summary(team: str = Query(...), user: CurrentUser = Depends(require_login)):
    """Quick sanity-check endpoint: how many goals are stored for a team,
    by category, GF and GA. Useful for confirming an upload actually landed."""
    conn = get_db()
    abbr = team.upper()
    gf = conn.execute(
        "SELECT gfCat as category, COUNT(*) as goals FROM goal_categories "
        "WHERE scoringTeamAbbrev = ? GROUP BY gfCat ORDER BY goals DESC", (abbr,)
    ).fetchall()
    ga = conn.execute(
        "SELECT gaCat as category, COUNT(*) as goals FROM goal_categories "
        "WHERE concedingTeamAbbrev = ? GROUP BY gaCat ORDER BY goals DESC", (abbr,)
    ).fetchall()
    conn.close()
    return {"team": abbr, "goalsFor": [dict(r) for r in gf], "goalsAgainst": [dict(r) for r in ga]}


def _gc_in_scope(date, gtype, date_from, date_to):
    if date_from and date < date_from:
        return False
    if date_to and date > date_to:
        return False
    if gtype == "rs":
        return date <= RS_END
    if gtype == "po":
        return date > RS_END
    return True


@app.get("/api/goal-categories/league-table")
def gc_league_table(
    side: str = Query(..., description="'gf' or 'ga'"),
    type: str = Query("all", description="'all' | 'rs' | 'po'"),
    date_from: str | None = Query(None, alias="from"),
    date_to: str | None = Query(None, alias="to"),
    team: str | None = Query(None, description="If set, one row per opponent instead of per team"),
    user: CurrentUser = Depends(require_login),
):
    """Database-driven replacement for what used to require a fresh GC.xlsx
    upload every session — same shape (rows + categories) the existing
    table-rendering code already expects, just sourced from goal_categories
    instead of a client-side-parsed Excel file. GP comes from the `games`
    table directly (not from goal data), so a team with zero goals in a
    stretch still shows the correct games-played count instead of vanishing."""
    if side not in ("gf", "ga"):
        raise HTTPException(400, "side must be 'gf' or 'ga'")
    conn = get_db()

    games = conn.execute("SELECT gameReferenceId, gameDate, homeTeamAbbrev, awayTeamAbbrev FROM games").fetchall()
    games = [g for g in games if _gc_in_scope(g["gameDate"], type, date_from, date_to)]
    valid_gids = {g["gameReferenceId"] for g in games}
    game_by_id = {g["gameReferenceId"]: g for g in games}

    goal_rows = conn.execute("SELECT gameReferenceId, scoringTeamAbbrev, concedingTeamAbbrev, gaCat, gfCat FROM goal_categories").fetchall()
    conn.close()

    cat_totals = defaultdict(int)
    rows = defaultdict(lambda: {"gp": 0, "total": 0, "cats": defaultdict(int)})
    team = team.upper() if team else None

    if team:
        for g in games:
            if team not in (g["homeTeamAbbrev"], g["awayTeamAbbrev"]):
                continue
            opp = g["awayTeamAbbrev"] if g["homeTeamAbbrev"] == team else g["homeTeamAbbrev"]
            rows[opp]["gp"] += 1
        for gr in goal_rows:
            if gr["gameReferenceId"] not in valid_gids:
                continue
            g = game_by_id[gr["gameReferenceId"]]
            if team not in (g["homeTeamAbbrev"], g["awayTeamAbbrev"]):
                continue
            opp = g["awayTeamAbbrev"] if g["homeTeamAbbrev"] == team else g["homeTeamAbbrev"]
            if side == "gf" and gr["scoringTeamAbbrev"] == team and gr["gfCat"] != "Shootout":
                cat = gr["gfCat"]
            elif side == "ga" and gr["concedingTeamAbbrev"] == team and gr["gaCat"] != "Shootout":
                cat = gr["gaCat"]
            else:
                continue
            rows[opp]["cats"][cat] += 1
            rows[opp]["total"] += 1
            cat_totals[cat] += 1
    else:
        gp_counts = defaultdict(int)
        for g in games:
            gp_counts[g["homeTeamAbbrev"]] += 1
            gp_counts[g["awayTeamAbbrev"]] += 1
        for abbr, count in gp_counts.items():
            rows[abbr]["gp"] = count
        for gr in goal_rows:
            if gr["gameReferenceId"] not in valid_gids:
                continue
            if side == "gf":
                abbr, cat = gr["scoringTeamAbbrev"], gr["gfCat"]
            else:
                abbr, cat = gr["concedingTeamAbbrev"], gr["gaCat"]
            if cat == "Shootout":
                continue
            rows[abbr]["cats"][cat] += 1
            rows[abbr]["total"] += 1
            cat_totals[cat] += 1

    categories = sorted(cat_totals.keys(), key=lambda c: -cat_totals[c])
    rows_out = [
        {"team": abbr, "gp": r["gp"], "total": r["total"], "cats": dict(r["cats"])}
        for abbr, r in rows.items() if r["gp"] > 0
    ]
    return {"rows": rows_out, "categories": categories}


@app.get("/api/goal-categories/discover")
def gc_discover_categories(side: str = Query(...), user: CurrentUser = Depends(require_login)):
    """Category -> list of systems, for populating the cross-search
    dropdowns without needing a client-side Excel parse to discover them."""
    if side not in ("gf", "ga"):
        raise HTTPException(400, "side must be 'gf' or 'ga'")
    conn = get_db()
    cat_col, sys_col = ("gfCat", "gfSys") if side == "gf" else ("gaCat", "gaSys")
    rows = conn.execute(
        f"SELECT DISTINCT {cat_col} as cat, {sys_col} as sys FROM goal_categories "
        f"WHERE {cat_col} IS NOT NULL AND {cat_col} != 'Shootout'"
    ).fetchall()
    conn.close()
    out = defaultdict(list)
    for r in rows:
        # Register the category itself even when this particular row has no
        # system value — a category whose goals ALL happen to have a blank
        # system was previously dropped from the dropdown entirely, because
        # `out[r["cat"]]` was only ever touched inside the `if r["sys"]`
        # check below, so a category with zero non-blank systems never got
        # a key created for it at all.
        _ = out[r["cat"]]
        if r["sys"] and r["sys"] not in out[r["cat"]]:
            out[r["cat"]].append(r["sys"])
    for cat in out:
        out[cat].sort()
    return out


@app.get("/api/goal-categories/cross-search")
def gc_cross_search(
    gf_cat: str | None = Query(None, alias="gfCat"),
    gf_sys: str | None = Query(None, alias="gfSys"),
    ga_cat: str | None = Query(None, alias="gaCat"),
    ga_sys: str | None = Query(None, alias="gaSys"),
    team: str | None = Query(None),
    date_from: str | None = Query(None, alias="from"),
    date_to: str | None = Query(None, alias="to"),
    user: CurrentUser = Depends(require_login),
):
    """Mix-and-match filter across GF Category/System and GA Category/System
    — same tool as the client-side cross-search, now reading straight from
    the database. When `team` is set, scoped to that team's own goals-for
    (gaCat/gaSys still describe the specific opponent's breakdown on those
    same goals) — same framing as the original."""
    conn = get_db()
    team = team.upper() if team else None

    games = conn.execute("SELECT gameReferenceId, gameDate, homeTeamAbbrev, awayTeamAbbrev FROM games").fetchall()
    games = [g for g in games if _gc_in_scope(g["gameDate"], "all", date_from, date_to)]
    valid_gids = {g["gameReferenceId"] for g in games}
    gid_to_date = {g["gameReferenceId"]: g["gameDate"] for g in games}

    goal_rows = conn.execute(
        "SELECT gameReferenceId, scoringTeamAbbrev, concedingTeamAbbrev, gfCat, gfSys, gaCat, gaSys FROM goal_categories"
    ).fetchall()
    conn.close()

    matches, by_team, gf_breakdown, ga_breakdown = [], defaultdict(int), defaultdict(int), defaultdict(int)
    for gr in goal_rows:
        if gr["gameReferenceId"] not in valid_gids:
            continue
        if gr["gfCat"] == "Shootout" or gr["gaCat"] == "Shootout":
            continue
        if team and gr["scoringTeamAbbrev"] != team:
            continue
        if gf_cat and gr["gfCat"] != gf_cat:
            continue
        if gf_sys and gr["gfSys"] != gf_sys:
            continue
        if ga_cat and gr["gaCat"] != ga_cat:
            continue
        if ga_sys and gr["gaSys"] != ga_sys:
            continue
        # The frontend's results table reads `date`, `scoringTeam`, and
        # `concedingTeam` specifically — the raw goal_categories row has
        # neither a date column at all (it lives on `games`, joined in
        # here) nor those exact key names (it has *Abbrev suffixes), so
        # without this the table rendered "undefined" for every row.
        m = dict(gr)
        m["date"] = gid_to_date.get(gr["gameReferenceId"])
        m["scoringTeam"] = gr["scoringTeamAbbrev"]
        m["concedingTeam"] = gr["concedingTeamAbbrev"]
        matches.append(m)
        by_team[gr["scoringTeamAbbrev"]] += 1
        gf_breakdown[f"{gr['gfCat']} — {gr['gfSys'] or '(none)'}"] += 1
        ga_breakdown[f"{gr['gaCat']} — {gr['gaSys'] or '(none)'}"] += 1

    return {
        "total": len(matches),
        "matches": sorted(matches, key=lambda m: m["gameReferenceId"] or 0),
        "byTeam": sorted([{"team": t, "count": c} for t, c in by_team.items()], key=lambda x: -x["count"]),
        "gfBreakdown": sorted([{"key": k, "count": c} for k, c in gf_breakdown.items()], key=lambda x: -x["count"]),
        "gaBreakdown": sorted([{"key": k, "count": c} for k, c in ga_breakdown.items()], key=lambda x: -x["count"]),
    }


# ── PSF game upload ───────────────────────────────────────────────────────
# The browser-based replacement for running `load_nhl.py --src ... --db ...`
# by hand for each new game. Reuses load_nhl.py's own parsing/loading
# functions directly (same file already sitting alongside api.py) rather
# than re-implementing that logic a second time — one copy of the actual
# PSF-parsing rules, called from both the command line and this endpoint.
import load_nhl
import tempfile

@app.post("/api/games/upload")
async def upload_psf_game(file: UploadFile = File(...), user: CurrentUser = Depends(require_admin)):
    """Accepts one game's PSF CSV — same file you'd otherwise point
    load_nhl.py at locally — and loads it into the database. Safe to
    re-upload the same game (load_nhl's own load_file replaces that
    game's rows rather than duplicating them, same as running the script
    twice locally already does).

    Filename matters: load_nhl.py pulls the date/teams/season/game ID out
    of the filename itself (playsequence-YYYYMMDD-LEAGUE-AWAYvsHOME-SEASON-
    GAMEID.csv), so this only works correctly if the uploaded file still
    has its original SportLogiq filename — same requirement as using the
    script locally, not something new introduced by this endpoint."""
    if not file.filename.endswith(".csv"):
        raise HTTPException(400, "Expected a .csv file (the original PSF playsequence export)")

    raw = await file.read()
    # Write to a temp path that PRESERVES the original filename, since
    # load_nhl's parse_filename() reads metadata out of the filename itself.
    tmp_dir = tempfile.mkdtemp()
    tmp_path = Path(tmp_dir) / file.filename
    tmp_path.write_bytes(raw)

    conn = get_db()
    try:
        meta = load_nhl.parse_filename(str(tmp_path))
        if meta.get("gameReferenceId") is None:
            raise HTTPException(
                400,
                f"Couldn't extract a game ID from filename '{file.filename}'. "
                "Expected pattern: playsequence-YYYYMMDD-LEAGUE-AWAYvsHOME-SEASON-GAMEID.csv "
                "— check this is the original, unrenamed file from SportLogiq.",
            )
        n_plays = load_nhl.load_file(conn, str(tmp_path))
        if n_plays == 0:
            raise HTTPException(400, f"'{file.filename}' parsed with a valid filename but contained zero play rows — is this an empty or corrupt file?")
        conn.commit()
        load_nhl.rebuild_players(conn)
        conn.commit()
        game_row = conn.execute(
            "SELECT gameReferenceId, gameDate, homeTeamAbbrev, awayTeamAbbrev, matchup FROM games "
            "WHERE sourceFile = ? ORDER BY gameReferenceId DESC LIMIT 1",
            (file.filename,),
        ).fetchone()

        # Derive this game's faceoff draws at upload time, so the Faceoffs tab
        # never needs an on-demand pass over the play-by-play. Handedness is a
        # league-wide majority vote, so it re-resolves after each new game.
        faceoff_draws = None
        if game_row:
            try:
                gid = game_row["gameReferenceId"]
                faceoff_draws = fo_build_game(
                    conn, gid, _fo_playoff_game_ids(conn, meta.get("season")))
                fo_rebuild_hands(conn)
                conn.commit()
            except Exception as e:
                # A PSF load succeeding matters more than the faceoff
                # derivation succeeding — report it, don't fail the upload.
                print(f"[upload] faceoff derivation failed for {file.filename}: {e}")
                faceoff_draws = f"failed: {e}"
    finally:
        conn.close()
        tmp_path.unlink(missing_ok=True)
        Path(tmp_dir).rmdir()

    return {
        "filename": file.filename,
        "playsLoaded": n_plays,
        "game": dict(game_row) if game_row else None,
        "faceoffDraws": faceoff_draws,
    }


# ============================================================================
# Shootout Scout — self-sufficient port of Erik's nhl_shootout_pipeline.py
# ============================================================================
# Originally a standalone site Erik maintained by hand: run the pipeline
# script against the NHL API, ship fresh players.json/goalies.json/
# alltime.json/so_order.json. Ported in here so Coaches Station owns the
# refresh itself. shootout.db (seeded from Erik's original database, which
# already has the full 2005-06—present history he built) ships on the
# persistent volume alongside nhl.db — see SHOOTOUT_DB_PATH below.
#
# /api/shootout/refresh pulls straight from the NHL API (same endpoints
# the old pipeline used) and is meant to be called periodically (a cron
# hitting it, or a manual "Refresh" button) rather than requiring anyone
# to run a script by hand. The four GET endpoints below compute the same
# shapes the old JSON exports used to, live from these tables, so the
# Shootout tab's frontend just fetches from here now instead of reading a
# static snapshot.

SHOOTOUT_DB_PATH = Path(os.environ.get("SHOOTOUT_DB_PATH", Path(__file__).parent / "shootout.db"))

SHOOTOUT_SCHEMA = """
CREATE TABLE IF NOT EXISTS skater_shootout (
    player_id   INTEGER NOT NULL,
    full_name   TEXT NOT NULL,
    team_abbrev TEXT,
    season      TEXT NOT NULL,
    goals       INTEGER DEFAULT 0,
    attempts    INTEGER DEFAULT 0,
    PRIMARY KEY (player_id, season)
);
CREATE TABLE IF NOT EXISTS goalie_shootout (
    goalie_id   INTEGER NOT NULL,
    full_name   TEXT NOT NULL,
    team_abbrev TEXT,
    season      TEXT NOT NULL,
    saves       INTEGER DEFAULT 0,
    shots_against INTEGER DEFAULT 0,
    wins        INTEGER DEFAULT 0,
    losses      INTEGER DEFAULT 0,
    PRIMARY KEY (goalie_id, season)
);
CREATE TABLE IF NOT EXISTS vs_goalie_splits (
    player_id   INTEGER NOT NULL,
    goalie_id   INTEGER NOT NULL,
    goalie_name TEXT,
    goals       INTEGER DEFAULT 0,
    attempts    INTEGER DEFAULT 0,
    PRIMARY KEY (player_id, goalie_id)
);
CREATE TABLE IF NOT EXISTS so_attempts (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    game_id     INTEGER NOT NULL,
    season      TEXT NOT NULL,
    game_date   TEXT,
    home_team   TEXT,
    away_team   TEXT,
    round_num   INTEGER,
    shooter_id  INTEGER NOT NULL,
    goalie_id   INTEGER NOT NULL,
    result      TEXT NOT NULL,
    shot_type   TEXT,
    miss_reason TEXT,
    UNIQUE(game_id, shooter_id, round_num)
);
CREATE INDEX IF NOT EXISTS idx_soa_shooter ON so_attempts(shooter_id);
CREATE INDEX IF NOT EXISTS idx_soa_goalie  ON so_attempts(goalie_id);
CREATE INDEX IF NOT EXISTS idx_soa_season  ON so_attempts(season);
CREATE TABLE IF NOT EXISTS active_rosters (
    player_id   INTEGER PRIMARY KEY,
    full_name   TEXT,
    team_abbrev TEXT NOT NULL,
    position    TEXT,
    is_goalie   INTEGER DEFAULT 0,
    jersey_number INTEGER
);
CREATE TABLE IF NOT EXISTS shootout_refresh_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ran_at TEXT NOT NULL,
    season TEXT,
    new_games INTEGER,
    note TEXT
);
"""


def get_shootout_db():
    conn = sqlite3.connect(SHOOTOUT_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


@app.on_event("startup")
def ensure_shootout_schema():
    """Same idempotent pattern as ensure_schema() above — safe to run on
    every startup, and this is what turns a freshly-copied shootout.db (or
    a brand new empty one) into something these endpoints can use."""
    SHOOTOUT_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(SHOOTOUT_DB_PATH)
    conn.executescript(SHOOTOUT_SCHEMA)
    conn.commit()
    conn.close()


def _shootout_current_season() -> str:
    """Same Oct-June season-string convention used elsewhere in this file."""
    today = dt.date.today()
    if today.month >= 9:
        return f"{today.year}{today.year + 1}"
    return f"{today.year - 1}{today.year}"


async def _so_fetch_shootout_stats(client: httpx.AsyncClient, season: str, report_type: str) -> list:
    """Pulls the NHL stats API's dedicated shootout report for one season.
    gameTypeId=2 = regular season only. Paginates (100 rows/page cap)."""
    base = f"{NHL_STATS_API_BASE}/stats/rest/en"
    endpoint = "skater" if report_type == "skater" else "goalie"
    all_rows = []
    start, limit = 0, 100
    while True:
        url = (
            f"{base}/{endpoint}/shootout"
            f"?isAggregate=false&isGame=false"
            f"&cayenneExp=seasonId={season}%20and%20gameTypeId=2"
            f"&start={start}&limit={limit}"
        )
        try:
            r = await client.get(url, timeout=20.0)
            r.raise_for_status()
            data = r.json()
        except Exception:
            break
        rows = data.get("data", [])
        all_rows.extend(rows)
        total = data.get("total", 0)
        start += limit
        if start >= total:
            break
    return all_rows


async def _so_refresh_season_aggregates(client: httpx.AsyncClient, conn: sqlite3.Connection, season: str) -> dict:
    """Upserts this season's skater/goalie shootout totals. Cheap — a
    handful of paginated requests, no per-game scanning."""
    skater_rows = await _so_fetch_shootout_stats(client, season, "skater")
    for r in skater_rows:
        pid = r.get("playerId")
        if not pid:
            continue
        conn.execute(
            """
            INSERT INTO skater_shootout (player_id, full_name, team_abbrev, season, goals, attempts)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(player_id, season) DO UPDATE SET
                full_name=excluded.full_name, team_abbrev=excluded.team_abbrev,
                goals=excluded.goals, attempts=excluded.attempts
            """,
            (pid, r.get("skaterFullName", f"Player #{pid}"), r.get("teamAbbrevs", ""), season,
             r.get("shootoutGoals", 0), r.get("shootoutShots", 0)),
        )

    goalie_rows = await _so_fetch_shootout_stats(client, season, "goalie")
    for r in goalie_rows:
        gid = r.get("playerId")
        if not gid:
            continue
        conn.execute(
            """
            INSERT INTO goalie_shootout (goalie_id, full_name, team_abbrev, season, saves, shots_against, wins, losses)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(goalie_id, season) DO UPDATE SET
                full_name=excluded.full_name, team_abbrev=excluded.team_abbrev,
                saves=excluded.saves, shots_against=excluded.shots_against,
                wins=excluded.wins, losses=excluded.losses
            """,
            (gid, r.get("goalieFullName", f"Goalie #{gid}"), r.get("teamAbbrevs", ""), season,
             r.get("shootoutSaves", 0), r.get("shootoutShotsAgainst", 0),
             r.get("shootoutWins", 0), r.get("shootoutLosses", 0)),
        )
    conn.commit()
    return {"skaters": len(skater_rows), "goalies": len(goalie_rows)}


_SHOOTOUT_TEAM_ABBREVS = [
    "ANA", "BOS", "BUF", "CGY", "CAR", "CHI", "COL", "CBJ", "DAL", "DET",
    "EDM", "FLA", "LAK", "MIN", "MTL", "NSH", "NJD", "NYI", "NYR", "OTT",
    "PHI", "PIT", "SJS", "SEA", "STL", "TBL", "TOR", "UTA", "VAN", "VGK", "WSH", "WPG",
]


async def _so_update_rosters(client: httpx.AsyncClient, conn: sqlite3.Connection, season: str) -> dict:
    """Refreshes active_rosters team-by-team from current NHL rosters —
    only players in this table show up in the Shootout tab's team panels;
    everyone else is still searchable/all-time from history alone.

    Deliberately NOT a wipe-then-rebuild: each team's existing rows are
    only cleared once that team's own fetch has actually succeeded, so a
    bad run (rate-limited, NHL API hiccup, etc.) can never leave the whole
    table empty — it just leaves whichever teams failed with their
    previous data intact. A short delay between requests mirrors the
    original pipeline's pacing and avoids tripping a rate limit that a
    tight, un-paced loop of 32 requests risked hitting."""
    ok_teams, failed_teams = [], []
    for team in _SHOOTOUT_TEAM_ABBREVS:
        try:
            r = await client.get(f"{NHL_API_BASE}/roster/{team}/current", timeout=15.0)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            failed_teams.append(f"{team}: {type(e).__name__}: {e}")
            await asyncio.sleep(0.3)
            continue
        conn.execute("DELETE FROM active_rosters WHERE team_abbrev=?", (team,))
        for group, is_goalie in (("forwards", 0), ("defensemen", 0), ("goalies", 1)):
            for p in data.get(group, []):
                pid = p["id"]
                name = f"{p['firstName']['default']} {p['lastName']['default']}"
                pos = p.get("positionCode", "")
                num = p.get("sweaterNumber")
                conn.execute(
                    """
                    INSERT OR REPLACE INTO active_rosters
                        (player_id, full_name, team_abbrev, position, is_goalie, jersey_number)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (pid, name, team, pos, is_goalie, num),
                )
                table = "goalie_shootout" if is_goalie else "skater_shootout"
                id_col = "goalie_id" if is_goalie else "player_id"
                conn.execute(
                    f"UPDATE {table} SET team_abbrev=? WHERE {id_col}=? AND season=?",
                    (team, pid, season),
                )
        conn.commit()
        ok_teams.append(team)
        await asyncio.sleep(0.3)
    return {"teamsUpdated": ok_teams, "teamsFailed": failed_teams}


async def _so_get_new_shootout_game_ids(client: httpx.AsyncClient, season: str, already: set) -> list:
    """Scans this season's schedule week-by-week for games that went to a
    shootout (lastPeriodType == 'SO'), skipping any game_id already
    processed. Bounded to the season's Oct 1 – Jun 30 window."""
    year1, year2 = int(season[:4]), int(season[4:])
    cursor = dt.date(year1, 10, 1)
    end = dt.date(year2, 6, 30)
    found = []
    while cursor <= end:
        try:
            r = await client.get(f"{NHL_API_BASE}/schedule/{cursor.isoformat()}", timeout=15.0)
            r.raise_for_status()
            data = r.json()
        except Exception:
            cursor += dt.timedelta(days=7)
            continue
        for day in data.get("gameWeek", []):
            for g in day.get("games", []):
                if g.get("gameType") != 2:
                    continue
                if g.get("gameState") not in ("OFF", "FINAL"):
                    continue
                outcome = g.get("gameOutcome") or {}
                if outcome.get("lastPeriodType") == "SO" and g["id"] not in already:
                    found.append(g["id"])
        cursor += dt.timedelta(days=7)
    return found


async def _so_get_splits_from_pbp(client: httpx.AsyncClient, game_id: int):
    """Per-game play-by-play parse — shooter/goalie pairs, shot type, miss
    reason, round number. Direct port of the pipeline's get_splits_from_pbp,
    adapted to httpx async. Returns (pairs, game_date, home_abbrev, away_abbrev)."""
    try:
        r = await client.get(f"{NHL_API_BASE}/gamecenter/{game_id}/play-by-play", timeout=15.0)
        r.raise_for_status()
        pbp = r.json()
    except Exception:
        return [], None, None, None

    game_date = pbp.get("gameDate")
    home_abbrev = pbp.get("homeTeam", {}).get("abbrev")
    away_abbrev = pbp.get("awayTeam", {}).get("abbrev")
    home_id = pbp.get("homeTeam", {}).get("id")
    away_id = pbp.get("awayTeam", {}).get("id")

    so_plays = []
    team_goalie = {}
    for play in pbp.get("plays", []):
        if play.get("periodDescriptor", {}).get("periodType") != "SO":
            continue
        d = play.get("details", {})
        type_key = play.get("typeDescKey", "")
        shooter_id = d.get("shootingPlayerId") or d.get("scoringPlayerId")
        goalie_id = d.get("goalieInNetId")
        owner_tid = d.get("eventOwnerTeamId")
        shot_type = d.get("shotType")
        miss_reason = d.get("reason")
        if type_key not in ("shot-on-goal", "goal", "missed-shot", "failed-shot-attempt"):
            continue
        if not shooter_id:
            continue
        so_plays.append((shooter_id, goalie_id, type_key, owner_tid, shot_type, miss_reason))
        if goalie_id and owner_tid:
            team_goalie[owner_tid] = goalie_id

    if len(team_goalie) < 2:
        so_goalie_ids = {gid for _, gid, _, _, _, _ in so_plays if gid}
        for spot in pbp.get("rosterSpots", []):
            if spot.get("positionCode") != "G":
                continue
            pid = spot.get("playerId")
            tid = spot.get("teamId")
            if pid not in so_goalie_ids:
                continue
            other_tid = away_id if tid == home_id else home_id
            if other_tid and other_tid not in team_goalie:
                team_goalie[other_tid] = pid

    if len(team_goalie) < 2:
        try:
            g_url = (
                f"{NHL_STATS_API_BASE}/stats/rest/en/goalie/shootout"
                f"?isAggregate=false&isGame=true&cayenneExp=gameId={game_id}&start=0&limit=10"
            )
            g = await client.get(g_url, timeout=15.0)
            g.raise_for_status()
            g_data = g.json()
            for row in g_data.get("data", []):
                gid = row.get("playerId")
                wins = row.get("shootoutWins", 0)
                loss = row.get("shootoutLosses", 0)
                if not gid or (wins == 0 and loss == 0):
                    continue
                goalie_team_id = next(
                    (s.get("teamId") for s in pbp.get("rosterSpots", [])
                     if s.get("playerId") == gid and s.get("positionCode") == "G"),
                    None,
                )
                if goalie_team_id:
                    shooting_tid = away_id if goalie_team_id == home_id else home_id
                    if shooting_tid and shooting_tid not in team_goalie:
                        team_goalie[shooting_tid] = gid
        except Exception:
            pass

    pairs = []
    round_counter = {}
    for shooter_id, goalie_id, type_key, owner_tid, shot_type, miss_reason in so_plays:
        if not goalie_id and owner_tid:
            goalie_id = team_goalie.get(owner_tid)
        if not shooter_id or not goalie_id:
            continue
        round_counter[owner_tid] = round_counter.get(owner_tid, 0) + 1
        result = "goal" if type_key == "goal" else ("miss" if type_key in ("missed-shot", "failed-shot-attempt") else "save")
        pairs.append((shooter_id, goalie_id, result, shot_type, miss_reason, round_counter[owner_tid]))
    return pairs, game_date, home_abbrev, away_abbrev


async def _so_process_new_games(client: httpx.AsyncClient, conn: sqlite3.Connection, season: str, game_ids: list) -> int:
    goalie_names = {}
    for row in conn.execute("SELECT goalie_id, full_name FROM goalie_shootout GROUP BY goalie_id"):
        goalie_names[row["goalie_id"]] = row["full_name"]
    for row in conn.execute("SELECT player_id, full_name FROM active_rosters WHERE is_goalie=1"):
        goalie_names[row["player_id"]] = row["full_name"]

    processed = 0
    for game_id in game_ids:
        pairs, game_date, home_abbrev, away_abbrev = await _so_get_splits_from_pbp(client, game_id)
        for shooter_id, goalie_id, result, shot_type, miss_reason, round_num in pairs:
            conn.execute(
                """
                INSERT OR IGNORE INTO so_attempts
                    (game_id, season, game_date, home_team, away_team, round_num, shooter_id, goalie_id, result, shot_type, miss_reason)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (game_id, season, game_date, home_abbrev, away_abbrev, round_num, shooter_id, goalie_id, result, shot_type, miss_reason),
            )
        if pairs:
            processed += 1
        conn.commit()

    # Rebuild vs_goalie_splits from so_attempts — pure local aggregation,
    # no further NHL API calls needed.
    conn.execute("DELETE FROM vs_goalie_splits")
    for row in conn.execute(
        """
        SELECT shooter_id, goalie_id,
               SUM(CASE WHEN result='goal' THEN 1 ELSE 0 END) goals,
               COUNT(*) attempts
        FROM so_attempts GROUP BY shooter_id, goalie_id
        """
    ):
        gname = goalie_names.get(row["goalie_id"], f"Goalie #{row['goalie_id']}")
        conn.execute(
            """
            INSERT INTO vs_goalie_splits (player_id, goalie_id, goalie_name, goals, attempts)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(player_id, goalie_id) DO UPDATE SET
                goalie_name=excluded.goalie_name, goals=excluded.goals, attempts=excluded.attempts
            """,
            (row["shooter_id"], row["goalie_id"], gname, row["goals"], row["attempts"]),
        )
    conn.commit()
    return processed


@app.post("/api/shootout/refresh")
async def shootout_refresh(
    season: str = Query(None, description="e.g. 20252026; defaults to the current season"),
    max_new_games: int = Query(60, ge=1, le=200, description="Caps how many new shootout games get processed in one call"),
    user: CurrentUser = Depends(require_admin),
):
    """Replaces hand-running nhl_shootout_pipeline.py. Refreshes this
    season's skater/goalie aggregate totals (a handful of requests),
    rebuilds active rosters, then scans the schedule for shootout games not
    yet in so_attempts and processes up to max_new_games of them. If
    remaining > 0 in the response, call again to pick up the rest — this
    stays bounded per call on purpose rather than risking one huge
    synchronous request timing out mid-season-backfill."""
    season = season or _shootout_current_season()
    conn = get_shootout_db()
    try:
        async with httpx.AsyncClient(follow_redirects=True) as client:
            agg = await _so_refresh_season_aggregates(client, conn, season)
            roster_result = await _so_update_rosters(client, conn, season)
            already = {row[0] for row in conn.execute(
                "SELECT DISTINCT game_id FROM so_attempts WHERE season=?", (season,)
            )}
            new_ids = await _so_get_new_shootout_game_ids(client, season, already)
            to_process = new_ids[:max_new_games]
            processed = await _so_process_new_games(client, conn, season, to_process)
        conn.execute(
            "INSERT INTO shootout_refresh_log (ran_at, season, new_games, note) VALUES (?, ?, ?, ?)",
            (dt.datetime.utcnow().isoformat(), season, processed,
             f"{len(new_ids)} new games found, {len(to_process)} processed; "
             f"rosters: {len(roster_result['teamsUpdated'])} ok, "
             f"{len(roster_result['teamsFailed'])} failed {roster_result['teamsFailed']}"),
        )
        conn.commit()
    finally:
        conn.close()
    return {
        "season": season,
        "skaterRowsRefreshed": agg["skaters"],
        "goalieRowsRefreshed": agg["goalies"],
        "rostersUpdated": roster_result["teamsUpdated"],
        "rostersFailed": roster_result["teamsFailed"],
        "newGamesFound": len(new_ids),
        "newGamesProcessed": processed,
        "remaining": max(0, len(new_ids) - len(to_process)),
    }


@app.get("/api/shootout/last-refresh")
def shootout_last_refresh(user: CurrentUser = Depends(require_login)):
    conn = get_shootout_db()
    row = conn.execute("SELECT * FROM shootout_refresh_log ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return dict(row) if row else {"ran_at": None}


def _so_last_refresh_str(conn: sqlite3.Connection) -> str | None:
    row = conn.execute("SELECT ran_at FROM shootout_refresh_log ORDER BY id DESC LIMIT 1").fetchone()
    return row["ran_at"] if row else None


@app.get("/api/shootout/players")
def shootout_players(response: Response, user: CurrentUser = Depends(require_login)):
    """Live equivalent of the old players.json export — shooter career/
    season/vs-goalie splits, shot-type breakdown, round performance."""
    conn = get_shootout_db()
    last_refresh = _so_last_refresh_str(conn)
    if last_refresh:
        response.headers["x-shootout-last-refresh"] = last_refresh

    active = {}
    for row in conn.execute("SELECT player_id, full_name, team_abbrev, is_goalie, jersey_number FROM active_rosters"):
        active[row["player_id"]] = {
            "team": row["team_abbrev"], "name": row["full_name"],
            "is_goalie": row["is_goalie"], "jersey_number": row["jersey_number"],
        }

    players_map = {}
    for pid, ar in active.items():
        if ar["is_goalie"]:
            continue
        players_map[pid] = {
            "id": pid, "name": ar["name"], "team": ar["team"], "number": ar.get("jersey_number"),
            "active": True, "career": [0, 0], "seasons": {}, "vs_goalie": {},
        }

    for row in conn.execute("SELECT player_id, full_name, team_abbrev, season, goals, attempts FROM skater_shootout ORDER BY season"):
        pid = row["player_id"]
        if pid not in players_map:
            players_map[pid] = {
                "id": pid, "name": row["full_name"], "team": "", "active": False,
                "career": [0, 0], "seasons": {}, "vs_goalie": {},
            }
        p = players_map[pid]
        p["career"][0] += row["goals"]
        p["career"][1] += row["attempts"]
        p["seasons"][row["season"]] = [row["goals"], row["attempts"]]

    for row in conn.execute("SELECT player_id, goalie_name, goals, attempts FROM vs_goalie_splits"):
        pid = row["player_id"]
        if pid in players_map:
            players_map[pid]["vs_goalie"][row["goalie_name"]] = [row["goals"], row["attempts"]]

    for row in conn.execute(
        """
        SELECT sa.shooter_id, gs.goalie_name, sa.shot_type,
               SUM(CASE WHEN sa.result='goal' THEN 1 ELSE 0 END) goals, COUNT(*) attempts
        FROM so_attempts sa
        LEFT JOIN vs_goalie_splits gs ON gs.player_id=sa.shooter_id AND gs.goalie_id=sa.goalie_id
        GROUP BY sa.shooter_id, sa.goalie_id, sa.shot_type
        """
    ):
        pid = row["shooter_id"]
        gname = row["goalie_name"]
        if pid not in players_map or not gname:
            continue
        p = players_map[pid]
        p.setdefault("vs_goalie_shots", {}).setdefault(gname, {})
        st = row["shot_type"] or "unknown"
        p["vs_goalie_shots"][gname][st] = [row["goals"], row["attempts"]]

    for row in conn.execute("SELECT shooter_id, shot_type, result, round_num, miss_reason FROM so_attempts"):
        pid = row["shooter_id"]
        if pid not in players_map:
            continue
        p = players_map[pid]
        st = row["shot_type"] or "unknown"
        p.setdefault("shot_types", {}).setdefault(st, [0, 0])
        if row["result"] == "goal":
            p["shot_types"][st][0] += 1
        p["shot_types"][st][1] += 1

        rn = str(row["round_num"] or "?")
        p.setdefault("by_round", {}).setdefault(rn, [0, 0])
        if row["result"] == "goal":
            p["by_round"][rn][0] += 1
        p["by_round"][rn][1] += 1

        if row["result"] == "miss" and row["miss_reason"]:
            p.setdefault("miss_reasons", {})
            p["miss_reasons"][row["miss_reason"]] = p["miss_reasons"].get(row["miss_reason"], 0) + 1

    conn.close()
    players_out = [p for p in players_map.values() if p["active"] or p["career"][1] > 0]
    players_out.sort(key=lambda p: (not p["active"], -p["career"][1]))
    return players_out


@app.get("/api/shootout/goalies")
def shootout_goalies(response: Response, user: CurrentUser = Depends(require_login)):
    """Live equivalent of the old goalies.json export."""
    conn = get_shootout_db()
    last_refresh = _so_last_refresh_str(conn)
    if last_refresh:
        response.headers["x-shootout-last-refresh"] = last_refresh

    active = {}
    for row in conn.execute("SELECT player_id, full_name, team_abbrev, is_goalie, jersey_number FROM active_rosters"):
        active[row["player_id"]] = {
            "team": row["team_abbrev"], "name": row["full_name"],
            "is_goalie": row["is_goalie"], "jersey_number": row["jersey_number"],
        }

    goalies_map = {}
    for pid, ar in active.items():
        if not ar["is_goalie"]:
            continue
        goalies_map[pid] = {
            "id": pid, "name": ar["name"], "team": ar["team"], "number": ar.get("jersey_number"),
            "active": True, "stopped": 0, "faced": 0, "wins": 0, "losses": 0,
        }

    for row in conn.execute("SELECT goalie_id, full_name, team_abbrev, season, saves, shots_against, wins, losses FROM goalie_shootout ORDER BY season"):
        gid = row["goalie_id"]
        if gid not in goalies_map:
            goalies_map[gid] = {
                "id": gid, "name": row["full_name"], "team": "", "active": False,
                "stopped": 0, "faced": 0, "wins": 0, "losses": 0,
            }
        g = goalies_map[gid]
        g["stopped"] += row["saves"]
        g["faced"] += row["shots_against"]
        g["wins"] += row["wins"]
        g["losses"] += row["losses"]

    for g in goalies_map.values():
        g["record"] = f"{g['wins']} W \u2013 {g['losses']} L in career shootouts"

    for row in conn.execute("SELECT goalie_id, shot_type, result, round_num, miss_reason FROM so_attempts"):
        gid = row["goalie_id"]
        if gid not in goalies_map:
            continue
        g = goalies_map[gid]
        st = row["shot_type"] or "unknown"
        g.setdefault("shots_by_type", {}).setdefault(st, [0, 0])
        if row["result"] != "goal":
            g["shots_by_type"][st][0] += 1
        g["shots_by_type"][st][1] += 1

        rn = str(row["round_num"] or "?")
        g.setdefault("by_round", {}).setdefault(rn, [0, 0])
        if row["result"] != "goal":
            g["by_round"][rn][0] += 1
        g["by_round"][rn][1] += 1

        if row["result"] == "miss" and row["miss_reason"]:
            g.setdefault("miss_reasons", {})
            g["miss_reasons"][row["miss_reason"]] = g["miss_reasons"].get(row["miss_reason"], 0) + 1

    conn.close()
    return [g for g in goalies_map.values() if g["active"] or g["faced"] > 0]


@app.get("/api/shootout/alltime")
def shootout_alltime(
    response: Response,
    top_n: int = Query(25, ge=1, le=100),
    min_attempts: int = Query(15, ge=1),
    min_faced: int = Query(20, ge=1),
    user: CurrentUser = Depends(require_login),
):
    """Live equivalent of the old alltime.json export."""
    conn = get_shootout_db()
    last_refresh = _so_last_refresh_str(conn)
    if last_refresh:
        response.headers["x-shootout-last-refresh"] = last_refresh

    def query(sql, params=()):
        cur = conn.execute(sql, params)
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]

    top_goals = query(
        f"""SELECT player_id, full_name AS name, team_abbrev AS team,
                   SUM(goals) AS goals, SUM(attempts) AS att
            FROM skater_shootout GROUP BY player_id ORDER BY goals DESC LIMIT {top_n}"""
    )
    top_pct = query(
        f"""SELECT player_id, full_name AS name, team_abbrev AS team,
                   SUM(goals) AS goals, SUM(attempts) AS att
            FROM skater_shootout GROUP BY player_id
            HAVING att >= {min_attempts}
            ORDER BY (CAST(SUM(goals) AS FLOAT)/SUM(attempts)) DESC LIMIT {top_n}"""
    )
    top_sv = query(
        f"""SELECT goalie_id, full_name AS name, team_abbrev AS team,
                   SUM(saves) AS stopped, SUM(shots_against) AS faced
            FROM goalie_shootout GROUP BY goalie_id
            HAVING faced >= {min_faced}
            ORDER BY (CAST(SUM(saves) AS FLOAT)/SUM(shots_against)) DESC LIMIT {top_n}"""
    )
    most_faced = query(
        f"""SELECT goalie_id, full_name AS name, team_abbrev AS team,
                   SUM(saves) AS stopped, SUM(shots_against) AS faced
            FROM goalie_shootout GROUP BY goalie_id ORDER BY faced DESC LIMIT {top_n}"""
    )
    conn.close()
    return {
        "top_goals": top_goals, "top_shooting_pct": top_pct,
        "top_goalie_save_pct": top_sv, "most_shots_faced": most_faced,
        "min_attempts_threshold": min_attempts, "min_faced_threshold": min_faced,
    }


@app.get("/api/shootout/so-order")
def shootout_so_order(response: Response, user: CurrentUser = Depends(require_login)):
    """Live equivalent of the old so_order.json export — per-team,
    per-season shootout order/participation history. Direct port of
    export_so_order's majority-vote goalie-team assignment logic."""
    conn = get_shootout_db()
    last_refresh = _so_last_refresh_str(conn)
    if last_refresh:
        response.headers["x-shootout-last-refresh"] = last_refresh

    game_info = {}
    for row in conn.execute(
        "SELECT DISTINCT game_id, game_date, home_team, away_team FROM so_attempts WHERE home_team IS NOT NULL AND away_team IS NOT NULL"
    ):
        game_info[row["game_id"]] = {"date": row["game_date"], "home": row["home_team"], "away": row["away_team"]}

    rows = conn.execute(
        """
        SELECT sa.game_id, sa.season, sa.game_date, sa.home_team, sa.away_team,
               COALESCE(ss.team_abbrev, ar.team_abbrev) AS fallback_team,
               COALESCE(
                   (SELECT sk.full_name FROM skater_shootout sk WHERE sk.player_id=sa.shooter_id LIMIT 1),
                   ar.full_name, 'Player #'||sa.shooter_id
               ) AS shooter_name,
               ar.jersey_number,
               sa.shooter_id, sa.goalie_id, sa.round_num, sa.result, sa.shot_type, sa.miss_reason
        FROM so_attempts sa
        LEFT JOIN skater_shootout ss ON ss.player_id=sa.shooter_id AND ss.season=sa.season
        LEFT JOIN active_rosters ar ON ar.player_id=sa.shooter_id
        ORDER BY sa.season DESC, sa.game_id, sa.round_num
        """
    ).fetchall()
    conn.close()

    game_attempts = defaultdict(list)
    for r in rows:
        game_attempts[r["game_id"]].append(r)

    out = defaultdict(lambda: defaultdict(list))
    for gid, attempts in game_attempts.items():
        info = game_info.get(gid)
        has_home_away = bool(info and info.get("home") and info.get("away"))
        season = attempts[0]["season"]

        if has_home_away:
            home, away = info["home"], info["away"]
            teams = {home, away}
            date = info.get("date")
        else:
            teams = {r["fallback_team"] for r in attempts if r["fallback_team"]}
            home, away = None, None
            date = None

        if len(teams) < 2:
            continue

        shooter_teams = {}
        if has_home_away:
            goalie_to_rows = defaultdict(list)
            for r in attempts:
                goalie_to_rows[r["goalie_id"]].append(r)
            goalie_team_in_game = {}
            if len(goalie_to_rows) == 2:
                for goalie_id, g_rows in goalie_to_rows.items():
                    home_count = sum(1 for r in g_rows if r["fallback_team"] == home)
                    away_count = sum(1 for r in g_rows if r["fallback_team"] == away)
                    if home_count > away_count:
                        goalie_team_in_game[goalie_id] = away
                    elif away_count > home_count:
                        goalie_team_in_game[goalie_id] = home
            for r in attempts:
                goalie_team = goalie_team_in_game.get(r["goalie_id"])
                if goalie_team:
                    shooter_teams[r["shooter_id"]] = away if goalie_team == home else home
                else:
                    ft = r["fallback_team"]
                    shooter_teams[r["shooter_id"]] = ft if ft in (home, away) else None
        else:
            for r in attempts:
                shooter_teams[r["shooter_id"]] = r["fallback_team"]

        goals = [r for r in attempts if r["result"] == "goal"]
        winner = shooter_teams.get(goals[-1]["shooter_id"]) if goals else None

        for team in teams:
            opponent = next((t for t in teams if t != team), "?")
            outcome = "W" if winner == team else ("L" if winner else "?")
            team_attempts = [
                {
                    "round": r["round_num"], "shooter": r["shooter_name"], "shooter_id": r["shooter_id"],
                    "number": r["jersey_number"], "result": r["result"],
                    "shot_type": r["shot_type"] or "unknown", "miss_reason": r["miss_reason"],
                }
                for r in attempts if shooter_teams.get(r["shooter_id"]) == team
            ]
            if team_attempts:
                out[team][season].append({
                    "game_id": gid, "date": date, "opponent": opponent, "outcome": outcome,
                    "attempts": sorted(team_attempts, key=lambda x: x["round"] or 99),
                })

    final_out = {}
    for team, seasons in out.items():
        final_out[team] = {}
        for season, games in seasons.items():
            final_out[team][season] = sorted(games, key=lambda g: g["game_id"])
    return final_out


# ==========================================================================
"""
Faceoff win-direction derivation from PSF playsequence data.

Reproduces Sportlogiq's documented logic, verified draw-for-draw against
Erik's validated build (reason + sector + distance + recovery coords).

KEY FACTS ESTABLISHED FROM THE DATA (not assumed):

  * A faceoff produces THREE rows: one `shorthand='Face-Off'` summary row
    (type hoz/hdz/nz, null zone/outcome/manpower) plus two consecutive
    team-perspective rows sharing a gameTime. Filter on `outcome IS NOT NULL`
    to keep the team rows. 5,477 summary + 10,954 team rows in ANA's 94 games.

  * The recovery is TAGGED, not inferred: it is the first `lpr` with
    `type='faceoff'` by the winning team. There is no search window and no
    time threshold to tune.

  * `flags` carries the OPPOSING centre's hand on every determined row
    (`lefthandedopponent` / `righthandedopponent`). Present on 10,890 of
    10,954 rows; the 64 without it are exactly the `outcome='undetermined'`
    rows. A player's own hand therefore comes from the PAIRED row.

  * `westfodot`/`eastfodot` are in the PHYSICAL rink frame, not the adjusted
    one — the same flag appears at yAdj +22.38 and yAdj -21.38. Dot identity
    MUST come from adjusted coordinates. Using the flags mirrors the DZ map.
"""

import math

# --- Dot keys ---------------------------------------------------------------
# Measured from adjusted coordinates and cross-checked against Erik's pixel
# layout. NOTE: DZ is inverted relative to NZ/OZ in Erik's naming — dz-east is
# the TOP dot (negative yAdj) while oz-east is the BOTTOM one. This is not a
# typo; deriving these by the "obvious" rule ships flipped DZ maps.
#           yAdj < 0 (top of rink as drawn)   yAdj > 0 (bottom)
DOT_KEYS = {
    ("dz", "top"): "dz-east",      ("dz", "bot"): "dz-west",
    ("nzd", "top"): "nz-def-west", ("nzd", "bot"): "nz-def-east",
    ("nzo", "top"): "nz-off-west", ("nzo", "bot"): "nz-off-east",
    ("oz", "top"): "oz-west",      ("oz", "bot"): "oz-east",
}
NZ_CENTRE_MAX_ABS_X = 10.0   # centre dot clusters at |x| < 1; wings sit at |x| ~ 20

SECTOR_NAMES = ["Fwd", "Fwd-Bot", "Bot", "Back-Bot", "Back", "Back-Top", "Top", "Fwd-Top"]
FORWARD_SECTORS = {0, 1, 7}


def dot_key(zone, x_adj, y_adj):
    """Dot identity in the acting player's own adjusted frame."""
    if zone is None or x_adj is None or y_adj is None:
        return None
    if zone == "nz" and abs(x_adj) < NZ_CENTRE_MAX_ABS_X:
        return "nz-center"
    if zone == "nz":
        group = "nzo" if x_adj > 0 else "nzd"
    elif zone in ("dz", "oz"):
        group = zone
    else:
        return None
    return DOT_KEYS.get((group, "bot" if y_adj > 0 else "top"))


def sector_of(dx, dy):
    """45-degree bin, measured off +x (toward the attacking net) rotating
    toward +y (bottom of the rink as drawn). 0=Fwd .. 7=Fwd-Top."""
    if abs(dx) < 1e-9 and abs(dy) < 1e-9:
        return None                      # zero_vector — recovery on the dot
    angle = math.degrees(math.atan2(dy, dx)) % 360
    return int(((angle + 22.5) % 360) // 45)


def hand_from_flags(flags):
    """Hand of the OPPOSING centre, as tagged on this row."""
    if not flags:
        return None
    f = flags.replace(" ", "")
    if "lefthandedopponent" in f:
        return "L"
    if "righthandedopponent" in f:
        return "R"
    return None


# --- Recovery resolution ----------------------------------------------------

RECOVERY_VARIANTS = {
    # name        require type=='faceoff' | require winner's team | require successful
    "tagged":   (True,  True,  True),
    "anytype":  (False, True,  True),
    "anyout":   (False, True,  False),
    "anyteam":  (False, False, True),
}


SELFPASS_VARIANTS = ("adjacent", "own_next", "scan_pass", "scan_any",
                     "poss", "poss_any")

STOPPAGE = ("faceoff", "whistle", "penalty")


def _find_self_pass(events, lpr_i, winner_pid, winner_team, mode):
    """Index of the winner's pass following his own faceoff recovery, or None.

    The reference build resolves 368 more draws than an adjacency rule does:
    opponent events (pressure, block, check) interleave between the lpr and
    the pass, so "immediately next" is too strict.
    """
    n = len(events)
    if mode == "adjacent":
        j = lpr_i + 1
        if j < n and events[j]["name"] == "pass" and events[j]["playerReferenceId"] == winner_pid:
            return j
        return None

    base_poss = events[lpr_i].get("currentPossession")
    for j in range(lpr_i + 1, n):
        e = events[j]
        if e["name"] in STOPPAGE:
            return None
        # A pass in a different possession belongs to a different play. Without
        # this bound the scan runs off the end of the sequence and picks up
        # recoveries 150+ feet away — longer than the rink.
        if mode in ("poss", "poss_any") and base_poss is not None:
            ep = e.get("currentPossession")
            if ep is not None and ep != base_poss:
                return None
        if mode == "own_next" and e["playerReferenceId"] == winner_pid:
            # the winner's very next action must itself be the pass
            return j if e["name"] == "pass" else None
        if mode == "scan_pass" and e["name"] == "lpr" and e["outcome"] == "successful" \
                and e["team"] != winner_team:
            return None                      # opponent took it back first
        if e["name"] == "pass" and e["playerReferenceId"] == winner_pid:
            return j
    return None


def _find_reception(events, pass_i, winner_team, bound_poss=None):
    for j in range(pass_i + 1, len(events)):
        e = events[j]
        if e["name"] in STOPPAGE:
            return None
        if bound_poss is not None:
            ep = e.get("currentPossession")
            if ep is not None and ep != bound_poss:
                return None
        if e["name"] == "reception" and e["team"] == winner_team:
            return j
    return None


def resolve_recovery(events, start_idx, winner_pid, winner_team,
                     variant="anytype", selfpass="poss"):
    """Walk forward for the faceoff recovery.

      teammate_lpr        recovery by anyone but the winner -> the lpr
      self_pass_reception winner recovers AND then passes -> the reception
      self_no_pass_lpr    winner recovers, no pass -> his own lpr

    `variant` selects the recovery lpr (see RECOVERY_VARIANTS); `selfpass`
    selects how the follow-up pass is found (see SELFPASS_VARIANTS).
    """
    need_tag, need_team, need_ok = RECOVERY_VARIANTS[variant]
    lpr_i = None
    for i in range(start_idx, len(events)):
        e = events[i]
        if e["name"] in STOPPAGE:
            break
        if e["name"] != "lpr":
            continue
        if need_tag and e["type"] != "faceoff":
            continue
        if need_team and e["team"] != winner_team:
            continue
        if need_ok and e["outcome"] != "successful":
            continue
        lpr_i = i
        break
    if lpr_i is None:
        return None

    lpr = events[lpr_i]
    if lpr["playerReferenceId"] != winner_pid:
        return "teammate_lpr", lpr["xAdjCoord"], lpr["yAdjCoord"]

    pass_i = _find_self_pass(events, lpr_i, winner_pid, winner_team, selfpass)
    if pass_i is not None:
        bound = lpr.get("currentPossession") if selfpass == "poss" else None
        rcp_i = _find_reception(events, pass_i, winner_team, bound)
        if rcp_i is not None:
            return ("self_pass_reception",
                    events[rcp_i]["xAdjCoord"], events[rcp_i]["yAdjCoord"])
    return "self_no_pass_lpr", lpr["xAdjCoord"], lpr["yAdjCoord"]


# --- Per-game extraction ----------------------------------------------------

def pair_faceoffs(events):
    """Yield (winner_idx, loser_idx_or_None) for each faceoff.

    Grouped by contiguous runs of faceoff rows rather than by exact gameTime.
    Two real cases in ANA's 82 games defeat an equality-on-gameTime rule:

      * the two team rows can carry gameTimes 0.167s apart (2025-12-15,
        2630.0333 vs 2630.2), so they never group together; and
      * a faceoff can have only ONE team-perspective row, with no opposing
        row at all (2025-12-03, 3334.43) — still a won draw.

    A run is bounded by any non-faceoff event, so back-to-back faceoffs stay
    separate. Undetermined-only runs yield nothing: nobody won the draw.
    """
    i, n = 0, len(events)
    while i < n:
        if events[i]["name"] != "faceoff":
            i += 1
            continue
        j = i
        while j < n and events[j]["name"] == "faceoff":
            j += 1
        run = [k for k in range(i, j) if events[k]["outcome"] is not None]
        wins = [k for k in run if events[k]["outcome"] == "successful"]
        others = [k for k in run if events[k]["outcome"] != "successful"]
        if len(wins) == 1:
            yield wins[0], (others[0] if others else None)
        i = j


def derive_game(events, max_period=4, variant="anytype",
                selfpass="poss", anomalies=None):
    """events: list of dict-like plays for one game, ordered by id.
    Returns a list of draw records — two per resolved faceoff (win + loss).
    Period 4 (OT) is included, matching the reference build; period 5 has no
    faceoffs but is excluded defensively.
    """
    out = []
    if anomalies is not None:
        groups = {}
        for i, e in enumerate(events):
            if e["name"] == "faceoff" and e["outcome"] is not None:
                groups.setdefault(e["gameTime"], []).append(i)
        for t, idxs in groups.items():
            outs = sorted(events[i]["outcome"] for i in idxs)
            if not (len(idxs) == 2 and outs.count("successful") == 1):
                anomalies.append({"gameTime": t, "rows": len(idxs), "outcomes": outs})
    for wi, li in pair_faceoffs(events):
        w = events[wi]
        l = events[li] if li is not None else None
        if w["period"] is None or w["period"] > max_period:
            continue

        rec = resolve_recovery(events, max(wi, li if li is not None else wi) + 1,
                               w["playerReferenceId"], w["team"], variant, selfpass)
        sector = is_fwd = reason = rx = ry = dist = None
        if rec:
            reason, rx, ry = rec
            if rx is not None and ry is not None:
                dx, dy = rx - w["xAdjCoord"], ry - w["yAdjCoord"]
                sector = sector_of(dx, dy)
                dist = round(math.hypot(dx, dy), 1)
                if sector is None:
                    reason = "zero_vector"
                else:
                    is_fwd = sector in FORWARD_SECTORS

        pairs = [(w, 1)] + ([(l, 0)] if l is not None else [])
        for row, won in pairs:
            other = l if won else w
            out.append({
                "faceoffId": row["id"],
                "period": row["period"],
                "gameTime": row["gameTime"],
                "team": row["team"],
                "playerReferenceId": row["playerReferenceId"],
                "oppPlayerReferenceId": other["playerReferenceId"] if other else None,
                # this row's flag names the OPPONENT's hand
                "oppHand": hand_from_flags(row["flags"]),
                "dotKey": dot_key(row["zone"], row["xAdjCoord"], row["yAdjCoord"]),
                "zone": row["zone"],
                "manpower": row["manpowerSituation"],
                "situation": "ES" if row["manpowerSituation"] == "evenStrength" else "ST",
                "won": won,
                "sector": sector if won else None,
                "isForward": (1 if is_fwd else 0) if (won and is_fwd is not None) else None,
                "resolution": reason if won else None,
                "recoveryX": rx if won else None,
                "recoveryY": ry if won else None,
                "recoveryDist": dist if won else None,
            })
    return out


def resolve_hands(draws):
    """Majority vote. A player's own hand is the flag carried on his
    OPPONENT's row, so we credit the opponent's oppHand to this player."""
    votes = {}
    by_faceoff = {}
    for d in draws:
        by_faceoff.setdefault((d["gameTime"], d["team"]), d)
    for d in draws:
        pid, opp_hand = d["oppPlayerReferenceId"], d["oppHand"]
        if pid is not None and opp_hand:
            v = votes.setdefault(pid, {"L": 0, "R": 0})
            v[opp_hand] += 1
    return {pid: ("L" if v["L"] >= v["R"] else "R") for pid, v in votes.items()}, votes


# ============================================================================
# FACEOFF WIN-DIRECTION — schema, backfill, validation
#
# Paste at the bottom of api.py, below the derivation helpers.
# All routes are sync `def` so FastAPI runs them in the threadpool; none of
# this sqlite work ever touches the event loop.
# ============================================================================

FO_SCHEMA = """
CREATE TABLE IF NOT EXISTS faceoff_draws (
    gameReferenceId      INTEGER NOT NULL,
    faceoffId            INTEGER NOT NULL,
    gameDate             TEXT,
    season               TEXT,
    isPlayoff            INTEGER NOT NULL DEFAULT 0,
    period               INTEGER,
    gameTime             REAL,
    teamAbbrev           TEXT,
    playerReferenceId    INTEGER NOT NULL,
    oppPlayerReferenceId INTEGER,
    oppHand              TEXT,
    dotKey               TEXT,
    zone                 TEXT,
    manpower             TEXT,
    situation            TEXT,
    won                  INTEGER NOT NULL,
    sector               INTEGER,
    isForward            INTEGER,
    resolution           TEXT,
    recoveryX            REAL,
    recoveryY            REAL,
    recoveryDist         REAL,
    PRIMARY KEY (gameReferenceId, faceoffId)
);
CREATE INDEX IF NOT EXISTS idx_fo_player  ON faceoff_draws (playerReferenceId, situation);
CREATE INDEX IF NOT EXISTS idx_fo_team    ON faceoff_draws (teamAbbrev, season, isPlayoff);
CREATE INDEX IF NOT EXISTS idx_fo_game    ON faceoff_draws (gameReferenceId);
CREATE INDEX IF NOT EXISTS idx_fo_opp     ON faceoff_draws (playerReferenceId, oppPlayerReferenceId);

CREATE TABLE IF NOT EXISTS faceoff_hands (
    playerReferenceId INTEGER PRIMARY KEY,
    hand              TEXT,
    votesL            INTEGER NOT NULL DEFAULT 0,
    votesR            INTEGER NOT NULL DEFAULT 0,
    source            TEXT
);
"""


@app.on_event("startup")
def ensure_faceoff_schema():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(FO_SCHEMA)
    conn.commit()
    conn.close()


# --- regular season vs playoffs ---------------------------------------------

def _fo_rs_game_count(season: str) -> int:
    """Regular-season games per team. The NHL moves from 82 to 84 beginning
    with 2026-27 under the new CBA, so this must not be a constant: at 82 the
    classifier would tag every team's 83rd and 84th game as a playoff game,
    which would switch the Playoffs selector on in early April, mid-season."""
    try:
        start_year = int(str(season)[:4])
    except (TypeError, ValueError):
        return 84
    return 82 if start_year <= 2025 else 84


def _fo_playoff_game_ids(conn, season: str) -> set:
    """A game is regular season only if it falls within the first N for BOTH
    teams, N being that season's schedule length. Self-correcting: needs no
    hardcoded cutoff date, and flags nothing mid-season because no team
    exceeds N until the playoffs start.

    (The front end's RS_END/PO_START constants are hardcoded to 2025-26 and
    will be wrong for 2026-27 — this deliberately does not depend on them.)
    """
    rows = conn.execute(
        """SELECT gameReferenceId, gameDate, awayTeamAbbrev, homeTeamAbbrev
           FROM games WHERE season = ? ORDER BY gameDate, gameReferenceId""",
        (season,),
    ).fetchall()
    rs_games = _fo_rs_game_count(season)
    counts, playoffs = {}, set()
    for r in rows:
        for ab in (r["awayTeamAbbrev"], r["homeTeamAbbrev"]):
            counts[ab] = counts.get(ab, 0) + 1
            if counts[ab] > rs_games:
                playoffs.add(r["gameReferenceId"])
    return playoffs


def _fo_rows_for_game(conn, game_id: int) -> list:
    return [dict(r) for r in conn.execute(
        """SELECT id, name, type, outcome, zone, team, gameTime, period,
                  xAdjCoord, yAdjCoord, playerReferenceId, manpowerSituation,
                  flags, currentPossession
           FROM plays WHERE gameReferenceId = ? ORDER BY id""",
        (game_id,),
    ).fetchall()]


# Reverse of TEAM_NAMES — the `plays.team` column stores full names while
# every endpoint addresses teams by abbreviation.
ABBR_BY_FULL = {v: k for k, v in TEAM_NAMES.items()}


def fo_build_game(conn, game_id: int, playoff_ids: set) -> int:
    """Derive and store every draw for one game. Idempotent per game."""
    g = conn.execute(
        "SELECT gameDate, season FROM games WHERE gameReferenceId = ?", (game_id,)
    ).fetchone()
    if not g:
        return 0
    events = _fo_rows_for_game(conn, game_id)
    draws = derive_game(events)
    conn.execute("DELETE FROM faceoff_draws WHERE gameReferenceId = ?", (game_id,))
    payload = [(
        game_id, d["faceoffId"], g["gameDate"], g["season"],
        1 if game_id in playoff_ids else 0, d["period"], d["gameTime"],
        ABBR_BY_FULL.get(d["team"], d["team"]), d["playerReferenceId"],
        d["oppPlayerReferenceId"], d["oppHand"], d["dotKey"], d["zone"],
        d["manpower"], d["situation"], d["won"], d["sector"], d["isForward"],
        d["resolution"], d["recoveryX"], d["recoveryY"], d["recoveryDist"],
    ) for d in draws]
    conn.executemany(
        """INSERT OR REPLACE INTO faceoff_draws VALUES
           (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", payload)
    return len(payload)


def fo_rebuild_hands(conn):
    """League-wide majority vote. A player's own hand is the flag carried on
    his OPPONENT's row, so credit each row's oppHand to its oppPlayer."""
    votes = {}
    for r in conn.execute(
        """SELECT oppPlayerReferenceId pid, oppHand h, COUNT(*) n
           FROM faceoff_draws
           WHERE oppPlayerReferenceId IS NOT NULL AND oppHand IS NOT NULL
           GROUP BY pid, h"""
    ).fetchall():
        v = votes.setdefault(r["pid"], {"L": 0, "R": 0})
        v[r["h"]] += r["n"]
    conn.execute("DELETE FROM faceoff_hands")
    conn.executemany(
        "INSERT INTO faceoff_hands VALUES (?,?,?,?,?)",
        [(pid, "L" if v["L"] >= v["R"] else "R", v["L"], v["R"], "psf")
         for pid, v in votes.items()],
    )
    return len(votes)


@app.post("/api/faceoffs/backfill")
def faceoffs_backfill(
    season: str | None = Query(None, description="Limit to one season"),
    team: str | None = Query(None, description="Limit to one team's games"),
    max_games: int = Query(150, ge=1, le=2000,
                           description="Games per call — keeps the request inside Fly's proxy timeout"),
    skip_existing: bool = Query(True, description="Skip games already derived"),
    rebuild_hands: bool = Query(True),
    admin: CurrentUser = Depends(require_admin),
):
    """Derives faceoff draws in resumable batches.

    The league is ~1,400 games; doing them in one request outruns the proxy
    timeout, so this processes `max_games` at a time and skips what it has
    already done. Call it repeatedly until `remaining` comes back 0.
    """
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        q = "SELECT g.gameReferenceId, g.season FROM games g WHERE 1=1"
        args = []
        if season:
            q += " AND g.season = ?"
            args.append(season)
        if team:
            q += " AND (g.awayTeamAbbrev = ? OR g.homeTeamAbbrev = ?)"
            args += [team, team]
        if skip_existing:
            q += (" AND NOT EXISTS (SELECT 1 FROM faceoff_draws d"
                  "                WHERE d.gameReferenceId = g.gameReferenceId)")
        pending = conn.execute(q + " ORDER BY g.gameDate", args).fetchall()
        batch = pending[:max_games]

        po_by_season, total = {}, 0
        for row in batch:
            s = row["season"]
            if s not in po_by_season:
                po_by_season[s] = _fo_playoff_game_ids(conn, s)
            total += fo_build_game(conn, row["gameReferenceId"], po_by_season[s])
        conn.commit()

        players = None
        if rebuild_hands:
            players = fo_rebuild_hands(conn)
            conn.commit()
        return {
            "gamesProcessed": len(batch),
            "remaining": len(pending) - len(batch),
            "draws": total,
            "playersWithHand": players,
            "playoffGames": {s: len(v) for s, v in po_by_season.items()},
            "note": ("call again until remaining is 0"
                     if len(pending) > len(batch) else "complete"),
        }
    finally:
        conn.close()


# --- validation against the reference build ---------------------------------

@app.post("/api/faceoffs/validate")
def faceoffs_validate(
    file: UploadFile | None = File(None, description="faceoff_answer_key_summary.json"),
    team: str = Query("ANA"),
    season: str = Query("20252026"),
    limit: int = Query(25, description="Max mismatches to list"),
    admin: CurrentUser = Depends(require_admin),
):
    """Diffs derived output against the reference build's answer key.

    The key is uploaded rather than read from disk: it is a one-off
    validation artifact, so there is no reason to carry 460KB of it in the
    production image (and no reason for this to depend on how the Dockerfile
    happens to copy files). Falls back to a copy beside api.py if present.

    Regular season only — the reference build was 82 games ending 2026-04-16
    (2025-26; from 2026-27 the season is 84).
    """
    if file is not None:
        try:
            key = json.loads(file.file.read())
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            raise HTTPException(400, f"uploaded file is not valid JSON: {e}")
    else:
        key_path = Path(__file__).parent / "faceoff_answer_key_summary.json"
        if not key_path.exists():
            raise HTTPException(400, "no answer key uploaded and none found beside api.py")
        key = json.loads(key_path.read_text())
    if not isinstance(key, dict) or "players" not in key:
        raise HTTPException(400, "answer key missing a 'players' object")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        gids = [r[0] for r in conn.execute(
            """SELECT gameReferenceId FROM games
               WHERE season = ? AND (awayTeamAbbrev = ? OR homeTeamAbbrev = ?)
               ORDER BY gameDate""", (season, team, team)).fetchall()]
        po = _fo_playoff_game_ids(conn, season)
        gids = [g for g in gids if g not in po]
        if not gids:
            raise HTTPException(404, f"no regular-season {team} games for {season}")
        ph = ",".join("?" for _ in gids)

        mine = {}
        for r in conn.execute(
            f"""SELECT playerReferenceId p, situation s, dotKey d,
                       SUM(won) wins, SUM(1-won) losses, COUNT(*) total
                FROM faceoff_draws WHERE gameReferenceId IN ({ph})
                  AND dotKey IS NOT NULL
                GROUP BY p, s, d""", gids).fetchall():
            mine[(str(r["p"]), r["s"], r["d"])] = (r["wins"], r["losses"], r["total"])

        sectors = {}
        for r in conn.execute(
            f"""SELECT playerReferenceId p, situation s, dotKey d, sector sec,
                       COUNT(*) n
                FROM faceoff_draws WHERE gameReferenceId IN ({ph})
                  AND won = 1 AND sector IS NOT NULL AND dotKey IS NOT NULL
                GROUP BY p, s, d, sec""", gids).fetchall():
            sectors.setdefault((str(r["p"]), r["s"], r["d"]), {})[r["sec"]] = r["n"]

        theirs, mismatches = {}, []
        for pid, pl in key["players"].items():
            for sit in ("ES", "ST"):
                for dk, dv in pl[sit]["dots"].items():
                    theirs[(pid, sit, dk)] = (dv["wins"], dv["losses"], dv["total"])

        checked = agreed = 0
        for k in set(mine) | set(theirs):
            checked += 1
            m, t = mine.get(k), theirs.get(k)
            if m == t:
                agreed += 1
            elif len(mismatches) < limit:
                mismatches.append({
                    "player": key["players"].get(k[0], {}).get("name", k[0]),
                    "pid": k[0], "situation": k[1], "dot": k[2],
                    "mine": m, "reference": t,
                })

        sec_checked = sec_agreed = sec_bad = 0
        for k, ref in theirs.items():
            pid, sit, dk = k
            rd = key["players"].get(pid, {}).get(sit, {}).get("dots", {}).get(dk)
            if not rd:
                continue
            want = {x["sector"]: x["count"] for x in rd["dirs"]}
            got = sectors.get(k, {})
            sec_checked += 1
            if want == got:
                sec_agreed += 1
            elif sec_bad < limit:
                sec_bad += 1
                mismatches.append({"player": key["players"][pid]["name"], "pid": pid,
                                   "situation": sit, "dot": dk, "kind": "sectors",
                                   "mine": got, "reference": want})

        hands = conn.execute(
            "SELECT playerReferenceId p, hand h FROM faceoff_hands").fetchall()
        hmap = {str(r["p"]): r["h"] for r in hands}
        h_ok = h_bad = 0
        hand_diffs = []
        for pid, pl in key["players"].items():
            if pid in hmap and pl.get("hand"):
                if hmap[pid] == pl["hand"]:
                    h_ok += 1
                else:
                    h_bad += 1
                    if len(hand_diffs) < limit:
                        hand_diffs.append({"pid": pid, "name": pl["name"],
                                           "mine": hmap[pid], "reference": pl["hand"]})

        return {
            "games": len(gids),
            "dotTotals": {"checked": checked, "agreed": agreed,
                          "pct": round(100 * agreed / checked, 2) if checked else None},
            "sectorCounts": {"checked": sec_checked, "agreed": sec_agreed,
                             "pct": round(100 * sec_agreed / sec_checked, 2) if sec_checked else None},
            "handedness": {"agreed": h_ok, "disagreed": h_bad, "diffs": hand_diffs},
            "mismatches": mismatches,
        }
    finally:
        conn.close()


@app.post("/api/faceoffs/calibrate")
def faceoffs_calibrate(
    file: UploadFile = File(..., description="faceoff_answer_key_summary.json"),
    team: str = Query("ANA"),
    season: str = Query("20252026"),
    admin: CurrentUser = Depends(require_admin),
):
    """Measures every candidate recovery rule against the reference build in
    one pass, so the rule is chosen from evidence rather than picked and then
    checked. Derives in memory only — does not touch faceoff_draws.

    Reports, per variant: share of wins with a resolved direction, and how
    many (player, situation, dot) sector histograms match the key exactly.
    """
    key = json.loads(file.file.read())
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        gids = [r[0] for r in conn.execute(
            """SELECT gameReferenceId FROM games
               WHERE season = ? AND (awayTeamAbbrev = ? OR homeTeamAbbrev = ?)
               ORDER BY gameDate""", (season, team, team)).fetchall()]
        gids = [g for g in gids if g not in _fo_playoff_game_ids(conn, season)]

        want = {}
        for pid, pl in key["players"].items():
            for sit in ("ES", "ST"):
                for dk, dv in pl[sit]["dots"].items():
                    want[(pid, sit, dk)] = {x["sector"]: x["count"] for x in dv["dirs"]}

        games = [_fo_rows_for_game(conn, g) for g in gids]
        results = {}
        combos = ([(v, "scan_pass") for v in RECOVERY_VARIANTS]
                  + [("anytype", sp) for sp in SELFPASS_VARIANTS])
        for variant, selfpass in combos:
            got, wins, resolved = {}, 0, 0
            for events in games:
                for d in derive_game(events, variant=variant, selfpass=selfpass):
                    if not d["won"]:
                        continue
                    wins += 1
                    if d["sector"] is None or d["dotKey"] is None:
                        continue
                    resolved += 1
                    k = (str(d["playerReferenceId"]), d["situation"], d["dotKey"])
                    got.setdefault(k, {})[d["sector"]] = got.setdefault(k, {}).get(d["sector"], 0) + 1
            exact = sum(1 for k, v in want.items() if got.get(k, {}) == v)
            results[f"{variant}/{selfpass}"] = {
                "wins": wins,
                "resolvedPct": round(100 * resolved / wins, 2) if wins else None,
                "dotsExact": exact,
                "dotsChecked": len(want),
                "dotsExactPct": round(100 * exact / len(want), 2) if want else None,
            }
        best = max(results, key=lambda v: results[v]["dotsExact"])
        return {"games": len(gids), "referenceResolvedPct": key["meta"].get("resolvedPct"),
                "variants": results, "best": best}
    finally:
        conn.close()


@app.post("/api/faceoffs/diff-draws")
def faceoffs_diff_draws(
    file: UploadFile = File(..., description="faceoff_answer_key_draws.jsonl"),
    team: str = Query("ANA"),
    season: str = Query("20252026"),
    variant: str = Query("anytype"),
    selfpass: str = Query("poss"),
    samples: int = Query(8, ge=0, le=40),
    admin: CurrentUser = Depends(require_admin),
):
    """Per-draw diff against the reference build. Reports a confusion matrix
    of resolution reasons and a sample of disagreeing draws, so the residual
    rule difference is identified rather than guessed at."""
    ref = {}
    for line in file.file.read().decode("utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        if r.get("out") != "win":
            continue
        ref[(str(r["pid"]), r["date"], round(float(r["t"]), 2))] = r

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """SELECT gameReferenceId, gameDate FROM games
               WHERE season = ? AND (awayTeamAbbrev = ? OR homeTeamAbbrev = ?)
               ORDER BY gameDate""", (season, team, team)).fetchall()
        po = _fo_playoff_game_ids(conn, season)
        rows = [r for r in rows if r["gameReferenceId"] not in po]

        confusion, sector_delta, samp = {}, {}, []
        anomalies, mine_keys = [], set()
        for g in rows:
            events = _fo_rows_for_game(conn, g["gameReferenceId"])
            for d in derive_game(events, variant=variant, selfpass=selfpass, anomalies=anomalies):
                if not d["won"]:
                    continue
                k = (str(d["playerReferenceId"]), g["gameDate"], round(d["gameTime"], 2))
                mine_keys.add(k)
                r = ref.get(k)
                if r is None:
                    confusion[("(no reference row)", str(d["resolution"]))] = \
                        confusion.get(("(no reference row)", str(d["resolution"])), 0) + 1
                    continue
                pair = (str(r["rsn"]), str(d["resolution"]))
                confusion[pair] = confusion.get(pair, 0) + 1
                if r["sec"] != d["sector"]:
                    key = f'{r["sec"]}->{d["sector"]}'
                    sector_delta[key] = sector_delta.get(key, 0) + 1
                    if len(samp) < samples:
                        samp.append({
                            "pid": k[0], "date": k[1], "gameTime": k[2], "dot": d["dotKey"],
                            "refReason": r["rsn"], "mineReason": d["resolution"],
                            "refSector": r["sec"], "mineSector": d["sector"],
                            "refXY": [r["rx"], r["ry"]],
                            "mineXY": [d["recoveryX"], d["recoveryY"]],
                            "refDist": r["rd"], "mineDist": d["recoveryDist"],
                        })
        missing = []
        for k in list(set(ref) - mine_keys)[:samples]:
            pid, date, t_ = k
            row = conn.execute(
                """SELECT gameReferenceId g FROM games WHERE gameDate = ?
                   AND (awayTeamAbbrev = ? OR homeTeamAbbrev = ?)""",
                (date, team, team)).fetchone()
            rows = []
            if row:
                rows = [dict(r) for r in conn.execute(
                    """SELECT id, name, outcome, zone, team, gameTime,
                              playerReferenceId, flags
                       FROM plays WHERE gameReferenceId = ? AND name = 'faceoff'
                         AND gameTime BETWEEN ? AND ? ORDER BY id""",
                    (row["g"], t_ - 0.5, t_ + 0.5)).fetchall()]
            missing.append({"pid": pid, "date": date, "gameTime": t_,
                            "faceoffRowsThere": rows})
        return {
            "games": len(rows),
            "referenceWins": len(ref), "myWins": len(mine_keys),
            "missingFromMine": len(set(ref) - mine_keys), "missingSample": missing,
            "extraInMine": len(mine_keys - set(ref)),
            "reasonConfusion": {f"ref={k[0]} | mine={k[1]}": v
                                for k, v in sorted(confusion.items(), key=lambda x: -x[1])},
            "sectorShifts": dict(sorted(sector_delta.items(), key=lambda x: -x[1])[:15]),
            "pairingAnomalies": anomalies[:samples],
            "pairingAnomalyCount": len(anomalies),
            "sampleMismatches": samp,
        }
    finally:
        conn.close()


# ============================================================================
# Payload for the Faceoffs tab — emits Erik's {meta, players} shape so his
# front end can consume it directly via its ?data= loader.
# ============================================================================

_FO_SEGMENTS = {"rs": " AND isPlayoff = 0", "po": " AND isPlayoff = 1", "both": ""}

# Strength views offered by the tab. Each draw contributes to its own bucket
# and to ALL, so PP and PK are full dot maps rather than bare win/loss counts.
FO_SITUATIONS = ("ES", "PP", "PK", "ALL")
MANPOWER_SIT = {"evenStrength": "ES", "powerPlay": "PP", "shortHanded": "PK"}


@app.get("/api/faceoffs/matchup/{us}/{them}")
def faceoffs_matchup(
    us: str, them: str,
    season: str = Query("20252026"),
    segment: str = Query("rs", description="rs | po | both"),
    user: CurrentUser = Depends(require_login),
):
    if segment not in _FO_SEGMENTS:
        raise HTTPException(400, "segment must be rs, po or both")
    us, them = us.upper(), them.upper()
    seg_sql = _FO_SEGMENTS[segment]

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        # Gate the RS/PO/Both selector on playoff games actually existing for
        # this season, rather than on a date: before the playoffs begin there
        # is nothing to show, and a team that missed them has nothing either.
        # BOTH teams must have playoff draws — this is a matchup view, and a
        # Playoffs button that yields an empty grid because one side missed
        # the playoffs is worse than no button. (EDM played 6 in 2025-26; VAN
        # played none, which is exactly the case that errored.)
        po_teams = {r[0] for r in conn.execute(
            "SELECT DISTINCT teamAbbrev FROM faceoff_draws "
            "WHERE season = ? AND isPlayoff = 1 AND teamAbbrev IN (?,?)",
            (season, us, them)).fetchall()}
        has_po = len(po_teams) == 2

        rows = conn.execute(
            f"""SELECT playerReferenceId pid, teamAbbrev team, situation sit, dotKey dot,
                       manpower, oppHand, oppPlayerReferenceId opp, won, sector, isForward
                FROM faceoff_draws
                WHERE season = ? AND teamAbbrev IN (?, ?) AND dotKey IS NOT NULL{seg_sql}""",
            (season, us, them)).fetchall()

        players: dict = {}
        for r in rows:
            p = players.setdefault(str(r["pid"]), {
                "team": r["team"],
                "dots": {s: {} for s in FO_SITUATIONS},
                "dotsByHand": {s: {"L": {}, "R": {}} for s in FO_SITUATIONS},
                "vs": {s: {"L": [0, 0], "R": [0, 0]} for s in FO_SITUATIONS},
                "h2h": {s: {} for s in FO_SITUATIONS},
                "rec": {"PP": [0, 0], "SH": [0, 0], "EV": [0, 0]},
            })
            # Every draw lands in its own strength bucket AND in ALL, so the
            # four buttons are real views rather than ES plus a lump.
            sits = [MANPOWER_SIT.get(r["manpower"]), "ALL"]
            for sit in sits:
                if not sit:
                    continue
                d = p["dots"][sit].setdefault(r["dot"], {"wins": 0, "losses": 0, "sectors": {}})
                if r["won"]:
                    d["wins"] += 1
                    if r["sector"] is not None:
                        d["sectors"][r["sector"]] = d["sectors"].get(r["sector"], 0) + 1
                else:
                    d["losses"] += 1
                if r["oppHand"] in ("L", "R"):
                    v = p["vs"][sit][r["oppHand"]]
                    v[0] += r["won"]; v[1] += 1
                    # Mirrors the "dots" accumulation above, scoped to draws
                    # against that opponent hand only — feeds the Player Map's
                    # opponent-handedness toggle without changing "dots" itself.
                    hd = p["dotsByHand"][sit][r["oppHand"]].setdefault(
                        r["dot"], {"wins": 0, "losses": 0, "sectors": {}})
                    if r["won"]:
                        hd["wins"] += 1
                        if r["sector"] is not None:
                            hd["sectors"][r["sector"]] = hd["sectors"].get(r["sector"], 0) + 1
                    else:
                        hd["losses"] += 1
                if r["opp"] is not None:
                    h = p["h2h"][sit].setdefault(str(r["opp"]), [0, 0])
                    h[0] += r["won"]; h[1] += 1
            bucket = {"powerPlay": "PP", "shortHanded": "SH", "evenStrength": "EV"}.get(r["manpower"])
            if bucket:
                p["rec"][bucket][0] += r["won"]; p["rec"][bucket][1] += 1

        hands = {str(r["playerReferenceId"]): r["hand"]
                 for r in conn.execute("SELECT playerReferenceId, hand FROM faceoff_hands")}
        names = {str(r["playerReferenceId"]): r for r in conn.execute(
            "SELECT playerReferenceId, firstName, lastName, position, jersey FROM players")}

        def _reshape_dots(raw_dots: dict) -> dict:
            dots = {}
            for dk, dv in raw_dots.items():
                w = dv["wins"]
                dirs = [{"sector": s, "count": c,
                         "shareOfWins": round(c / w, 5) if w else 0,
                         "forward": s in FORWARD_SECTORS}
                        for s, c in sorted(dv["sectors"].items())]
                dots[dk] = {
                    "wins": w, "losses": dv["losses"], "total": w + dv["losses"],
                    "winPct": round(w / (w + dv["losses"]), 5) if (w + dv["losses"]) else 0,
                    "winsMissingDirection": w - sum(dv["sectors"].values()),
                    "dirs": dirs,
                }
            return dots

        out = {}
        for pid, p in players.items():
            nm = names.get(pid)
            rec = {
                "name": f'{nm["firstName"]} {nm["lastName"]}'.strip() if nm else pid,
                "team": p["team"], "teamName": TEAM_NAMES.get(p["team"], p["team"]),
                "hand": hands.get(pid), "handSource": "psf" if pid in hands else None,
                "number": nm["jersey"] if nm else None,
                "position": nm["position"] if nm else None,
                "rec": p["rec"],
            }
            for sit in FO_SITUATIONS:
                rec[sit] = {
                    "dots": _reshape_dots(p["dots"][sit]),
                    "dotsByHand": {
                        "L": _reshape_dots(p["dotsByHand"][sit]["L"]),
                        "R": _reshape_dots(p["dotsByHand"][sit]["R"]),
                    },
                    "vsL": p["vs"][sit]["L"], "vsR": p["vs"][sit]["R"],
                    "h2h": p["h2h"][sit],
                }
            out[pid] = rec

        # Fields Erik's front end already reads off META, so his render code
        # needs no reshaping: coverage, dataset line, resolution summary.
        agg = conn.execute(
            f"""SELECT COUNT(DISTINCT gameReferenceId) g, MIN(gameDate) f, MAX(gameDate) l,
                       SUM(won) w,
                       SUM(CASE WHEN won = 1 AND sector IS NOT NULL THEN 1 ELSE 0 END) rw
                FROM faceoff_draws
                WHERE season = ? AND teamAbbrev IN (?,?){seg_sql}""",
            (season, us, them)).fetchone()
        res = {r["resolution"]: r["n"] for r in conn.execute(
            f"""SELECT resolution, COUNT(*) n FROM faceoff_draws
                WHERE season = ? AND teamAbbrev IN (?,?) AND won = 1
                  AND resolution IS NOT NULL{seg_sql}
                GROUP BY resolution""", (season, us, them)).fetchall()}

        counts = {}
        for r in conn.execute(
            f"""SELECT teamAbbrev t, COUNT(DISTINCT gameReferenceId) g, COUNT(*) d
                FROM faceoff_draws WHERE season = ? AND teamAbbrev IN (?,?){seg_sql}
                GROUP BY t""", (season, us, them)).fetchall():
            counts[r["t"]] = {"games": r["g"], "draws": r["d"]}

        return {
            "meta": {
                "season": season, "us": us, "them": them, "segment": segment,
                "hasPlayoffs": bool(has_po),
                "teams": {ab: {"name": TEAM_NAMES.get(ab, ab),
                               "fullSeason": True, "rosterFromApi": False,
                               "players": sum(1 for p in out.values() if p["team"] == ab),
                               **counts.get(ab, {"games": 0, "draws": 0})}
                          for ab in (us, them)},
                "players": len(out),
                "games": agg["g"] or 0,
                "firstGame": agg["f"], "lastGame": agg["l"],
                "generated": dt.date.today().isoformat(),
                "apiUsed": False,
                "totalWins": agg["w"] or 0,
                "resolvedPct": (round(100 * agg["rw"] / agg["w"], 1) if agg["w"] else None),
                "resolution": res,
                "validation": {},
            },
            "players": out,
        }
    finally:
        conn.close()
